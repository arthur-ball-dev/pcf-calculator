"""
DEFRA UK Government Conversion Factors connector.

TASK-DATA-P5-003: DEFRA Data Connector
TASK-DATA-P7-008: Fix DEFRA Connector URL (BUG-DATA-002)
TASK-DATA-P8-BUG-003: Fix DEFRA header detection and column mapping

This module implements the DEFRAEmissionFactorsIngestion class that:
- Downloads DEFRA Excel workbook from gov.uk
- Parses emission factors from multiple sheets (Fuels, Electricity, Materials, etc.)
- Transforms data to internal schema with correct scope assignment
- Supports upsert pattern for incremental updates

Features:
- Multi-sheet parsing for 6 categories
- Partial sheet name matching (handles year-specific sheet names)
- Unit extraction from column names (e.g., "kg CO2e per kWh" -> "kWh")
- Scope assignment per category (Scope 1, 2, 3)
- Geography set to GB (Great Britain)
- Data quality rating of 0.88 (high confidence UK government source)

Usage:
    from backend.services.data_ingestion.defra_ingestion import (
        DEFRAEmissionFactorsIngestion
    )

    ingestion = DEFRAEmissionFactorsIngestion(
        db=async_session,
        data_source_id="defra-source-uuid"
    )
    result = await ingestion.execute_sync()
"""

import io
import re
from typing import List, Dict, Any, Optional

import httpx
from openpyxl import load_workbook

from backend.services.data_ingestion.base import BaseDataIngestion


class DEFRAEmissionFactorsIngestion(BaseDataIngestion):
    """
    DEFRA UK Government Conversion Factors connector.

    Downloads and parses the DEFRA greenhouse gas conversion factors
    Excel workbook, extracting emission factors from multiple sheets
    and transforming them to the internal schema.

    Attributes:
        DEFRA_URL: URL to download the DEFRA conversion factors Excel file
        SHEET_CONFIGS: Configuration for each sheet type including scope,
            category, and column mappings
        reference_year: Reference year for the emission factors (default: 2024)
    """

    # DEFRA file URL (update annually when new edition released)
    # TASK-DATA-P7-008: Fixed URL - UK Government migrated to new CDN structure
    # BUG-DATA-002: Old URL returned 404 due to path change from
    # /government/uploads/system/uploads/attachment_data/ to /media/{content-id}/
    DEFRA_URL = (
        "https://assets.publishing.service.gov.uk/media/6722567487df31a87d8c497e/"
        "ghg-conversion-factors-2024-full_set__for_advanced_users__v1_1.xlsx"
    )

    # Sheets to process and their mappings
    # TASK-DATA-P8-BUG-003: Updated column names to match actual DEFRA file
    # The actual DEFRA columns use "kg CO2e" not "kg CO2e per unit"
    SHEET_CONFIGS: Dict[str, Dict[str, Any]] = {
        "Fuels": {
            "scope": "Scope 1",
            "category": "combustion",
            "activity_col": "Fuel",
            # TASK-DATA-P8-BUG-003: Primary and fallback CO2e column names
            "co2e_col": "kg CO2e",  # Actual DEFRA column name
            "co2e_col_fallback": ["kg CO2e per unit", "Total kg CO2e"],
            "unit_col": "Unit",
            # Required columns to identify header row (not just activity col)
            "required_cols": ["Fuel", "kg CO2e"],
        },
        "Electricity": {
            "scope": "Scope 2",
            "category": "electricity",
            "activity_col": "Activity",
            "co2e_col": "kg CO2e",
            "co2e_col_fallback": ["kg CO2e per kWh"],
            "unit_col": None,  # Fixed unit extracted from column name
            "required_cols": ["Activity", "kg CO2e"],
        },
        "Material use": {
            "scope": "Scope 3",
            "category": "materials",
            "activity_col": "Material",
            "co2e_col": "kg CO2e",
            "co2e_col_fallback": ["kg CO2e per kg", "Total kg CO2e"],
            "unit_col": None,
            "required_cols": ["Material", "kg CO2e"],
        },
        "Waste disposal": {
            "scope": "Scope 3",
            "category": "waste",
            "activity_col": "Waste type",
            "co2e_col": "kg CO2e",
            "co2e_col_fallback": ["kg CO2e per tonne", "Total kg CO2e"],
            "unit_col": None,
            "required_cols": ["Waste", "kg CO2e"],
        },
        "Business travel- air": {
            "scope": "Scope 3",
            "category": "transport",
            "activity_col": "Type of flight",
            "co2e_col": "kg CO2e",
            "co2e_col_fallback": ["kg CO2e per passenger km"],
            "unit_col": None,
            "required_cols": ["Type of flight", "kg CO2e"],
        },
        "Freighting goods": {
            "scope": "Scope 3",
            "category": "transport",
            "activity_col": "Vehicle type",
            "co2e_col": "kg CO2e",
            "co2e_col_fallback": ["kg CO2e per tonne.km"],
            "unit_col": None,
            "required_cols": ["Vehicle", "kg CO2e"],
        },
    }

    # Minimum number of required columns that must be present to identify header row
    MIN_REQUIRED_COLS = 2

    def __init__(self, *args: Any, **kwargs: Any) -> None:
        """
        Initialize the DEFRA ingestion connector.

        Sets reference_year to 2024 (current DEFRA edition).
        Update this when new annual edition is released.

        Args:
            *args: Positional arguments passed to BaseDataIngestion
            **kwargs: Keyword arguments passed to BaseDataIngestion
        """
        super().__init__(*args, **kwargs)
        self.reference_year: int = 2024

    async def fetch_raw_data(self) -> bytes:
        """
        Download DEFRA Excel file.

        Uses httpx with redirect following enabled since DEFRA site
        may redirect to CDN or different URL.

        Returns:
            Raw bytes of the Excel file

        Raises:
            httpx.HTTPStatusError: On HTTP error responses
            httpx.ConnectError: On connection failures
        """
        async with httpx.AsyncClient(timeout=120.0) as client:
            response = await client.get(
                self.DEFRA_URL,
                follow_redirects=True
            )
            response.raise_for_status()
            return response.content

    async def parse_data(self, raw_data: bytes) -> List[Dict[str, Any]]:
        """
        Parse DEFRA Excel workbook.

        Iterates through configured sheet types, finds matching sheets
        using partial name matching, and extracts records from each.

        Args:
            raw_data: Raw bytes of the Excel file

        Returns:
            List of parsed records with sheet metadata included
        """
        workbook = load_workbook(io.BytesIO(raw_data), read_only=True)
        records: List[Dict[str, Any]] = []

        for sheet_name, config in self.SHEET_CONFIGS.items():
            # Handle partial sheet name matches
            matching_sheet = self._find_sheet(workbook, sheet_name)
            if not matching_sheet:
                continue

            sheet = workbook[matching_sheet]
            sheet_records = self._parse_sheet(sheet, config, matching_sheet)
            records.extend(sheet_records)

        workbook.close()
        return records

    def _find_sheet(
        self,
        workbook: Any,
        target_name: str
    ) -> Optional[str]:
        """
        Find sheet by partial name match.

        DEFRA sheet names may include year or other suffixes,
        so we use case-insensitive partial matching.

        Args:
            workbook: openpyxl Workbook object
            target_name: Target sheet name to search for

        Returns:
            Actual sheet name if found, None otherwise
        """
        target_lower = target_name.lower()
        for sheet_name in workbook.sheetnames:
            if target_lower in sheet_name.lower():
                return sheet_name
        return None

    def _parse_sheet(
        self,
        sheet: Any,
        config: Dict[str, Any],
        sheet_name: str
    ) -> List[Dict[str, Any]]:
        """
        Parse a single DEFRA sheet.

        Finds the header row by looking for multiple expected column names,
        then extracts all data rows below it.

        TASK-DATA-P8-BUG-003: Improved header detection to require multiple
        columns, not just the activity column (which matches sheet titles).

        Args:
            sheet: openpyxl worksheet object
            config: Sheet configuration with column mappings
            sheet_name: Actual sheet name (for metadata)

        Returns:
            List of parsed records from this sheet
        """
        records: List[Dict[str, Any]] = []
        headers: Optional[List[str]] = None

        for row_idx, row in enumerate(sheet.iter_rows(values_only=True)):
            # Find header row (contains expected column names)
            if headers is None:
                if self._is_header_row(row, config):
                    headers = [str(h).strip() if h else "" for h in row]
                    continue
                continue

            # Skip empty rows
            if not any(row):
                continue

            # Create record from row data
            record = dict(zip(headers, row))
            record["_sheet_name"] = sheet_name
            record["_row_idx"] = row_idx
            record["_config"] = config
            records.append(record)

        return records

    def _is_header_row(self, row: tuple, config: Dict[str, Any]) -> bool:
        """
        Check if row is the header row.

        TASK-DATA-P8-BUG-003: Improved header detection that requires
        multiple expected columns to be present, not just the activity column.
        This prevents matching sheet titles (e.g., "Fuels" at row 1) instead
        of the actual data header row (e.g., row 21).

        The function checks for:
        1. Required columns from config (at least MIN_REQUIRED_COLS must match)
        2. "kg CO2e" column must be present (distinguishes data header from titles)

        Args:
            row: Tuple of cell values
            config: Sheet configuration with required_cols

        Returns:
            True if this appears to be the header row
        """
        if not row or not any(row):
            return False

        # Convert row to lowercase strings for matching
        row_cells = [str(cell).strip().lower() if cell else "" for cell in row]
        row_str = " ".join(row_cells)

        # Get required columns from config, fall back to activity_col
        required_cols = config.get("required_cols", [config["activity_col"]])

        # Count how many required columns are present
        matches = 0
        for col in required_cols:
            col_lower = col.lower()
            # Check if any cell contains this column name
            for cell in row_cells:
                if col_lower in cell:
                    matches += 1
                    break

        # TASK-DATA-P8-BUG-003: Require at least MIN_REQUIRED_COLS matches
        # AND ensure "kg co2e" is present (this is the key differentiator
        # from sheet titles which don't have CO2e columns)
        has_co2e_col = "kg co2e" in row_str

        return matches >= self.MIN_REQUIRED_COLS and has_co2e_col

    async def transform_data(
        self,
        parsed_data: List[Dict[str, Any]]
    ) -> List[Dict[str, Any]]:
        """
        Transform DEFRA data to internal schema.

        Extracts activity name, CO2e factor, unit, and applies
        scope/category from sheet configuration.

        Args:
            parsed_data: List of parsed records with metadata

        Returns:
            List of transformed records matching EmissionFactor schema
        """
        transformed: List[Dict[str, Any]] = []

        for record in parsed_data:
            config = record.get("_config", {})

            # Extract activity name
            activity_name = self._find_column_value(
                record, config.get("activity_col", "")
            )
            if not activity_name:
                continue

            # Extract CO2e factor - try primary column then fallbacks
            co2e_col = config.get("co2e_col", "")
            co2e_value = self._find_column_value(record, co2e_col)

            # TASK-DATA-P8-BUG-003: Try fallback column names if primary fails
            if not co2e_value:
                fallbacks = config.get("co2e_col_fallback", [])
                for fallback_col in fallbacks:
                    co2e_value = self._find_column_value(record, fallback_col)
                    if co2e_value:
                        break

            if not co2e_value:
                continue

            try:
                co2e_float = float(co2e_value)
            except (ValueError, TypeError):
                continue

            # Skip invalid/zero factors
            if co2e_float <= 0:
                continue

            # Determine unit
            unit = self._determine_unit(record, config, co2e_col)

            # Create external ID
            sheet_name = record.get("_sheet_name", "unknown")
            external_id = f"DEFRA_{sheet_name}_{activity_name}".replace(" ", "_")
            external_id = re.sub(r'[^\w\-]', '_', external_id)[:200]

            transformed.append({
                "activity_name": str(activity_name).strip(),
                "co2e_factor": co2e_float,
                "unit": unit,
                "data_source": "DEFRA",  # Set data source for BOM display
                "scope": config.get("scope", "Scope 3"),
                "category": config.get("category", "other"),
                "geography": "GB",
                "reference_year": self.reference_year,
                "data_quality_rating": 0.88,
                "external_id": external_id,
                "metadata": {
                    "source_sheet": sheet_name,
                    "source_row": record.get("_row_idx"),
                }
            })

        return transformed

    def _find_column_value(
        self,
        record: Dict[str, Any],
        column_hint: str
    ) -> Any:
        """
        Find value by partial column name match.

        Searches record keys for a case-insensitive partial match
        with the column hint, ignoring internal metadata keys.

        Args:
            record: Record dictionary with column values
            column_hint: Partial column name to search for

        Returns:
            Value from matching column, or None if not found
        """
        column_hint_lower = column_hint.lower()
        for key, value in record.items():
            if key.startswith("_"):
                continue
            if column_hint_lower in key.lower():
                return value
        return None

    def _determine_unit(
        self,
        record: Dict[str, Any],
        config: Dict[str, Any],
        co2e_col: str
    ) -> str:
        """
        Determine unit from column name or explicit unit column.

        First tries explicit unit column if configured.
        Falls back to extracting unit from CO2e column name
        (e.g., "kg CO2e per kWh" -> "kWh").

        Args:
            record: Record dictionary
            config: Sheet configuration
            co2e_col: CO2e column name

        Returns:
            Unit string, defaults to "unit" if not determinable
        """
        # Try explicit unit column first
        if config.get("unit_col"):
            unit_val = self._find_column_value(record, config["unit_col"])
            if unit_val:
                return str(unit_val)

        # Extract from CO2e column name
        # e.g., "kg CO2e per kWh" -> "kWh"
        match = re.search(r'per\s+(\w+)', co2e_col.lower())
        if match:
            return match.group(1)

        return "unit"


__all__ = [
    "DEFRAEmissionFactorsIngestion",
]
