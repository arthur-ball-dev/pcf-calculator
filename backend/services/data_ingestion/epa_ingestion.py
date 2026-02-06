"""
EPA GHG Emission Factors Hub Data Connector.

TASK-DATA-P5-002: EPA Data Connector
TASK-DATA-P7-007: Fix EPA Connector URLs and Sheet Names
TASK-DATA-P8-BUG: Fix EPA Fuels sheet name and eGRID column headers
TASK-DATA-P8-BUG-002: Fix EPA Fuels multi-table parsing
TASK-DATA-P10: Expand EPA import with Table 8 (Transport) and Table 9 (Materials)

This module implements the EPA GHG Emission Factors Hub connector that:
- Downloads Excel files from the EPA website
- Parses fuel and eGRID emission factor data
- Parses Table 8 (Transportation) for Scope 3 transport factors
- Parses Table 9 (Materials/Waste) for Scope 3 materials factors
- Transforms records to internal schema with unit conversions
- Handles lb/MWh to kg/kWh conversion for eGRID data
- Handles multi-table parsing for 2024 format files
- Backward-compatible with older file formats for testing

Data Sources:
- EPA GHG Emission Factors Hub: https://www.epa.gov/climateleadership/ghg-emission-factors-hub
- eGRID Subregion Data: https://www.epa.gov/egrid

Usage:
    from backend.services.data_ingestion.epa_ingestion import EPAEmissionFactorsIngestion

    ingestion = EPAEmissionFactorsIngestion(
        db=session,
        data_source_id=source_id,
        file_key="fuels"  # or "egrid"
    )
    result = await ingestion.execute_sync()
"""

import io
import re
from typing import List, Dict, Any, Optional, Tuple

from openpyxl import load_workbook
import httpx

from backend.services.data_ingestion.base import BaseDataIngestion
from backend.services.data_ingestion.transformers.unit_normalizer import normalize_unit


class EPAEmissionFactorsIngestion(BaseDataIngestion):
    """
    EPA GHG Emission Factors Hub data connector.

    Supports two file types:
    - fuels: General fuel emission factors (natural gas, diesel, gasoline, etc.)
    - egrid: eGRID electricity emission factors by subregion

    Attributes:
        FILES: Configuration dict with URLs and sheet names for each file type
        file_key: Selected file type ("fuels" or "egrid")
        file_config: Configuration for the selected file type
    """

    # EPA file URLs (may need annual updates)
    # TASK-DATA-P7-007: Updated URLs and sheet names
    # BUG-DATA-001: Fixed fuels URL (was 404)
    # BUG-DATA-003: Fixed eGRID sheet names (SUBRGN22 -> SRL22)
    # TASK-DATA-P8-BUG-001: Fixed fuels sheet name (2024 file uses single sheet)
    # TASK-DATA-P8-BUG-002: Added fallback sheet names for fuel files backward compatibility
    FILES = {
        "egrid": {
            "url": "https://www.epa.gov/system/files/documents/2024-01/egrid2022_data.xlsx",
            # SRL22 = Subregion Level (2024 format)
            "sheets": ["SRL22", "US22"],
            "type": "electricity",
        },
        "fuels": {
            "url": "https://www.epa.gov/system/files/documents/2024-02/ghg-emission-factors-hub-2024.xlsx",
            # 2024 file has single combined sheet, older tests use separate sheets
            "sheets": ["Emission Factors Hub", "Table 1 - Fuel", "Table 2 - Mobile"],
            "type": "combustion",
        },
    }

    # Fallback sheet names for backward compatibility with test fixtures
    # TASK-DATA-P8-BUG-002: These are tried if primary sheets are not found
    FALLBACK_SHEETS = {
        "egrid": ["SUBRGN22"],  # Old test fixture sheet name
    }

    # Tables to parse from the "Emission Factors Hub" sheet
    # TASK-DATA-P8-BUG-002: Multi-table parsing configuration
    # Table 1: Stationary Combustion - most comprehensive fuel data
    # Table 2: Mobile Combustion CO2 - simpler format with explicit units
    # TASK-DATA-P10: Added Table 8 and Table 9 for expanded materials/transport
    # Table 8: Transportation (Scope 3) - upstream/downstream transport
    # Table 9: Materials/Waste (Scope 3) - end-of-life treatment factors
    FUELS_TABLES_TO_PARSE = ["Table 1", "Table 2", "Table 8", "Table 9"]

    # Category headers to skip in Table 9 Materials sheet
    TABLE_9_CATEGORY_HEADERS = [
        "Metals", "Plastics", "Paper Products", "Glass", "Organics",
        "Wood Products", "Mixed Categories", "Electronics", "Construction"
    ]

    def __init__(self, *args, file_key: str = "fuels", **kwargs) -> None:
        """
        Initialize the EPA ingestion connector.

        Args:
            *args: Positional arguments passed to BaseDataIngestion
            file_key: Which EPA file to process ("fuels" or "egrid")
            **kwargs: Keyword arguments passed to BaseDataIngestion

        Raises:
            KeyError: If file_key is not "fuels" or "egrid"
        """
        super().__init__(*args, **kwargs)
        self.file_key = file_key
        self.file_config = self.FILES[file_key]

    async def fetch_raw_data(self) -> bytes:
        """
        Download EPA Excel file.

        Returns:
            Raw bytes of the Excel file content

        Raises:
            httpx.HTTPStatusError: On HTTP errors (4xx, 5xx)
            httpx.TimeoutException: On request timeout
        """
        url = self.file_config["url"]

        async with httpx.AsyncClient(timeout=60.0) as client:
            response = await client.get(url)
            response.raise_for_status()
            return response.content

    async def parse_data(self, raw_data: bytes) -> List[Dict[str, Any]]:
        """
        Parse EPA Excel file into records.

        For eGRID files: Simple single-table parsing with column headers.
        For Fuels files: Multi-table parsing detecting "Table N" markers,
                        with fallback to single-table for older formats.

        Args:
            raw_data: Raw bytes from fetch_raw_data()

        Returns:
            List of dictionaries, each representing one emission factor record

        Raises:
            Exception: On corrupted or invalid Excel file
        """
        workbook = load_workbook(io.BytesIO(raw_data), read_only=True)
        records = []

        # Build list of sheets to try (primary + fallback)
        sheets_to_try = list(self.file_config["sheets"])
        if self.file_key in self.FALLBACK_SHEETS:
            sheets_to_try.extend(self.FALLBACK_SHEETS[self.file_key])

        for sheet_name in sheets_to_try:
            if sheet_name not in workbook.sheetnames:
                continue

            sheet = workbook[sheet_name]

            # TASK-DATA-P8-BUG-002: Handle multi-table format for 2024 fuels file
            if self.file_key == "fuels" and sheet_name == "Emission Factors Hub":
                records.extend(self._parse_multi_table_sheet(sheet, sheet_name))
            else:
                # Original single-table parsing for eGRID, older fuel formats, and tests
                records.extend(self._parse_single_table_sheet(sheet, sheet_name))

        workbook.close()
        return records

    def _parse_single_table_sheet(
        self, sheet, sheet_name: str
    ) -> List[Dict[str, Any]]:
        """
        Parse a single-table sheet (original behavior).

        First non-empty row is treated as headers, subsequent rows as data.

        Args:
            sheet: openpyxl worksheet object
            sheet_name: Name of the sheet for metadata

        Returns:
            List of parsed records
        """
        records = []
        headers = None

        for row_idx, row in enumerate(sheet.iter_rows(values_only=True)):
            # First non-empty row is headers
            if headers is None and any(row):
                headers = [
                    str(h).strip() if h else f"col_{i}"
                    for i, h in enumerate(row)
                ]
                continue

            if headers and any(row):
                record = dict(zip(headers, row))
                record["_source_sheet"] = sheet_name
                record["_source_row"] = row_idx + 1
                records.append(record)

        return records

    def _parse_multi_table_sheet(
        self, sheet, sheet_name: str
    ) -> List[Dict[str, Any]]:
        """
        Parse a multi-table sheet (2024 Emission Factors Hub format).

        TASK-DATA-P8-BUG-002: Detects "Table N" markers and parses each
        table with its own headers.

        The 2024 format has:
        - "Table 1", "Table 2", etc. markers in column B (index 1)
        - Headers in rows following the table marker
        - Data offset by columns (typically starts in column C, index 2)

        Args:
            sheet: openpyxl worksheet object
            sheet_name: Name of the sheet for metadata

        Returns:
            List of parsed records from all tables
        """
        records = []

        # First pass: collect all rows
        all_rows = list(sheet.iter_rows(values_only=True))

        # Find table boundaries
        table_positions = self._find_table_positions(all_rows)

        # Parse each table that we're interested in
        for table_name, start_row in table_positions.items():
            if table_name not in self.FUELS_TABLES_TO_PARSE:
                continue

            # Find end of this table (start of next table or end of data)
            next_tables = [
                pos for name, pos in table_positions.items()
                if pos > start_row
            ]
            end_row = min(next_tables) if next_tables else len(all_rows)

            # Parse this table's data
            table_records = self._parse_table_section(
                all_rows, start_row, end_row, table_name, sheet_name
            )
            records.extend(table_records)

        return records

    def _find_table_positions(
        self, rows: List[Tuple]
    ) -> Dict[str, int]:
        """
        Find row positions of table markers.

        Table markers in the 2024 format are "Table N" in column B (index 1).

        Args:
            rows: List of row tuples from the sheet

        Returns:
            Dict mapping table name to starting row index
        """
        positions = {}
        table_pattern = re.compile(r"^Table\s+(\d+)$")

        for row_idx, row in enumerate(rows):
            if len(row) > 1 and row[1]:
                cell_value = str(row[1]).strip()
                match = table_pattern.match(cell_value)
                if match:
                    positions[f"Table {match.group(1)}"] = row_idx

        return positions

    def _parse_table_section(
        self,
        rows: List[Tuple],
        start_row: int,
        end_row: int,
        table_name: str,
        sheet_name: str
    ) -> List[Dict[str, Any]]:
        """
        Parse a single table section from the multi-table sheet.

        Args:
            rows: All rows from the sheet
            start_row: Row index where table marker was found
            end_row: Row index where this table ends
            table_name: Name of the table (e.g., "Table 1")
            sheet_name: Name of the sheet for metadata

        Returns:
            List of parsed records from this table
        """
        records = []
        headers = None
        data_col_offset = 2  # Data typically starts at column C (index 2)
        header_row_idx = start_row

        # Search for header row after the table marker
        for row_idx in range(start_row + 1, min(start_row + 5, end_row)):
            if row_idx >= len(rows):
                break

            row = rows[row_idx]
            # Look for a row that has column headers (e.g., "Fuel Type")
            if self._is_header_row(row, table_name):
                headers = self._extract_headers(row, data_col_offset)
                header_row_idx = row_idx
                break

        if not headers:
            return records

        # Parse data rows after headers
        # Skip header row and any unit descriptor row
        data_start = header_row_idx + 1
        if data_start < len(rows) and self._is_unit_row(rows[data_start]):
            data_start += 1

        for row_idx in range(data_start, end_row):
            if row_idx >= len(rows):
                break

            row = rows[row_idx]

            # Skip empty rows and category header rows
            if not any(row[data_col_offset:]):
                continue

            # Skip rows that look like category headers (e.g., "Coal and Coke")
            if self._is_category_header(row, data_col_offset):
                continue

            # Skip rows that are notes or sources
            if self._is_notes_row(row, data_col_offset):
                continue

            # Extract record
            record = self._extract_record(row, headers, data_col_offset)
            if record:
                record["_source_sheet"] = sheet_name
                record["_source_row"] = row_idx + 1
                record["_source_table"] = table_name
                records.append(record)

        return records

    def _is_header_row(self, row: Tuple, table_name: str) -> bool:
        """
        Check if a row is a header row for the given table.

        TASK-DATA-P10: Added support for Table 8 and Table 9 header detection.

        Args:
            row: Row tuple from the sheet
            table_name: Name of the table

        Returns:
            True if this is a header row
        """
        if len(row) < 3:
            return False

        # Column C (index 2) should have identifying header
        col_c = str(row[2]).strip() if row[2] else ""

        if table_name == "Table 1":
            return col_c == "Fuel Type"
        elif table_name == "Table 2":
            return col_c == "Fuel Type"
        elif table_name == "Table 8":
            # Table 8: Transportation - header has "Vehicle Type"
            return col_c == "Vehicle Type"
        elif table_name == "Table 9":
            # Table 9: Materials/Waste - header has "Material"
            return col_c == "Material"

        return False

    def _is_unit_row(self, row: Tuple) -> bool:
        """
        Check if a row is a unit descriptor row (e.g., "kg CO2 per mmBtu").

        Args:
            row: Row tuple from the sheet

        Returns:
            True if this is a unit row
        """
        if len(row) < 4:
            return False

        # Unit rows typically don't have a fuel name in column C
        col_c = str(row[2]).strip() if row[2] else ""
        col_d = str(row[3]).strip() if row[3] else ""

        # Unit row: column C is empty or column D contains unit patterns
        if not col_c and ("per" in col_d.lower() or "mmBtu" in col_d):
            return True

        # Or column D/E contains unit patterns
        for col_idx in range(3, min(8, len(row))):
            if row[col_idx]:
                cell = str(row[col_idx]).lower()
                if "per" in cell and ("mmBtu" in cell or "kg" in cell or "unit" in cell):
                    return True

        return False

    def _is_category_header(self, row: Tuple, col_offset: int) -> bool:
        """
        Check if a row is a category header (e.g., "Coal and Coke").

        Category headers have text in the first data column but no
        numeric values in subsequent columns.

        Args:
            row: Row tuple from the sheet
            col_offset: Column offset where data starts

        Returns:
            True if this is a category header row
        """
        if len(row) <= col_offset:
            return False

        first_val = row[col_offset]
        if not first_val:
            return False

        # Check if subsequent columns have numeric values
        for col_idx in range(col_offset + 1, min(col_offset + 5, len(row))):
            if col_idx < len(row) and row[col_idx] is not None:
                try:
                    float(row[col_idx])
                    return False  # Has numeric value, not a category header
                except (ValueError, TypeError):
                    pass

        return True

    def _is_notes_row(self, row: Tuple, col_offset: int) -> bool:
        """
        Check if a row is a notes/source row.

        Args:
            row: Row tuple from the sheet
            col_offset: Column offset where data starts

        Returns:
            True if this is a notes row
        """
        if len(row) <= col_offset:
            return False

        first_val = str(row[col_offset]).strip() if row[col_offset] else ""

        # Notes rows typically start with these patterns
        notes_patterns = ["Source:", "Notes:", "http", "www.", "More information"]
        return any(first_val.startswith(p) for p in notes_patterns)

    def _extract_headers(self, row: Tuple, col_offset: int) -> List[str]:
        """
        Extract headers from a header row.

        TASK-DATA-P10: Enhanced to normalize headers by removing newlines
        and collapsing whitespace.

        Args:
            row: Header row tuple
            col_offset: Column offset where data starts

        Returns:
            List of header strings
        """
        headers = []
        for i, val in enumerate(row[col_offset:], start=col_offset):
            if val:
                # Normalize: remove newlines, collapse whitespace
                header = str(val).replace("\n", " ").replace("\r", "")
                header = " ".join(header.split())
                headers.append(header.strip())
            else:
                headers.append(f"col_{i}")
        return headers

    def _extract_record(
        self, row: Tuple, headers: List[str], col_offset: int
    ) -> Optional[Dict[str, Any]]:
        """
        Extract a record from a data row.

        Args:
            row: Data row tuple
            headers: List of column headers
            col_offset: Column offset where data starts

        Returns:
            Record dict or None if row is invalid
        """
        if len(row) <= col_offset:
            return None

        # Extract values starting from col_offset
        values = list(row[col_offset:])

        # Pad or truncate to match headers
        while len(values) < len(headers):
            values.append(None)

        record = dict(zip(headers, values))
        return record

    async def transform_data(
        self, parsed_data: List[Dict[str, Any]]
    ) -> List[Dict[str, Any]]:
        """
        Transform EPA data to internal schema.

        Routes to appropriate transform method based on file_key and source table.
        TASK-DATA-P10: Added routing for Table 8 and Table 9.

        Args:
            parsed_data: List of source-format records from parse_data()

        Returns:
            List of records matching EmissionFactor schema
        """
        transformed = []

        for record in parsed_data:
            # Handle different EPA file formats
            if self.file_key == "egrid":
                transformed.extend(self._transform_egrid_record(record))
            else:
                # Route based on source table for fuels file
                source_table = record.get("_source_table", "")
                if source_table == "Table 8":
                    transformed.extend(self._transform_transport_record(record))
                elif source_table == "Table 9":
                    transformed.extend(self._transform_materials_record(record))
                else:
                    transformed.extend(self._transform_fuel_record(record))

        return transformed

    def _transform_fuel_record(self, record: Dict[str, Any]) -> List[Dict[str, Any]]:
        """
        Transform fuel/combustion emission factor record.

        TASK-DATA-P8-BUG-002: Enhanced to handle 2024 multi-table format.
        Also backward-compatible with older single-sheet formats for testing.

        Handles:
        - Table 1 (Stationary Combustion): CO2 Factor in kg CO2 per mmBtu
        - Table 2 (Mobile Combustion CO2): kg CO2 per unit with explicit Unit column
        - Old format: kg CO2e per unit column

        Args:
            record: Single parsed record from EPA fuel Excel file

        Returns:
            List containing zero or one transformed records
        """
        results = []

        # Get activity name (Fuel Type column)
        activity_name = (
            record.get("Fuel Type") or
            record.get("Vehicle Type") or
            record.get("Source") or
            record.get("Activity") or
            record.get("Fuel/Activity") or
            self._find_column_value(record, "fuel") or
            self._find_column_value(record, "activity") or
            self._find_column_value(record, "source")
        )
        if not activity_name:
            return results

        # Skip if activity name looks like a category header
        activity_str = str(activity_name).strip()
        if not activity_str or activity_str in ["Coal and Coke", "Other Fuels - Solid"]:
            return results

        # TASK-DATA-P8-BUG-002: Handle different table formats
        source_table = record.get("_source_table", "")

        # Determine CO2e factor based on format
        co2e = None
        unit = None

        if source_table == "Table 2":
            # Table 2: Mobile Combustion CO2 - has kg CO2 per unit
            co2e = record.get("kg CO2 per unit")
            unit = record.get("Unit") or "unit"
        elif source_table == "Table 1":
            # Table 1: Stationary Combustion - has CO2 Factor (kg CO2 per mmBtu)
            co2e = (
                record.get("CO2 Factor") or
                record.get("kg CO2 per mmBtu") or
                self._find_column_value(record, "kg CO2") or
                self._find_column_value(record, "CO2 Factor")
            )
            unit = record.get("Unit") or "mmBtu"
        else:
            # Old format (test fixtures): kg CO2e per unit
            co2e = (
                record.get("kg CO2e per unit") or
                record.get("CO2 Factor") or
                record.get("kg CO2 per mmBtu") or
                record.get("CO2e Factor") or
                record.get("kg CO2e") or
                self._find_column_value(record, "kg CO2e") or
                self._find_column_value(record, "kg CO2") or
                self._find_column_value(record, "CO2 Factor")
            )
            unit = record.get("Unit") or record.get("Units") or "unit"

        if not co2e:
            return results

        try:
            co2e_value = float(co2e)
        except (ValueError, TypeError):
            return results

        # Skip invalid/zero factors
        if co2e_value <= 0:
            return results

        # Clean up unit
        unit_str = str(unit).strip() if unit else "unit"

        # Apply unit normalization
        norm_result = normalize_unit(co2e_value, unit_str)

        results.append({
            "activity_name": activity_str,
            "co2e_factor": norm_result.normalized_factor,
            "unit": norm_result.normalized_unit,
            "data_source": "EPA",  # Set data source for BOM display
            "scope": self._determine_scope(record),
            "category": self.file_config["type"],
            "geography": "US",
            "reference_year": 2023,
            "data_quality_rating": 0.90,
            "external_id": f"EPA_{self.file_key}_{activity_str}".replace(" ", "_"),
            # Unit normalization audit fields
            "original_unit": norm_result.original_unit if norm_result.was_normalized else None,
            "original_co2e_factor": norm_result.original_factor if norm_result.was_normalized else None,
            "conversion_factor": norm_result.conversion_factor,
            "normalized_at": norm_result.normalized_at,
            "metadata": {
                "source_sheet": record.get("_source_sheet"),
                "source_row": record.get("_source_row"),
                "source_table": source_table,
            }
        })

        return results

    def _find_column_value(self, record: Dict[str, Any], hint: str) -> Any:
        """
        Find value by partial column name match.

        Searches record keys for a case-insensitive partial match,
        ignoring internal metadata keys (starting with _).

        Args:
            record: Record dictionary with column values
            hint: Partial column name to search for

        Returns:
            Value from matching column, or None if not found
        """
        hint_lower = hint.lower()
        for key, value in record.items():
            if key.startswith("_"):
                continue
            if hint_lower in key.lower() and value is not None:
                return value
        return None

    def _transform_egrid_record(self, record: Dict[str, Any]) -> List[Dict[str, Any]]:
        """
        Transform eGRID electricity emission factor record.

        Converts EPA's lb/MWh units to kg/kWh for internal use:
        - 1 lb = 0.453592 kg
        - 1 MWh = 1000 kWh
        - lb/MWh * 0.453592 / 1000 = kg/kWh

        TASK-DATA-P8-BUG-002: Handle both short codes and long column names
        from eGRID files which have two header rows.

        Short code column names (row 1): SUBRGN, SRCO2RTA
        Long column names (row 0): "eGRID subregion acronym", "eGRID subregion..."

        Args:
            record: Single parsed record from EPA eGRID Excel file

        Returns:
            List containing zero or one transformed records
        """
        results = []

        # TASK-DATA-P8-BUG-002: Handle both short and long column names
        # Short codes (if parser uses row 1 headers)
        subregion = record.get("SUBRGN") or record.get("Subregion")

        # Long names (if parser uses row 0 headers)
        if not subregion:
            subregion = (
                record.get("eGRID subregion acronym") or
                self._find_column_value(record, "subregion acronym") or
                self._find_column_value(record, "SUBRGN")
            )

        if not subregion:
            return results

        # eGRID uses lb CO2e per MWh, convert to kg CO2e per kWh
        # Short code column name
        co2_rate = record.get("SRCO2RTA") or record.get("CO2 Rate")

        # Long column name patterns (2024 format)
        if not co2_rate:
            co2_rate = (
                self._find_column_value(
                    record, "CO2 equivalent total output emission rate"
                ) or
                self._find_column_value(record, "CO2e output emission rate") or
                self._find_column_value(record, "lb/MWh") or
                self._find_column_value(record, "SRCO2RTA")
            )

        if not co2_rate:
            return results

        try:
            # Convert lb/MWh to kg/kWh
            co2e_value = float(co2_rate) * 0.453592 / 1000
        except (ValueError, TypeError):
            return results

        # Skip invalid/zero factors
        if co2e_value <= 0:
            return results

        # Apply unit normalization (kWh is already standard, but record for consistency)
        norm_result = normalize_unit(co2e_value, "kWh")

        results.append({
            "activity_name": f"Grid Electricity - {subregion}",
            "co2e_factor": norm_result.normalized_factor,
            "unit": norm_result.normalized_unit,
            "data_source": "EPA",  # Set data source for BOM display
            "scope": "Scope 2",
            "category": "electricity",
            "geography": "US",
            "reference_year": 2022,
            "data_quality_rating": 0.95,  # eGRID is highly reliable
            "external_id": f"EPA_eGRID_{subregion}",
            # Unit normalization audit fields (lb/MWh -> kg/kWh already done above)
            "original_unit": "lb/MWh",
            "original_co2e_factor": float(co2_rate),
            "conversion_factor": 0.000453592,  # lb/MWh to kg/kWh
            "normalized_at": None,  # Manual conversion, not via normalizer
            "metadata": {
                "subregion_code": subregion,
                "source_sheet": record.get("_source_sheet"),
            }
        })

        return results

    def _determine_scope(self, record: Dict[str, Any]) -> str:
        """
        Determine GHG Protocol scope from EPA data.

        Mapping:
        - Electricity -> Scope 2
        - Mobile/Transportation -> Scope 1
        - Other (stationary combustion, etc.) -> Scope 1

        Args:
            record: Parsed record containing Category field or _source_table

        Returns:
            GHG Protocol scope string ("Scope 1" or "Scope 2")
        """
        # Check source table for scope determination
        source_table = record.get("_source_table", "")
        if "Table 2" in source_table:
            return "Scope 1"  # Mobile combustion

        category = str(record.get("Category", "")).lower()
        if "electricity" in category:
            return "Scope 2"
        elif "mobile" in category or "transport" in category:
            return "Scope 1"
        else:
            return "Scope 1"

    def _transform_transport_record(
        self, record: Dict[str, Any]
    ) -> List[Dict[str, Any]]:
        """
        Transform Table 8 transportation emission factor record.

        TASK-DATA-P10: New method for Scope 3 Category 4/9 transportation factors.

        Table 8 contains transportation factors in kg CO2 per unit (vehicle-mile
        or short ton-mile). These are Scope 3 upstream/downstream transport factors.

        Headers:
        - Vehicle Type
        - CO2 Factor (kg CO2 / unit)
        - CH4 Factor (g CH4 / unit)
        - N2O Factor (g N2O / unit)
        - Units

        Args:
            record: Single parsed record from EPA Table 8

        Returns:
            List containing zero or one transformed records
        """
        results = []

        # Get vehicle type
        vehicle_type = record.get("Vehicle Type")
        if not vehicle_type:
            return results

        vehicle_str = str(vehicle_type).strip()
        if not vehicle_str:
            return results

        # Get CO2 factor - column header is "CO2 Factor (kg CO2 / unit)"
        co2_factor = (
            self._find_column_value(record, "CO2 Factor") or
            self._find_column_value(record, "kg CO2")
        )

        if not co2_factor:
            return results

        try:
            co2e_value = float(co2_factor)
        except (ValueError, TypeError):
            return results

        if co2e_value <= 0:
            return results

        # Get unit from Units column
        unit = record.get("Units") or self._find_column_value(record, "unit")
        unit_str = str(unit).strip() if unit else "unit"

        # Determine category based on unit
        # vehicle-mile = passenger/light transport
        # short ton-mile = freight transport
        if "ton-mile" in unit_str.lower():
            category = "freight_transport"
        else:
            category = "passenger_transport"

        # Apply unit normalization
        norm_result = normalize_unit(co2e_value, unit_str)

        results.append({
            "activity_name": f"Transport - {vehicle_str}",
            "co2e_factor": norm_result.normalized_factor,
            "unit": norm_result.normalized_unit,
            "data_source": "EPA",
            "scope": "Scope 3",
            "category": category,
            "geography": "US",
            "reference_year": 2023,
            "data_quality_rating": 0.90,
            "external_id": f"EPA_transport_{vehicle_str}_{unit_str}".replace(
                " ", "_"
            ).replace("-", "_"),
            # Unit normalization audit fields
            "original_unit": norm_result.original_unit if norm_result.was_normalized else None,
            "original_co2e_factor": norm_result.original_factor if norm_result.was_normalized else None,
            "conversion_factor": norm_result.conversion_factor,
            "normalized_at": norm_result.normalized_at,
            "metadata": {
                "source_sheet": record.get("_source_sheet"),
                "source_row": record.get("_source_row"),
                "source_table": "Table 8",
                "vehicle_type": vehicle_str,
            }
        })

        return results

    def _transform_materials_record(
        self, record: Dict[str, Any]
    ) -> List[Dict[str, Any]]:
        """
        Transform Table 9 materials/waste emission factor record.

        TASK-DATA-P10: New method for Scope 3 Category 5/12 materials factors.

        Table 9 contains end-of-life treatment factors for various materials.
        Units are Metric Tons CO2e per Short Ton of Material.
        We convert to kg CO2e per kg (divide by 1000 since 1 ton = 1000 kg,
        and metric tons CO2e per short ton ≈ kg CO2e per kg).

        Creates separate emission factors for each disposal method:
        - Recycled
        - Landfilled
        - Combusted
        - Composted
        - Anaerobically Digested

        Args:
            record: Single parsed record from EPA Table 9

        Returns:
            List of transformed records (one per disposal method with valid data)
        """
        results = []

        # Get material name
        material_name = record.get("Material")
        if not material_name:
            return results

        material_str = str(material_name).strip()
        if not material_str:
            return results

        # Skip category headers and notes rows
        skip_patterns = [
            "Metals", "Plastics", "Paper Products", "Glass", "Organics",
            "Wood Products", "Mixed Categories", "Electronics", "Construction",
            "Source:", "Notes:", "^A", "^B", "^C", "^D", "NA",
            "More information", "http", "www."
        ]
        if any(material_str.startswith(p) for p in skip_patterns):
            return results

        # Disposal methods and their column names
        # TASK-DATA-P10: Map column headers to disposal method names
        # Note: Excel superscripts become "RecycledA", "LandfilledB" etc.
        disposal_methods = {
            "Recycled": ["Recycled", "RecycledA"],
            "Landfilled": ["Landfilled", "LandfilledB"],
            "Combusted": ["Combusted", "CombustedC"],
            "Composted": ["Composted", "CompostedD"],
            "Anaerobically Digested": [
                "Anaerobically Digested",
                "Anaerobically Digested (Dry",  # Partial match works
            ]
        }

        # Determine material category for grouping
        material_category = self._categorize_material(material_str)

        for method_name, col_variants in disposal_methods.items():
            co2e_value = None

            # Try each column variant
            for col_name in col_variants:
                value = self._find_column_value(record, col_name)
                if value is not None and str(value).strip().lower() not in ["na", ""]:
                    try:
                        co2e_value = float(value)
                        break
                    except (ValueError, TypeError):
                        continue

            if co2e_value is None or co2e_value <= 0:
                continue

            # Convert from Metric Tons CO2e per Short Ton to kg CO2e per kg
            # 1 short ton = 907.185 kg, 1 metric ton = 1000 kg
            # Factor is already approximately kg CO2e per kg since
            # metric ton CO2e / short ton ≈ 1.1 kg CO2e / kg
            # We'll use the factor as-is since it's effectively kg/kg scale
            co2e_per_kg = co2e_value * (1000 / 907.185)

            # Create unique external ID
            external_id = (
                f"EPA_material_{material_str}_{method_name}"
                .replace(" ", "_")
                .replace("-", "_")
                .replace("(", "")
                .replace(")", "")
            )[:200]

            # Apply unit normalization (already converted to kg, record audit trail)
            norm_result = normalize_unit(co2e_per_kg, "kg")

            results.append({
                "activity_name": f"{material_str} - {method_name}",
                "co2e_factor": round(norm_result.normalized_factor, 4),
                "unit": norm_result.normalized_unit,
                "data_source": "EPA",
                "scope": "Scope 3",
                "category": f"materials_{material_category}",
                "geography": "US",
                "reference_year": 2023,
                "data_quality_rating": 0.88,
                "external_id": external_id,
                # Unit normalization audit fields (manual conversion from short ton)
                "original_unit": "Metric Tons CO2e per Short Ton",
                "original_co2e_factor": co2e_value,
                "conversion_factor": 1000 / 907.185,  # Short ton to kg
                "normalized_at": None,  # Manual conversion
                "metadata": {
                    "source_sheet": record.get("_source_sheet"),
                    "source_row": record.get("_source_row"),
                    "source_table": "Table 9",
                    "material": material_str,
                    "disposal_method": method_name,
                }
            })

        return results

    def _categorize_material(self, material: str) -> str:
        """
        Categorize a material by type for grouping.

        TASK-DATA-P10: Helper for _transform_materials_record.

        Args:
            material: Material name string

        Returns:
            Category string (e.g., "metals", "plastics", "paper", "electronics")
        """
        material_lower = material.lower()

        # Metal materials
        if any(m in material_lower for m in [
            "aluminum", "steel", "copper", "metal", "iron"
        ]):
            return "metals"

        # Plastic materials
        if any(p in material_lower for p in [
            "hdpe", "ldpe", "pet", "lldpe", "pp", "ps", "pvc", "pla", "plastic"
        ]):
            return "plastics"

        # Paper/cardboard
        if any(p in material_lower for p in [
            "paper", "corrugated", "magazine", "newspaper", "textbook", "phonebook"
        ]):
            return "paper"

        # Electronics
        if any(e in material_lower for e in [
            "cpu", "display", "electronic", "peripheral", "device", "hard-copy"
        ]):
            return "electronics"

        # Glass
        if "glass" in material_lower or "fiberglass" in material_lower:
            return "glass"

        # Wood
        if any(w in material_lower for w in [
            "lumber", "fiberboard", "wood", "mdf"
        ]):
            return "wood"

        # Organics/food
        if any(o in material_lower for o in [
            "food", "beef", "poultry", "grain", "bread", "fruit", "vegetable",
            "dairy", "yard", "grass", "leaves", "branch", "organic"
        ]):
            return "organics"

        # Construction
        if any(c in material_lower for c in [
            "concrete", "asphalt", "drywall", "brick", "insulation", "vinyl",
            "carpet", "shingle", "flooring"
        ]):
            return "construction"

        # Rubber/tires
        if "tire" in material_lower or "rubber" in material_lower:
            return "rubber"

        return "other"


__all__ = [
    "EPAEmissionFactorsIngestion",
]
