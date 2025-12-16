"""
Pytest fixtures for test data files.

TASK-DATA-P5-002: EPA Data Connector - Phase A Tests

This module provides pytest fixtures that programmatically create
test data files using openpyxl. The fixtures are available to all
tests in the backend/tests directory.
"""

import pytest
from io import BytesIO
from pathlib import Path


@pytest.fixture(scope="session")
def fixtures_dir() -> Path:
    """Return the path to the fixtures directory."""
    return Path(__file__).parent


@pytest.fixture(scope="session")
def epa_sample_excel_path(fixtures_dir) -> Path:
    """
    Create a sample EPA emission factors Excel file for testing.

    This fixture creates the file once per test session and returns
    the path to it. The file mimics the structure of the actual
    EPA GHG Emission Factors Hub spreadsheet.

    Returns:
        Path to the created epa_sample.xlsx file
    """
    try:
        from openpyxl import Workbook
    except ImportError:
        pytest.skip("openpyxl not available")

    output_path = fixtures_dir / "epa_sample.xlsx"

    # Create workbook
    wb = Workbook()

    # =========================================================================
    # Sheet 1: "Table 1 - Fuel" - Stationary combustion emission factors
    # =========================================================================
    ws_fuel = wb.active
    ws_fuel.title = "Table 1 - Fuel"

    # Header row (matches EPA format)
    fuel_headers = ["Fuel Type", "kg CO2e per unit", "Unit", "Category"]
    for col, header in enumerate(fuel_headers, 1):
        ws_fuel.cell(row=1, column=col, value=header)

    # Sample fuel emission factors (representative of actual EPA data)
    fuel_data = [
        ("Natural Gas", 2.75, "kg", "Stationary Combustion"),
        ("Diesel", 10.21, "L", "Stationary Combustion"),
        ("Gasoline", 8.89, "L", "Stationary Combustion"),
        ("Propane", 6.35, "kg", "Stationary Combustion"),
        ("Residual Fuel Oil", 11.27, "L", "Stationary Combustion"),
        ("Coal (Bituminous)", 2.563, "kg", "Stationary Combustion"),
        ("Coal (Sub-bituminous)", 2.412, "kg", "Stationary Combustion"),
        ("Kerosene", 9.75, "L", "Stationary Combustion"),
        ("LPG", 5.68, "L", "Stationary Combustion"),
        ("Fuel Oil No. 1", 10.15, "L", "Stationary Combustion"),
    ]

    for row_idx, (fuel_type, factor, unit, category) in enumerate(fuel_data, 2):
        ws_fuel.cell(row=row_idx, column=1, value=fuel_type)
        ws_fuel.cell(row=row_idx, column=2, value=factor)
        ws_fuel.cell(row=row_idx, column=3, value=unit)
        ws_fuel.cell(row=row_idx, column=4, value=category)

    # =========================================================================
    # Sheet 2: "Table 2 - Mobile" - Mobile combustion emission factors
    # =========================================================================
    ws_mobile = wb.create_sheet("Table 2 - Mobile")

    mobile_headers = ["Vehicle Type", "CO2e Factor", "Unit", "Category"]
    for col, header in enumerate(mobile_headers, 1):
        ws_mobile.cell(row=1, column=col, value=header)

    mobile_data = [
        ("Passenger Car", 0.21, "vehicle-mile", "Mobile - On-Road"),
        ("Light-Duty Truck", 0.29, "vehicle-mile", "Mobile - On-Road"),
        ("Heavy-Duty Truck", 1.07, "vehicle-mile", "Mobile - On-Road"),
        ("Motorcycle", 0.12, "vehicle-mile", "Mobile - On-Road"),
        ("Bus", 0.89, "vehicle-mile", "Mobile - On-Road"),
    ]

    for row_idx, (vtype, factor, unit, category) in enumerate(mobile_data, 2):
        ws_mobile.cell(row=row_idx, column=1, value=vtype)
        ws_mobile.cell(row=row_idx, column=2, value=factor)
        ws_mobile.cell(row=row_idx, column=3, value=unit)
        ws_mobile.cell(row=row_idx, column=4, value=category)

    # Save the workbook
    wb.save(output_path)
    wb.close()

    return output_path


@pytest.fixture(scope="session")
def epa_sample_excel_bytes(epa_sample_excel_path) -> bytes:
    """
    Return the sample EPA Excel file as bytes.

    This is useful for tests that need to mock HTTP responses
    with the file content.
    """
    return epa_sample_excel_path.read_bytes()


@pytest.fixture(scope="session")
def egrid_sample_excel_path(fixtures_dir) -> Path:
    """
    Create a sample EPA eGRID Excel file for testing.

    This fixture creates an eGRID-format Excel file with subregion
    electricity emission factors. The SRCO2RTA column contains
    values in lb/MWh that need conversion to kg/kWh.

    Returns:
        Path to the created egrid_sample.xlsx file
    """
    try:
        from openpyxl import Workbook
    except ImportError:
        pytest.skip("openpyxl not available")

    output_path = fixtures_dir / "egrid_sample.xlsx"

    wb = Workbook()

    # =========================================================================
    # Sheet 1: "SUBRGN22" - Subregion emission factors
    # =========================================================================
    ws_subrgn = wb.active
    ws_subrgn.title = "SUBRGN22"

    # Header row (matches eGRID format)
    # SRCO2RTA = Subregion annual CO2 output emission rate (lb/MWh)
    subrgn_headers = ["SUBRGN", "SRCO2RTA", "SRNOXRTA", "SRSO2RTA", "SRPMRTA"]
    for col, header in enumerate(subrgn_headers, 1):
        ws_subrgn.cell(row=1, column=col, value=header)

    # Sample eGRID subregion data (all US EPA subregions)
    # SRCO2RTA values are in lb CO2/MWh
    subrgn_data = [
        ("AKGD", 1098.5, 2.3, 1.2, 0.05),   # Alaska Grid
        ("AKMS", 987.2, 1.8, 0.9, 0.04),    # Alaska Miscellaneous
        ("CAMX", 531.2, 0.8, 0.4, 0.02),    # WECC California
        ("ERCT", 925.3, 1.5, 0.9, 0.03),    # ERCOT (Texas)
        ("FRCC", 892.4, 1.2, 0.7, 0.03),    # FRCC All (Florida)
        ("MROE", 1456.8, 2.8, 2.1, 0.06),   # MRO East
        ("MROW", 1234.5, 2.4, 1.8, 0.05),   # MRO West
        ("NEWE", 456.7, 0.5, 0.3, 0.02),    # NPCC New England
        ("NWPP", 678.9, 1.0, 0.5, 0.02),    # WECC Northwest
        ("NYCW", 543.2, 0.6, 0.3, 0.02),    # NPCC NYC/Westchester
        ("NYLI", 789.3, 0.9, 0.5, 0.03),    # NPCC Long Island
        ("NYUP", 234.5, 0.3, 0.2, 0.01),    # NPCC Upstate NY
        ("RFCE", 789.3, 1.3, 0.8, 0.03),    # RFC East
        ("RFCM", 1345.6, 2.5, 1.9, 0.05),   # RFC Michigan
        ("RFCW", 1245.7, 2.1, 1.8, 0.05),   # RFC West
        ("RMPA", 1123.4, 1.9, 1.4, 0.04),   # WECC Rockies
        ("SRMV", 856.2, 1.1, 0.6, 0.03),    # SERC Mississippi Valley
        ("SRMW", 1567.8, 3.0, 2.3, 0.06),   # SERC Midwest
        ("SRSO", 1012.8, 1.9, 1.1, 0.04),   # SERC South
        ("SRTV", 934.5, 1.6, 1.0, 0.04),    # SERC Tennessee Valley
    ]

    for row_idx, row_data in enumerate(subrgn_data, 2):
        for col_idx, value in enumerate(row_data, 1):
            ws_subrgn.cell(row=row_idx, column=col_idx, value=value)

    # =========================================================================
    # Sheet 2: "US22" - US national average
    # =========================================================================
    ws_us = wb.create_sheet("US22")

    for col, header in enumerate(subrgn_headers, 1):
        ws_us.cell(row=1, column=col, value=header)

    # US national average
    ws_us.cell(row=2, column=1, value="US")
    ws_us.cell(row=2, column=2, value=857.4)  # US average lb/MWh
    ws_us.cell(row=2, column=3, value=1.2)
    ws_us.cell(row=2, column=4, value=0.8)
    ws_us.cell(row=2, column=5, value=0.03)

    # Save the workbook
    wb.save(output_path)
    wb.close()

    return output_path


@pytest.fixture(scope="session")
def egrid_sample_excel_bytes(egrid_sample_excel_path) -> bytes:
    """
    Return the sample eGRID Excel file as bytes.

    This is useful for tests that need to mock HTTP responses
    with the file content.
    """
    return egrid_sample_excel_path.read_bytes()


@pytest.fixture
def malformed_excel_bytes() -> bytes:
    """
    Create a malformed Excel file for error testing.

    This fixture creates an Excel file with wrong sheet names
    and missing data to test error handling.
    """
    try:
        from openpyxl import Workbook
    except ImportError:
        pytest.skip("openpyxl not available")

    wb = Workbook()
    ws = wb.active
    ws.title = "Wrong Sheet Name"
    # Leave sheet empty - no data

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


@pytest.fixture
def invalid_file_bytes() -> bytes:
    """Return invalid (non-Excel) bytes for error testing."""
    return b"This is not a valid Excel file content"


@pytest.fixture
def empty_file_bytes() -> bytes:
    """Return empty bytes for error testing."""
    return b""
