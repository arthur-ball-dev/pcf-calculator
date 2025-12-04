"""
DEFRA test fixtures for data ingestion tests.

TASK-DATA-P5-003: DEFRA Data Connector - Phase A Tests

This module provides functions to create DEFRA-style test data files
programmatically using openpyxl. These fixtures match the structure
of the official DEFRA Conversion Factors workbook.

Usage:
    from backend.tests.fixtures.defra_fixtures import create_defra_sample_xlsx

    xlsx_bytes = create_defra_sample_xlsx()
"""

from io import BytesIO
from typing import Optional


def create_defra_sample_xlsx() -> bytes:
    """
    Create a minimal DEFRA-style Excel workbook for testing.

    Creates a workbook with all 6 configured sheet types:
    - Fuels (Scope 1)
    - UK electricity (Scope 2)
    - Material use (Scope 3)
    - Waste disposal (Scope 3)
    - Business travel- air (Scope 3)
    - Freighting goods (Scope 3)

    Returns:
        bytes: Excel workbook content
    """
    try:
        from openpyxl import Workbook
    except ImportError:
        raise ImportError("openpyxl is required to create DEFRA fixtures")

    wb = Workbook()

    # Remove default sheet
    default_sheet = wb.active
    wb.remove(default_sheet)

    # =========================================================================
    # Sheet 1: Fuels (Scope 1 - Direct Combustion)
    # =========================================================================
    fuels = wb.create_sheet("Fuels")
    fuels.append(["Category", "Fuel", "Unit", "kg CO2e per unit"])
    fuels.append(["Gaseous fuels", "Natural Gas", "kWh", 0.18287])
    fuels.append(["Gaseous fuels", "LPG", "kWh", 0.21448])
    fuels.append(["Liquid fuels", "Diesel (average biofuel blend)", "litre", 2.70554])
    fuels.append(["Liquid fuels", "Petrol (average biofuel blend)", "litre", 2.31481])
    fuels.append(["Solid fuels", "Coal (industrial)", "kg", 2.41674])

    # =========================================================================
    # Sheet 2: UK electricity (Scope 2 - Purchased Electricity)
    # =========================================================================
    electricity = wb.create_sheet("UK electricity")
    electricity.append(["Activity", "kg CO2e per kWh"])
    electricity.append(["Electricity: UK", 0.21233])
    electricity.append(["Electricity: UK (T&D losses)", 0.01879])
    electricity.append(["Electricity: UK (generation only)", 0.19354])

    # =========================================================================
    # Sheet 3: Material use (Scope 3 - Purchased Goods)
    # =========================================================================
    materials = wb.create_sheet("Material use")
    materials.append(["Category", "Material", "kg CO2e per kg"])
    materials.append(["Metals", "Primary steel", 1.85])
    materials.append(["Metals", "Aluminium (primary)", 11.59])
    materials.append(["Plastics", "PVC (general)", 3.10])

    # =========================================================================
    # Sheet 4: Waste disposal (Scope 3 - Waste Treatment)
    # =========================================================================
    waste = wb.create_sheet("Waste disposal")
    waste.append(["Waste Category", "Waste type", "kg CO2e per tonne"])
    waste.append(["Commercial waste", "Mixed waste (closed loop)", 467.0])
    waste.append(["Construction waste", "Demolition waste", 1.0])

    # =========================================================================
    # Sheet 5: Business travel- air (Scope 3 - Business Travel)
    # =========================================================================
    travel = wb.create_sheet("Business travel- air")
    travel.append(["Class", "Type of flight", "kg CO2e per passenger km"])
    travel.append(["Economy class", "Domestic", 0.24587])
    travel.append(["Economy class", "Short-haul (to/from UK)", 0.15353])
    travel.append(["Business class", "Long-haul (to/from UK)", 0.42951])

    # =========================================================================
    # Sheet 6: Freighting goods (Scope 3 - Upstream Transportation)
    # =========================================================================
    freight = wb.create_sheet("Freighting goods")
    freight.append(["Mode", "Vehicle type", "kg CO2e per tonne.km"])
    freight.append(["Road freight", "HGV (all diesel)", 0.10691])
    freight.append(["Rail freight", "Freight train", 0.02792])
    freight.append(["Sea freight", "Container ship (average)", 0.01615])

    # Save to BytesIO and return bytes
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


def create_minimal_defra_xlsx(sheet_name: str = "Fuels") -> bytes:
    """
    Create a minimal DEFRA workbook with only one sheet.

    Useful for testing edge cases like missing sheets.

    Args:
        sheet_name: Name of the single sheet to include

    Returns:
        bytes: Excel workbook content
    """
    try:
        from openpyxl import Workbook
    except ImportError:
        raise ImportError("openpyxl is required to create DEFRA fixtures")

    wb = Workbook()
    sheet = wb.active
    sheet.title = sheet_name

    if sheet_name == "Fuels":
        sheet.append(["Fuel", "kg CO2e per unit"])
        sheet.append(["Natural Gas", 0.18287])
    elif sheet_name == "Material use":
        sheet.append(["Material", "kg CO2e per kg"])
        sheet.append(["Primary steel", 1.85])
    else:
        sheet.append(["Activity", "Value"])
        sheet.append(["Test Activity", 1.0])

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


def create_defra_xlsx_with_updated_values(value_multiplier: float = 1.0) -> bytes:
    """
    Create a DEFRA workbook with adjustable values.

    Useful for testing upsert behavior with different values.

    Args:
        value_multiplier: Multiplier to apply to all emission factors

    Returns:
        bytes: Excel workbook content
    """
    try:
        from openpyxl import Workbook
    except ImportError:
        raise ImportError("openpyxl is required to create DEFRA fixtures")

    wb = Workbook()
    default_sheet = wb.active
    wb.remove(default_sheet)

    fuels = wb.create_sheet("Fuels")
    fuels.append(["Fuel", "kg CO2e per unit"])
    fuels.append(["Natural Gas", 0.18287 * value_multiplier])
    fuels.append(["LPG", 0.21448 * value_multiplier])

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


def save_defra_sample_to_file(output_path: str) -> None:
    """
    Save the DEFRA sample workbook to a file.

    Args:
        output_path: Path where to save the xlsx file
    """
    xlsx_bytes = create_defra_sample_xlsx()
    with open(output_path, 'wb') as f:
        f.write(xlsx_bytes)


__all__ = [
    'create_defra_sample_xlsx',
    'create_minimal_defra_xlsx',
    'create_defra_xlsx_with_updated_values',
    'save_defra_sample_to_file',
]
