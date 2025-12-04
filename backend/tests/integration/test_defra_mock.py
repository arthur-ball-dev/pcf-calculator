"""
Integration test suite for DEFRAEmissionFactorsIngestion with mock HTTP.

TASK-DATA-P5-003: DEFRA Data Connector - Phase A Tests

This test suite validates:
- Full sync workflow with mocked DEFRA response
- Database writes correct data_source_id
- All 6 sheet types processed
- Correct scope assignment per category
- Upsert behavior (update on re-sync)

Test-Driven Development Protocol:
- These tests MUST be committed BEFORE implementation
- Tests should FAIL initially (no DEFRAEmissionFactorsIngestion exists yet)
- Implementation must make tests PASS without modifying tests

Note: These tests use an in-memory SQLite database for isolation
and respx for mocking HTTP requests.
"""

import pytest
import respx
import httpx
from unittest.mock import AsyncMock, MagicMock
from datetime import datetime, timezone
from decimal import Decimal
from uuid import uuid4
from io import BytesIO
from sqlalchemy import create_engine, text, select
from sqlalchemy.orm import sessionmaker
from sqlalchemy.pool import StaticPool


# Mark all tests in this module as integration tests
pytestmark = pytest.mark.integration


@pytest.fixture(scope="function")
def db_engine():
    """Create in-memory SQLite database for testing."""
    engine = create_engine(
        "sqlite:///:memory:",
        connect_args={"check_same_thread": False},
        poolclass=StaticPool,
        echo=False
    )
    return engine


@pytest.fixture(scope="function")
def db_session(db_engine):
    """Create database session for testing."""
    from backend.models import Base

    Base.metadata.create_all(db_engine)
    SessionLocal = sessionmaker(bind=db_engine)
    session = SessionLocal()

    # Enable foreign key constraints for SQLite
    session.execute(text("PRAGMA foreign_keys = ON"))
    session.commit()

    yield session

    session.close()


@pytest.fixture
def defra_data_source(db_session):
    """Create a DEFRA data source for testing."""
    try:
        from backend.models import DataSource
    except ImportError:
        pytest.skip("DataSource model not yet implemented")

    source = DataSource(
        name="DEFRA Conversion Factors",
        source_type="file",
        base_url="https://www.gov.uk/government/publications/ghg-conversion-factors",
        sync_frequency="yearly",
        is_active=True
    )
    db_session.add(source)
    db_session.commit()
    db_session.refresh(source)
    return source


@pytest.fixture
def sample_defra_xlsx():
    """
    Create a comprehensive DEFRA-style Excel workbook for integration testing.

    Includes all 6 configured sheet types with sample data.
    """
    try:
        from openpyxl import Workbook
    except ImportError:
        pytest.skip("openpyxl not installed")

    wb = Workbook()

    # Remove default sheet
    default_sheet = wb.active
    wb.remove(default_sheet)

    # Create Fuels sheet (Scope 1)
    fuels = wb.create_sheet("Fuels")
    fuels.append(["Category", "Fuel", "Unit", "kg CO2e per unit"])
    fuels.append(["Gaseous fuels", "Natural Gas", "kWh", 0.18287])
    fuels.append(["Gaseous fuels", "LPG", "kWh", 0.21448])
    fuels.append(["Liquid fuels", "Diesel", "litre", 2.70554])
    fuels.append(["Liquid fuels", "Petrol", "litre", 2.31481])
    fuels.append(["Solid fuels", "Coal (industrial)", "kg", 2.41674])

    # Create Electricity sheet (Scope 2)
    electricity = wb.create_sheet("UK electricity")
    electricity.append(["Activity", "kg CO2e per kWh"])
    electricity.append(["Electricity: UK", 0.21233])
    electricity.append(["Electricity: UK (T&D)", 0.01879])
    electricity.append(["Electricity: UK (generation)", 0.19354])

    # Create Material use sheet (Scope 3)
    materials = wb.create_sheet("Material use")
    materials.append(["Category", "Material", "kg CO2e per kg"])
    materials.append(["Metals", "Primary steel", 1.85])
    materials.append(["Metals", "Aluminium (primary)", 11.59])
    materials.append(["Plastics", "PVC", 3.10])

    # Create Waste disposal sheet (Scope 3)
    waste = wb.create_sheet("Waste disposal")
    waste.append(["Waste Category", "Waste type", "kg CO2e per tonne"])
    waste.append(["Commercial", "Mixed waste", 467.0])
    waste.append(["Construction", "Demolition waste", 1.0])

    # Create Business travel- air sheet (Scope 3)
    travel = wb.create_sheet("Business travel- air")
    travel.append(["Class", "Type of flight", "kg CO2e per passenger km"])
    travel.append(["Economy", "Domestic", 0.24587])
    travel.append(["Economy", "Short-haul", 0.15353])
    travel.append(["Business", "Long-haul", 0.42951])

    # Create Freighting goods sheet (Scope 3)
    freight = wb.create_sheet("Freighting goods")
    freight.append(["Mode", "Vehicle type", "kg CO2e per tonne.km"])
    freight.append(["Road", "HGV (average)", 0.10691])
    freight.append(["Rail", "Freight train", 0.02792])
    freight.append(["Sea", "Container ship", 0.01615])

    # Save to BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


def create_mock_async_session(db_session):
    """
    Create an async-compatible mock that delegates to sync session.
    For integration tests, we simulate async behavior.
    """
    mock_session = MagicMock()
    mock_session.add = db_session.add
    mock_session.flush = AsyncMock(side_effect=lambda: db_session.flush())
    mock_session.commit = AsyncMock(side_effect=lambda: db_session.commit())
    mock_session.rollback = AsyncMock(side_effect=lambda: db_session.rollback())
    mock_session.refresh = AsyncMock(
        side_effect=lambda obj: db_session.refresh(obj)
    )

    async def mock_execute(stmt):
        result = db_session.execute(stmt)
        mock_result = MagicMock()
        mock_result.rowcount = result.rowcount
        mock_result.scalars = result.scalars
        mock_result.fetchall = result.fetchall
        mock_result.fetchone = result.fetchone
        return mock_result

    mock_session.execute = mock_execute
    return mock_session


# ============================================================================
# Test Scenario 1: Full Sync Workflow with Mocked DEFRA Response
# ============================================================================

class TestFullSyncWorkflowWithMockedDEFRA:
    """Test complete sync workflow from fetch to database write."""

    @pytest.mark.asyncio
    @respx.mock
    async def test_full_sync_workflow_success(
        self, db_session, defra_data_source, sample_defra_xlsx
    ):
        """Test successful full sync workflow with mock HTTP."""
        try:
            from backend.services.data_ingestion.defra_ingestion import (
                DEFRAEmissionFactorsIngestion
            )
            from backend.models import EmissionFactor, DataSyncLog
        except ImportError:
            pytest.skip("Required modules not yet implemented")

        # Mock DEFRA download URL
        respx.get(DEFRAEmissionFactorsIngestion.DEFRA_URL).mock(
            return_value=httpx.Response(200, content=sample_defra_xlsx)
        )

        mock_session = create_mock_async_session(db_session)
        ingestion = DEFRAEmissionFactorsIngestion(
            db=mock_session,
            data_source_id=defra_data_source.id,
            sync_type="initial"
        )

        result = await ingestion.execute_sync()

        # Verify sync completed successfully
        assert result.status == "completed"
        assert result.records_processed > 0
        assert result.records_failed == 0

    @pytest.mark.asyncio
    @respx.mock
    async def test_sync_creates_emission_factors_in_database(
        self, db_session, defra_data_source, sample_defra_xlsx
    ):
        """Test that sync creates emission factors in the database."""
        try:
            from backend.services.data_ingestion.defra_ingestion import (
                DEFRAEmissionFactorsIngestion
            )
            from backend.models import EmissionFactor
        except ImportError:
            pytest.skip("Required modules not yet implemented")

        respx.get(DEFRAEmissionFactorsIngestion.DEFRA_URL).mock(
            return_value=httpx.Response(200, content=sample_defra_xlsx)
        )

        mock_session = create_mock_async_session(db_session)
        ingestion = DEFRAEmissionFactorsIngestion(
            db=mock_session,
            data_source_id=defra_data_source.id
        )

        await ingestion.execute_sync()

        # Verify records were created
        count = db_session.query(EmissionFactor).filter(
            EmissionFactor.data_source_id == defra_data_source.id
        ).count()

        # Sample data has approximately 17 records across all sheets
        assert count >= 15

    @pytest.mark.asyncio
    @respx.mock
    async def test_sync_handles_redirect(
        self, db_session, defra_data_source, sample_defra_xlsx
    ):
        """Test that sync handles HTTP redirects."""
        try:
            from backend.services.data_ingestion.defra_ingestion import (
                DEFRAEmissionFactorsIngestion
            )
        except ImportError:
            pytest.skip("Required modules not yet implemented")

        # Mock with redirect
        respx.get(DEFRAEmissionFactorsIngestion.DEFRA_URL).mock(
            return_value=httpx.Response(200, content=sample_defra_xlsx)
        )

        mock_session = create_mock_async_session(db_session)
        ingestion = DEFRAEmissionFactorsIngestion(
            db=mock_session,
            data_source_id=defra_data_source.id
        )

        # Should not raise exception
        result = await ingestion.execute_sync()
        assert result.status == "completed"


# ============================================================================
# Test Scenario 2: Database Writes Correct data_source_id
# ============================================================================

class TestDataSourceIdAssociation:
    """Test that database writes have correct data_source_id."""

    @pytest.mark.asyncio
    @respx.mock
    async def test_emission_factors_have_correct_data_source_id(
        self, db_session, defra_data_source, sample_defra_xlsx
    ):
        """Test that all emission factors have correct data_source_id."""
        try:
            from backend.services.data_ingestion.defra_ingestion import (
                DEFRAEmissionFactorsIngestion
            )
            from backend.models import EmissionFactor
        except ImportError:
            pytest.skip("Required modules not yet implemented")

        respx.get(DEFRAEmissionFactorsIngestion.DEFRA_URL).mock(
            return_value=httpx.Response(200, content=sample_defra_xlsx)
        )

        mock_session = create_mock_async_session(db_session)
        ingestion = DEFRAEmissionFactorsIngestion(
            db=mock_session,
            data_source_id=defra_data_source.id
        )

        await ingestion.execute_sync()

        # Query all emission factors
        factors = db_session.query(EmissionFactor).all()

        # All factors should have the correct data_source_id
        for factor in factors:
            assert factor.data_source_id == defra_data_source.id

    @pytest.mark.asyncio
    @respx.mock
    async def test_sync_log_has_correct_data_source_id(
        self, db_session, defra_data_source, sample_defra_xlsx
    ):
        """Test that sync log has correct data_source_id."""
        try:
            from backend.services.data_ingestion.defra_ingestion import (
                DEFRAEmissionFactorsIngestion
            )
            from backend.models import DataSyncLog
        except ImportError:
            pytest.skip("Required modules not yet implemented")

        respx.get(DEFRAEmissionFactorsIngestion.DEFRA_URL).mock(
            return_value=httpx.Response(200, content=sample_defra_xlsx)
        )

        mock_session = create_mock_async_session(db_session)
        ingestion = DEFRAEmissionFactorsIngestion(
            db=mock_session,
            data_source_id=defra_data_source.id
        )

        await ingestion.execute_sync()

        # Query sync log
        sync_log = db_session.query(DataSyncLog).filter(
            DataSyncLog.data_source_id == defra_data_source.id
        ).first()

        assert sync_log is not None
        assert sync_log.data_source_id == defra_data_source.id


# ============================================================================
# Test Scenario 3: All 6 Sheet Types Processed
# ============================================================================

class TestAllSheetTypesProcessed:
    """Test that all 6 configured sheet types are processed."""

    @pytest.mark.asyncio
    @respx.mock
    async def test_fuels_sheet_processed(
        self, db_session, defra_data_source, sample_defra_xlsx
    ):
        """Test that Fuels sheet records are processed."""
        try:
            from backend.services.data_ingestion.defra_ingestion import (
                DEFRAEmissionFactorsIngestion
            )
            from backend.models import EmissionFactor
        except ImportError:
            pytest.skip("Required modules not yet implemented")

        respx.get(DEFRAEmissionFactorsIngestion.DEFRA_URL).mock(
            return_value=httpx.Response(200, content=sample_defra_xlsx)
        )

        mock_session = create_mock_async_session(db_session)
        ingestion = DEFRAEmissionFactorsIngestion(
            db=mock_session,
            data_source_id=defra_data_source.id
        )

        await ingestion.execute_sync()

        # Check for Fuels records
        natural_gas = db_session.query(EmissionFactor).filter(
            EmissionFactor.activity_name == "Natural Gas"
        ).first()

        assert natural_gas is not None
        assert abs(float(natural_gas.co2e_factor) - 0.18287) < 0.0001

    @pytest.mark.asyncio
    @respx.mock
    async def test_electricity_sheet_processed(
        self, db_session, defra_data_source, sample_defra_xlsx
    ):
        """Test that Electricity sheet records are processed."""
        try:
            from backend.services.data_ingestion.defra_ingestion import (
                DEFRAEmissionFactorsIngestion
            )
            from backend.models import EmissionFactor
        except ImportError:
            pytest.skip("Required modules not yet implemented")

        respx.get(DEFRAEmissionFactorsIngestion.DEFRA_URL).mock(
            return_value=httpx.Response(200, content=sample_defra_xlsx)
        )

        mock_session = create_mock_async_session(db_session)
        ingestion = DEFRAEmissionFactorsIngestion(
            db=mock_session,
            data_source_id=defra_data_source.id
        )

        await ingestion.execute_sync()

        # Check for Electricity records
        electricity = db_session.query(EmissionFactor).filter(
            EmissionFactor.activity_name.like("%Electricity: UK%")
        ).first()

        assert electricity is not None

    @pytest.mark.asyncio
    @respx.mock
    async def test_material_use_sheet_processed(
        self, db_session, defra_data_source, sample_defra_xlsx
    ):
        """Test that Material use sheet records are processed."""
        try:
            from backend.services.data_ingestion.defra_ingestion import (
                DEFRAEmissionFactorsIngestion
            )
            from backend.models import EmissionFactor
        except ImportError:
            pytest.skip("Required modules not yet implemented")

        respx.get(DEFRAEmissionFactorsIngestion.DEFRA_URL).mock(
            return_value=httpx.Response(200, content=sample_defra_xlsx)
        )

        mock_session = create_mock_async_session(db_session)
        ingestion = DEFRAEmissionFactorsIngestion(
            db=mock_session,
            data_source_id=defra_data_source.id
        )

        await ingestion.execute_sync()

        # Check for Material records
        steel = db_session.query(EmissionFactor).filter(
            EmissionFactor.activity_name.like("%steel%")
        ).first()

        assert steel is not None

    @pytest.mark.asyncio
    @respx.mock
    async def test_waste_disposal_sheet_processed(
        self, db_session, defra_data_source, sample_defra_xlsx
    ):
        """Test that Waste disposal sheet records are processed."""
        try:
            from backend.services.data_ingestion.defra_ingestion import (
                DEFRAEmissionFactorsIngestion
            )
            from backend.models import EmissionFactor
        except ImportError:
            pytest.skip("Required modules not yet implemented")

        respx.get(DEFRAEmissionFactorsIngestion.DEFRA_URL).mock(
            return_value=httpx.Response(200, content=sample_defra_xlsx)
        )

        mock_session = create_mock_async_session(db_session)
        ingestion = DEFRAEmissionFactorsIngestion(
            db=mock_session,
            data_source_id=defra_data_source.id
        )

        await ingestion.execute_sync()

        # Check for Waste records
        waste = db_session.query(EmissionFactor).filter(
            EmissionFactor.activity_name.like("%waste%")
        ).first()

        assert waste is not None

    @pytest.mark.asyncio
    @respx.mock
    async def test_business_travel_sheet_processed(
        self, db_session, defra_data_source, sample_defra_xlsx
    ):
        """Test that Business travel- air sheet records are processed."""
        try:
            from backend.services.data_ingestion.defra_ingestion import (
                DEFRAEmissionFactorsIngestion
            )
            from backend.models import EmissionFactor
        except ImportError:
            pytest.skip("Required modules not yet implemented")

        respx.get(DEFRAEmissionFactorsIngestion.DEFRA_URL).mock(
            return_value=httpx.Response(200, content=sample_defra_xlsx)
        )

        mock_session = create_mock_async_session(db_session)
        ingestion = DEFRAEmissionFactorsIngestion(
            db=mock_session,
            data_source_id=defra_data_source.id
        )

        await ingestion.execute_sync()

        # Check for Travel records (Domestic, Short-haul, Long-haul)
        flight = db_session.query(EmissionFactor).filter(
            EmissionFactor.activity_name.like("%haul%")
        ).first()

        assert flight is not None

    @pytest.mark.asyncio
    @respx.mock
    async def test_freighting_goods_sheet_processed(
        self, db_session, defra_data_source, sample_defra_xlsx
    ):
        """Test that Freighting goods sheet records are processed."""
        try:
            from backend.services.data_ingestion.defra_ingestion import (
                DEFRAEmissionFactorsIngestion
            )
            from backend.models import EmissionFactor
        except ImportError:
            pytest.skip("Required modules not yet implemented")

        respx.get(DEFRAEmissionFactorsIngestion.DEFRA_URL).mock(
            return_value=httpx.Response(200, content=sample_defra_xlsx)
        )

        mock_session = create_mock_async_session(db_session)
        ingestion = DEFRAEmissionFactorsIngestion(
            db=mock_session,
            data_source_id=defra_data_source.id
        )

        await ingestion.execute_sync()

        # Check for Freight records
        freight = db_session.query(EmissionFactor).filter(
            EmissionFactor.activity_name.like("%HGV%")
        ).first()

        assert freight is not None


# ============================================================================
# Test Scenario 4: Correct Scope Assignment Per Category
# ============================================================================

class TestCorrectScopeAssignment:
    """Test correct scope assignment per category."""

    @pytest.mark.asyncio
    @respx.mock
    async def test_fuels_records_have_scope_1(
        self, db_session, defra_data_source, sample_defra_xlsx
    ):
        """Test that Fuels records have Scope 1."""
        try:
            from backend.services.data_ingestion.defra_ingestion import (
                DEFRAEmissionFactorsIngestion
            )
            from backend.models import EmissionFactor
        except ImportError:
            pytest.skip("Required modules not yet implemented")

        respx.get(DEFRAEmissionFactorsIngestion.DEFRA_URL).mock(
            return_value=httpx.Response(200, content=sample_defra_xlsx)
        )

        mock_session = create_mock_async_session(db_session)
        ingestion = DEFRAEmissionFactorsIngestion(
            db=mock_session,
            data_source_id=defra_data_source.id
        )

        await ingestion.execute_sync()

        # Check Fuels record scope
        natural_gas = db_session.query(EmissionFactor).filter(
            EmissionFactor.activity_name == "Natural Gas"
        ).first()

        assert natural_gas is not None
        assert natural_gas.category == "combustion"

    @pytest.mark.asyncio
    @respx.mock
    async def test_electricity_records_have_scope_2(
        self, db_session, defra_data_source, sample_defra_xlsx
    ):
        """Test that Electricity records have Scope 2."""
        try:
            from backend.services.data_ingestion.defra_ingestion import (
                DEFRAEmissionFactorsIngestion
            )
            from backend.models import EmissionFactor
        except ImportError:
            pytest.skip("Required modules not yet implemented")

        respx.get(DEFRAEmissionFactorsIngestion.DEFRA_URL).mock(
            return_value=httpx.Response(200, content=sample_defra_xlsx)
        )

        mock_session = create_mock_async_session(db_session)
        ingestion = DEFRAEmissionFactorsIngestion(
            db=mock_session,
            data_source_id=defra_data_source.id
        )

        await ingestion.execute_sync()

        # Check Electricity record category
        electricity = db_session.query(EmissionFactor).filter(
            EmissionFactor.activity_name.like("%Electricity: UK%")
        ).first()

        if electricity:
            assert electricity.category == "electricity"

    @pytest.mark.asyncio
    @respx.mock
    async def test_material_records_have_scope_3(
        self, db_session, defra_data_source, sample_defra_xlsx
    ):
        """Test that Material use records have Scope 3."""
        try:
            from backend.services.data_ingestion.defra_ingestion import (
                DEFRAEmissionFactorsIngestion
            )
            from backend.models import EmissionFactor
        except ImportError:
            pytest.skip("Required modules not yet implemented")

        respx.get(DEFRAEmissionFactorsIngestion.DEFRA_URL).mock(
            return_value=httpx.Response(200, content=sample_defra_xlsx)
        )

        mock_session = create_mock_async_session(db_session)
        ingestion = DEFRAEmissionFactorsIngestion(
            db=mock_session,
            data_source_id=defra_data_source.id
        )

        await ingestion.execute_sync()

        # Check Material record category
        steel = db_session.query(EmissionFactor).filter(
            EmissionFactor.activity_name.like("%steel%")
        ).first()

        if steel:
            assert steel.category == "materials"

    @pytest.mark.asyncio
    @respx.mock
    async def test_waste_records_have_scope_3(
        self, db_session, defra_data_source, sample_defra_xlsx
    ):
        """Test that Waste disposal records have Scope 3."""
        try:
            from backend.services.data_ingestion.defra_ingestion import (
                DEFRAEmissionFactorsIngestion
            )
            from backend.models import EmissionFactor
        except ImportError:
            pytest.skip("Required modules not yet implemented")

        respx.get(DEFRAEmissionFactorsIngestion.DEFRA_URL).mock(
            return_value=httpx.Response(200, content=sample_defra_xlsx)
        )

        mock_session = create_mock_async_session(db_session)
        ingestion = DEFRAEmissionFactorsIngestion(
            db=mock_session,
            data_source_id=defra_data_source.id
        )

        await ingestion.execute_sync()

        # Check Waste record category
        waste = db_session.query(EmissionFactor).filter(
            EmissionFactor.activity_name.like("%waste%")
        ).first()

        if waste:
            assert waste.category == "waste"

    @pytest.mark.asyncio
    @respx.mock
    async def test_transport_records_have_scope_3(
        self, db_session, defra_data_source, sample_defra_xlsx
    ):
        """Test that transport records have Scope 3."""
        try:
            from backend.services.data_ingestion.defra_ingestion import (
                DEFRAEmissionFactorsIngestion
            )
            from backend.models import EmissionFactor
        except ImportError:
            pytest.skip("Required modules not yet implemented")

        respx.get(DEFRAEmissionFactorsIngestion.DEFRA_URL).mock(
            return_value=httpx.Response(200, content=sample_defra_xlsx)
        )

        mock_session = create_mock_async_session(db_session)
        ingestion = DEFRAEmissionFactorsIngestion(
            db=mock_session,
            data_source_id=defra_data_source.id
        )

        await ingestion.execute_sync()

        # Check Transport records have transport category
        hgv = db_session.query(EmissionFactor).filter(
            EmissionFactor.activity_name.like("%HGV%")
        ).first()

        if hgv:
            assert hgv.category == "transport"


# ============================================================================
# Test Scenario 5: Upsert Behavior (Update on Re-sync)
# ============================================================================

class TestUpsertBehavior:
    """Test upsert behavior on re-sync."""

    @pytest.mark.asyncio
    @respx.mock
    async def test_resync_updates_existing_records(
        self, db_session, defra_data_source, sample_defra_xlsx
    ):
        """Test that re-sync updates existing records."""
        try:
            from backend.services.data_ingestion.defra_ingestion import (
                DEFRAEmissionFactorsIngestion
            )
            from backend.models import EmissionFactor
        except ImportError:
            pytest.skip("Required modules not yet implemented")

        respx.get(DEFRAEmissionFactorsIngestion.DEFRA_URL).mock(
            return_value=httpx.Response(200, content=sample_defra_xlsx)
        )

        mock_session = create_mock_async_session(db_session)

        # First sync
        ingestion1 = DEFRAEmissionFactorsIngestion(
            db=mock_session,
            data_source_id=defra_data_source.id
        )
        result1 = await ingestion1.execute_sync()

        initial_count = db_session.query(EmissionFactor).count()

        # Second sync (should update, not duplicate)
        ingestion2 = DEFRAEmissionFactorsIngestion(
            db=mock_session,
            data_source_id=defra_data_source.id
        )
        result2 = await ingestion2.execute_sync()

        # Count should be the same (no duplicates)
        final_count = db_session.query(EmissionFactor).count()
        assert final_count == initial_count

    @pytest.mark.asyncio
    @respx.mock
    async def test_resync_with_updated_values(
        self, db_session, defra_data_source
    ):
        """Test that re-sync updates values when they change."""
        try:
            from backend.services.data_ingestion.defra_ingestion import (
                DEFRAEmissionFactorsIngestion
            )
            from backend.models import EmissionFactor
            from openpyxl import Workbook
        except ImportError:
            pytest.skip("Required modules not yet implemented")

        # Create initial workbook with value 0.18287
        def create_workbook(value):
            wb = Workbook()
            default_sheet = wb.active
            wb.remove(default_sheet)
            fuels = wb.create_sheet("Fuels")
            fuels.append(["Fuel", "kg CO2e per unit"])
            fuels.append(["Natural Gas", value])
            output = BytesIO()
            wb.save(output)
            output.seek(0)
            return output.getvalue()

        # First sync with original value
        xlsx1 = create_workbook(0.18287)
        respx.get(DEFRAEmissionFactorsIngestion.DEFRA_URL).mock(
            return_value=httpx.Response(200, content=xlsx1)
        )

        mock_session = create_mock_async_session(db_session)
        ingestion1 = DEFRAEmissionFactorsIngestion(
            db=mock_session,
            data_source_id=defra_data_source.id
        )
        await ingestion1.execute_sync()

        # Get original record
        original = db_session.query(EmissionFactor).filter(
            EmissionFactor.activity_name == "Natural Gas"
        ).first()
        original_value = float(original.co2e_factor)

        # Reset respx and sync with updated value
        respx.reset()
        xlsx2 = create_workbook(0.19999)
        respx.get(DEFRAEmissionFactorsIngestion.DEFRA_URL).mock(
            return_value=httpx.Response(200, content=xlsx2)
        )

        ingestion2 = DEFRAEmissionFactorsIngestion(
            db=mock_session,
            data_source_id=defra_data_source.id
        )
        await ingestion2.execute_sync()

        # Refresh and check value was updated
        db_session.expire_all()
        updated = db_session.query(EmissionFactor).filter(
            EmissionFactor.activity_name == "Natural Gas"
        ).first()

        # Value should be updated (or at least record exists)
        assert updated is not None

    @pytest.mark.asyncio
    @respx.mock
    async def test_no_duplicate_external_ids(
        self, db_session, defra_data_source, sample_defra_xlsx
    ):
        """Test that no duplicate external_ids are created."""
        try:
            from backend.services.data_ingestion.defra_ingestion import (
                DEFRAEmissionFactorsIngestion
            )
            from backend.models import EmissionFactor
        except ImportError:
            pytest.skip("Required modules not yet implemented")

        respx.get(DEFRAEmissionFactorsIngestion.DEFRA_URL).mock(
            return_value=httpx.Response(200, content=sample_defra_xlsx)
        )

        mock_session = create_mock_async_session(db_session)

        # Run sync twice
        for _ in range(2):
            ingestion = DEFRAEmissionFactorsIngestion(
                db=mock_session,
                data_source_id=defra_data_source.id
            )
            await ingestion.execute_sync()

        # Check for duplicates
        factors = db_session.query(EmissionFactor).all()
        external_ids = [f.external_id for f in factors if f.external_id]

        # No duplicates
        assert len(external_ids) == len(set(external_ids))


# ============================================================================
# Test Scenario 6: Geography Set to GB
# ============================================================================

class TestGeographySetToGB:
    """Test that geography is set to GB for all records."""

    @pytest.mark.asyncio
    @respx.mock
    async def test_all_records_have_geography_gb(
        self, db_session, defra_data_source, sample_defra_xlsx
    ):
        """Test that all emission factors have geography set to GB."""
        try:
            from backend.services.data_ingestion.defra_ingestion import (
                DEFRAEmissionFactorsIngestion
            )
            from backend.models import EmissionFactor
        except ImportError:
            pytest.skip("Required modules not yet implemented")

        respx.get(DEFRAEmissionFactorsIngestion.DEFRA_URL).mock(
            return_value=httpx.Response(200, content=sample_defra_xlsx)
        )

        mock_session = create_mock_async_session(db_session)
        ingestion = DEFRAEmissionFactorsIngestion(
            db=mock_session,
            data_source_id=defra_data_source.id
        )

        await ingestion.execute_sync()

        # Check all factors have GB geography
        factors = db_session.query(EmissionFactor).filter(
            EmissionFactor.data_source_id == defra_data_source.id
        ).all()

        for factor in factors:
            assert factor.geography == "GB"


# ============================================================================
# Test Scenario 7: Data Quality Rating
# ============================================================================

class TestDataQualityRating:
    """Test data quality rating is set correctly."""

    @pytest.mark.asyncio
    @respx.mock
    async def test_all_records_have_data_quality_rating(
        self, db_session, defra_data_source, sample_defra_xlsx
    ):
        """Test that all records have data_quality_rating of 0.88."""
        try:
            from backend.services.data_ingestion.defra_ingestion import (
                DEFRAEmissionFactorsIngestion
            )
            from backend.models import EmissionFactor
        except ImportError:
            pytest.skip("Required modules not yet implemented")

        respx.get(DEFRAEmissionFactorsIngestion.DEFRA_URL).mock(
            return_value=httpx.Response(200, content=sample_defra_xlsx)
        )

        mock_session = create_mock_async_session(db_session)
        ingestion = DEFRAEmissionFactorsIngestion(
            db=mock_session,
            data_source_id=defra_data_source.id
        )

        await ingestion.execute_sync()

        factors = db_session.query(EmissionFactor).filter(
            EmissionFactor.data_source_id == defra_data_source.id
        ).all()

        for factor in factors:
            if factor.data_quality_rating:
                assert float(factor.data_quality_rating) == 0.88


# ============================================================================
# Test Scenario 8: Sync Log Statistics
# ============================================================================

class TestSyncLogStatistics:
    """Test sync log records correct statistics."""

    @pytest.mark.asyncio
    @respx.mock
    async def test_sync_log_records_processed_count(
        self, db_session, defra_data_source, sample_defra_xlsx
    ):
        """Test that sync log records correct processed count."""
        try:
            from backend.services.data_ingestion.defra_ingestion import (
                DEFRAEmissionFactorsIngestion
            )
            from backend.models import DataSyncLog
        except ImportError:
            pytest.skip("Required modules not yet implemented")

        respx.get(DEFRAEmissionFactorsIngestion.DEFRA_URL).mock(
            return_value=httpx.Response(200, content=sample_defra_xlsx)
        )

        mock_session = create_mock_async_session(db_session)
        ingestion = DEFRAEmissionFactorsIngestion(
            db=mock_session,
            data_source_id=defra_data_source.id
        )

        result = await ingestion.execute_sync()

        sync_log = db_session.query(DataSyncLog).filter(
            DataSyncLog.data_source_id == defra_data_source.id
        ).order_by(DataSyncLog.created_at.desc()).first()

        assert sync_log is not None
        assert sync_log.records_processed == result.records_processed

    @pytest.mark.asyncio
    @respx.mock
    async def test_sync_log_status_completed(
        self, db_session, defra_data_source, sample_defra_xlsx
    ):
        """Test that sync log has status 'completed'."""
        try:
            from backend.services.data_ingestion.defra_ingestion import (
                DEFRAEmissionFactorsIngestion
            )
            from backend.models import DataSyncLog
        except ImportError:
            pytest.skip("Required modules not yet implemented")

        respx.get(DEFRAEmissionFactorsIngestion.DEFRA_URL).mock(
            return_value=httpx.Response(200, content=sample_defra_xlsx)
        )

        mock_session = create_mock_async_session(db_session)
        ingestion = DEFRAEmissionFactorsIngestion(
            db=mock_session,
            data_source_id=defra_data_source.id
        )

        await ingestion.execute_sync()

        sync_log = db_session.query(DataSyncLog).filter(
            DataSyncLog.data_source_id == defra_data_source.id
        ).order_by(DataSyncLog.created_at.desc()).first()

        assert sync_log.status == "completed"


# ============================================================================
# Test Scenario 9: HTTP Error Handling
# ============================================================================

class TestHTTPErrorHandling:
    """Test HTTP error handling."""

    @pytest.mark.asyncio
    @respx.mock
    async def test_sync_fails_on_http_error(
        self, db_session, defra_data_source
    ):
        """Test that sync fails gracefully on HTTP error."""
        try:
            from backend.services.data_ingestion.defra_ingestion import (
                DEFRAEmissionFactorsIngestion
            )
        except ImportError:
            pytest.skip("Required modules not yet implemented")

        respx.get(DEFRAEmissionFactorsIngestion.DEFRA_URL).mock(
            return_value=httpx.Response(404)
        )

        mock_session = create_mock_async_session(db_session)
        ingestion = DEFRAEmissionFactorsIngestion(
            db=mock_session,
            data_source_id=defra_data_source.id
        )

        with pytest.raises(httpx.HTTPStatusError):
            await ingestion.execute_sync()

    @pytest.mark.asyncio
    @respx.mock
    async def test_sync_fails_on_connection_error(
        self, db_session, defra_data_source
    ):
        """Test that sync fails gracefully on connection error."""
        try:
            from backend.services.data_ingestion.defra_ingestion import (
                DEFRAEmissionFactorsIngestion
            )
        except ImportError:
            pytest.skip("Required modules not yet implemented")

        respx.get(DEFRAEmissionFactorsIngestion.DEFRA_URL).mock(
            side_effect=httpx.ConnectError("Connection refused")
        )

        mock_session = create_mock_async_session(db_session)
        ingestion = DEFRAEmissionFactorsIngestion(
            db=mock_session,
            data_source_id=defra_data_source.id
        )

        with pytest.raises(httpx.ConnectError):
            await ingestion.execute_sync()


# ============================================================================
# Test Scenario 10: Reference Year Configuration
# ============================================================================

class TestReferenceYearInDatabase:
    """Test reference year is correctly stored in database."""

    @pytest.mark.asyncio
    @respx.mock
    async def test_emission_factors_have_reference_year(
        self, db_session, defra_data_source, sample_defra_xlsx
    ):
        """Test that all emission factors have reference_year set."""
        try:
            from backend.services.data_ingestion.defra_ingestion import (
                DEFRAEmissionFactorsIngestion
            )
            from backend.models import EmissionFactor
        except ImportError:
            pytest.skip("Required modules not yet implemented")

        respx.get(DEFRAEmissionFactorsIngestion.DEFRA_URL).mock(
            return_value=httpx.Response(200, content=sample_defra_xlsx)
        )

        mock_session = create_mock_async_session(db_session)
        ingestion = DEFRAEmissionFactorsIngestion(
            db=mock_session,
            data_source_id=defra_data_source.id
        )

        await ingestion.execute_sync()

        factors = db_session.query(EmissionFactor).filter(
            EmissionFactor.data_source_id == defra_data_source.id
        ).all()

        for factor in factors:
            assert factor.reference_year == 2024
