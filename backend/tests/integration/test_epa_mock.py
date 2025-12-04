"""
Integration test suite for EPAEmissionFactorsIngestion with mock HTTP.

TASK-DATA-P5-002: EPA Data Connector - Phase A Tests

This test suite validates:
- Full sync workflow with mocked EPA HTTP response
- Database writes correct data_source_id
- Upsert behavior (update on re-sync)
- Sync log lifecycle (in_progress -> completed)

Test-Driven Development Protocol:
- These tests MUST be committed BEFORE implementation
- Tests should FAIL initially (no EPAEmissionFactorsIngestion class exists yet)
- Implementation must make tests PASS without modifying tests

Note: These tests use an in-memory SQLite database for isolation
and respx/httpx for mocking HTTP requests to EPA endpoints.
"""

import pytest
import respx
import httpx
from unittest.mock import AsyncMock, MagicMock
from datetime import datetime, timezone
from decimal import Decimal
from uuid import uuid4
from io import BytesIO

from sqlalchemy import create_engine, text
from sqlalchemy.orm import sessionmaker
from sqlalchemy.pool import StaticPool


# Mark all tests in this module as integration tests
pytestmark = pytest.mark.integration


@pytest.fixture(scope="function")
def db_engine():
    """Create in-memory SQLite database for testing."""
    engine = create_engine(
        "sqlite:///:memory:",
        connect_args={"check_same_thread": False},
        poolclass=StaticPool,
        echo=False
    )
    return engine


@pytest.fixture(scope="function")
def db_session(db_engine):
    """Create database session for testing."""
    from backend.models import Base

    Base.metadata.create_all(db_engine)
    SessionLocal = sessionmaker(bind=db_engine)
    session = SessionLocal()

    # Enable foreign key constraints for SQLite
    session.execute(text("PRAGMA foreign_keys = ON"))
    session.commit()

    yield session

    session.close()


@pytest.fixture
def data_source(db_session):
    """Create a test EPA data source."""
    try:
        from backend.models import DataSource
    except ImportError:
        pytest.skip("DataSource model not yet implemented")

    source = DataSource(
        name="EPA GHG Emission Factors Hub",
        source_type="file",
        base_url="https://www.epa.gov/climateleadership/ghg-emission-factors-hub",
        sync_frequency="biweekly",
        is_active=True
    )
    db_session.add(source)
    db_session.commit()
    db_session.refresh(source)
    return source


@pytest.fixture
def sample_epa_fuel_excel():
    """
    Create a sample EPA fuel emission factors Excel file.

    This fixture creates a valid Excel file in memory that mimics
    the structure of the actual EPA emission factors spreadsheet.
    """
    try:
        from openpyxl import Workbook
    except ImportError:
        pytest.skip("openpyxl not available")

    wb = Workbook()

    # Create "Table 1 - Fuel" sheet
    ws = wb.active
    ws.title = "Table 1 - Fuel"

    # Header row (mimics EPA format)
    headers = ["Fuel Type", "kg CO2e per unit", "Unit", "Category"]
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)

    # Sample fuel data (representative of real EPA data)
    fuel_data = [
        ("Natural Gas", 2.75, "kg", "Stationary Combustion"),
        ("Diesel", 10.21, "L", "Stationary Combustion"),
        ("Gasoline", 8.89, "L", "Mobile"),
        ("Propane", 6.35, "kg", "Stationary Combustion"),
        ("Residual Fuel Oil", 11.27, "L", "Stationary Combustion"),
        ("Coal (Bituminous)", 2.563, "kg", "Stationary Combustion"),
        ("Kerosene", 9.75, "L", "Stationary Combustion"),
        ("Aviation Gasoline", 8.31, "L", "Mobile - Aviation"),
        ("Jet Fuel", 9.57, "L", "Mobile - Aviation"),
        ("LPG (Liquefied Petroleum Gas)", 5.68, "L", "Stationary Combustion"),
    ]

    for row_idx, (fuel_type, factor, unit, category) in enumerate(fuel_data, 2):
        ws.cell(row=row_idx, column=1, value=fuel_type)
        ws.cell(row=row_idx, column=2, value=factor)
        ws.cell(row=row_idx, column=3, value=unit)
        ws.cell(row=row_idx, column=4, value=category)

    # Create "Table 2 - Mobile" sheet (additional data source)
    ws2 = wb.create_sheet("Table 2 - Mobile")
    mobile_headers = ["Vehicle Type", "CO2e Factor", "Unit", "Category"]
    for col, header in enumerate(mobile_headers, 1):
        ws2.cell(row=1, column=col, value=header)

    mobile_data = [
        ("Passenger Car", 0.21, "vehicle-mile", "Mobile - On-Road"),
        ("Light-Duty Truck", 0.29, "vehicle-mile", "Mobile - On-Road"),
        ("Heavy-Duty Truck", 1.07, "vehicle-mile", "Mobile - On-Road"),
    ]

    for row_idx, (vtype, factor, unit, category) in enumerate(mobile_data, 2):
        ws2.cell(row=row_idx, column=1, value=vtype)
        ws2.cell(row=row_idx, column=2, value=factor)
        ws2.cell(row=row_idx, column=3, value=unit)
        ws2.cell(row=row_idx, column=4, value=category)

    # Save to bytes
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


@pytest.fixture
def sample_epa_egrid_excel():
    """
    Create a sample EPA eGRID Excel file.

    This fixture creates a valid Excel file that mimics the structure
    of the actual EPA eGRID data spreadsheet with subregion data.
    """
    try:
        from openpyxl import Workbook
    except ImportError:
        pytest.skip("openpyxl not available")

    wb = Workbook()

    # Create "SUBRGN22" sheet (subregion data)
    ws = wb.active
    ws.title = "SUBRGN22"

    # Header row - eGRID uses specific column names
    # SRCO2RTA = Subregion CO2 Rate (lb/MWh)
    headers = ["SUBRGN", "SRCO2RTA", "SRNOXRTA", "SRSO2RTA", "SRPMRTA"]
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)

    # Sample eGRID subregion data (SRCO2RTA is in lb/MWh)
    egrid_data = [
        ("AKGD", 1098.5, 2.3, 1.2, 0.05),   # Alaska Grid
        ("AKMS", 987.2, 1.8, 0.9, 0.04),    # Alaska Miscellaneous
        ("CAMX", 531.2, 0.8, 0.4, 0.02),    # WECC California
        ("ERCT", 925.3, 1.5, 0.9, 0.03),    # ERCOT
        ("FRCC", 892.4, 1.2, 0.7, 0.03),    # FRCC All
        ("MROE", 1456.8, 2.8, 2.1, 0.06),   # MRO East
        ("MROW", 1234.5, 2.4, 1.8, 0.05),   # MRO West
        ("NEWE", 456.7, 0.5, 0.3, 0.02),    # NPCC New England
        ("NWPP", 678.9, 1.0, 0.5, 0.02),    # WECC Northwest
        ("RFCE", 789.3, 1.3, 0.8, 0.03),    # RFC East
        ("RFCM", 1345.6, 2.5, 1.9, 0.05),   # RFC Michigan
        ("RFCW", 1245.7, 2.1, 1.8, 0.05),   # RFC West
        ("RMPA", 1123.4, 1.9, 1.4, 0.04),   # WECC Rockies
        ("SRMV", 856.2, 1.1, 0.6, 0.03),    # SERC Mississippi Valley
        ("SRMW", 1567.8, 3.0, 2.3, 0.06),   # SERC Midwest
        ("SRSO", 1012.8, 1.9, 1.1, 0.04),   # SERC South
        ("SRTV", 934.5, 1.6, 1.0, 0.04),    # SERC Tennessee Valley
        ("SRVC", 823.1, 1.4, 0.9, 0.03),    # SERC Virginia/Carolina
        ("SPNO", 1456.9, 2.7, 2.0, 0.05),   # SPP North
        ("SPSO", 1289.4, 2.2, 1.6, 0.05),   # SPP South
    ]

    for row_idx, row_data in enumerate(egrid_data, 2):
        for col_idx, value in enumerate(row_data, 1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    # Create "US22" sheet (US average data)
    ws2 = wb.create_sheet("US22")
    for col, header in enumerate(headers, 1):
        ws2.cell(row=1, column=col, value=header)
    ws2.cell(row=2, column=1, value="US")
    ws2.cell(row=2, column=2, value=857.4)  # US average
    ws2.cell(row=2, column=3, value=1.2)
    ws2.cell(row=2, column=4, value=0.8)
    ws2.cell(row=2, column=5, value=0.03)

    # Save to bytes
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


def create_mock_async_session(db_session):
    """
    Create an async-compatible mock that delegates to sync session.
    For integration tests, we simulate async behavior.
    """
    mock_session = MagicMock()
    mock_session.add = db_session.add
    mock_session.flush = AsyncMock(side_effect=lambda: db_session.flush())
    mock_session.commit = AsyncMock(side_effect=lambda: db_session.commit())
    mock_session.rollback = AsyncMock(side_effect=lambda: db_session.rollback())
    mock_session.refresh = AsyncMock(side_effect=lambda obj: db_session.refresh(obj))

    async def mock_execute(stmt):
        result = db_session.execute(stmt)
        mock_result = MagicMock()
        mock_result.rowcount = result.rowcount
        mock_result.scalars = result.scalars
        mock_result.fetchall = result.fetchall
        mock_result.fetchone = result.fetchone
        return mock_result

    mock_session.execute = mock_execute
    return mock_session


# ============================================================================
# Test Scenario 1: Full Sync Workflow with Mocked EPA Response
# ============================================================================

class TestFullSyncWorkflow:
    """Test complete EPA sync workflow from fetch to database write."""

    @pytest.mark.asyncio
    @respx.mock
    async def test_full_sync_workflow_with_fuel_data(
        self, db_session, data_source, sample_epa_fuel_excel
    ):
        """Test successful full sync workflow with mock EPA fuel data."""
        try:
            from backend.services.data_ingestion.epa_ingestion import (
                EPAEmissionFactorsIngestion
            )
            from backend.models import EmissionFactor, DataSyncLog
        except ImportError:
            pytest.skip("Required modules not yet implemented")

        # Get the URL from the ingestion class
        ingestion_temp = EPAEmissionFactorsIngestion(
            db=MagicMock(),
            data_source_id=data_source.id,
            file_key="fuels"
        )
        mock_url = ingestion_temp.file_config["url"]

        # Mock HTTP response with sample EPA file
        respx.get(mock_url).mock(
            return_value=httpx.Response(200, content=sample_epa_fuel_excel)
        )

        mock_session = create_mock_async_session(db_session)
        ingestion = EPAEmissionFactorsIngestion(
            db=mock_session,
            data_source_id=data_source.id,
            file_key="fuels",
            sync_type="initial"
        )

        result = await ingestion.execute_sync()

        # Verify sync completed successfully
        assert result.status == "completed"
        assert result.records_processed >= 10  # At least 10 fuel records
        assert result.records_created >= 10
        assert result.records_failed == 0

    @pytest.mark.asyncio
    @respx.mock
    async def test_full_sync_workflow_with_egrid_data(
        self, db_session, data_source, sample_epa_egrid_excel
    ):
        """Test successful full sync workflow with mock EPA eGRID data."""
        try:
            from backend.services.data_ingestion.epa_ingestion import (
                EPAEmissionFactorsIngestion
            )
            from backend.models import EmissionFactor
        except ImportError:
            pytest.skip("Required modules not yet implemented")

        # Get the URL from the ingestion class
        ingestion_temp = EPAEmissionFactorsIngestion(
            db=MagicMock(),
            data_source_id=data_source.id,
            file_key="egrid"
        )
        mock_url = ingestion_temp.file_config["url"]

        # Mock HTTP response with sample eGRID file
        respx.get(mock_url).mock(
            return_value=httpx.Response(200, content=sample_epa_egrid_excel)
        )

        mock_session = create_mock_async_session(db_session)
        ingestion = EPAEmissionFactorsIngestion(
            db=mock_session,
            data_source_id=data_source.id,
            file_key="egrid",
            sync_type="initial"
        )

        result = await ingestion.execute_sync()

        # Verify sync completed successfully
        assert result.status == "completed"
        assert result.records_processed >= 20  # 20+ subregions
        assert result.records_created >= 20
        assert result.records_failed == 0

    @pytest.mark.asyncio
    @respx.mock
    async def test_sync_handles_http_error_gracefully(
        self, db_session, data_source
    ):
        """Test sync handles HTTP errors gracefully."""
        try:
            from backend.services.data_ingestion.epa_ingestion import (
                EPAEmissionFactorsIngestion
            )
        except ImportError:
            pytest.skip("Required modules not yet implemented")

        # Get the URL from the ingestion class
        ingestion_temp = EPAEmissionFactorsIngestion(
            db=MagicMock(),
            data_source_id=data_source.id,
            file_key="fuels"
        )
        mock_url = ingestion_temp.file_config["url"]

        # Mock HTTP 404 error
        respx.get(mock_url).mock(
            return_value=httpx.Response(404)
        )

        mock_session = create_mock_async_session(db_session)
        ingestion = EPAEmissionFactorsIngestion(
            db=mock_session,
            data_source_id=data_source.id,
            file_key="fuels"
        )

        # Should raise an HTTP error
        with pytest.raises(httpx.HTTPStatusError):
            await ingestion.execute_sync()

    @pytest.mark.asyncio
    @respx.mock
    async def test_sync_handles_network_timeout(
        self, db_session, data_source
    ):
        """Test sync handles network timeout gracefully."""
        try:
            from backend.services.data_ingestion.epa_ingestion import (
                EPAEmissionFactorsIngestion
            )
        except ImportError:
            pytest.skip("Required modules not yet implemented")

        # Get the URL from the ingestion class
        ingestion_temp = EPAEmissionFactorsIngestion(
            db=MagicMock(),
            data_source_id=data_source.id,
            file_key="fuels"
        )
        mock_url = ingestion_temp.file_config["url"]

        # Mock network timeout
        respx.get(mock_url).mock(side_effect=httpx.TimeoutException("Timeout"))

        mock_session = create_mock_async_session(db_session)
        ingestion = EPAEmissionFactorsIngestion(
            db=mock_session,
            data_source_id=data_source.id,
            file_key="fuels"
        )

        # Should raise a timeout exception
        with pytest.raises(httpx.TimeoutException):
            await ingestion.execute_sync()


# ============================================================================
# Test Scenario 2: Database Writes Correct data_source_id
# ============================================================================

class TestDatabaseWrites:
    """Test that database writes have correct data_source_id."""

    @pytest.mark.asyncio
    @respx.mock
    async def test_emission_factors_have_correct_data_source_id(
        self, db_session, data_source, sample_epa_fuel_excel
    ):
        """Test that emission factors are linked to correct data source."""
        try:
            from backend.services.data_ingestion.epa_ingestion import (
                EPAEmissionFactorsIngestion
            )
            from backend.models import EmissionFactor
        except ImportError:
            pytest.skip("Required modules not yet implemented")

        # Get the URL from the ingestion class
        ingestion_temp = EPAEmissionFactorsIngestion(
            db=MagicMock(),
            data_source_id=data_source.id,
            file_key="fuels"
        )
        mock_url = ingestion_temp.file_config["url"]

        respx.get(mock_url).mock(
            return_value=httpx.Response(200, content=sample_epa_fuel_excel)
        )

        mock_session = create_mock_async_session(db_session)
        ingestion = EPAEmissionFactorsIngestion(
            db=mock_session,
            data_source_id=data_source.id,
            file_key="fuels"
        )

        await ingestion.execute_sync()

        # Verify all emission factors have correct data_source_id
        factors = db_session.query(EmissionFactor).filter(
            EmissionFactor.data_source_id == data_source.id
        ).all()

        assert len(factors) >= 10  # At least 10 fuel records
        for factor in factors:
            assert factor.data_source_id == data_source.id

    @pytest.mark.asyncio
    @respx.mock
    async def test_emission_factors_have_external_id(
        self, db_session, data_source, sample_epa_fuel_excel
    ):
        """Test that emission factors have external_id set."""
        try:
            from backend.services.data_ingestion.epa_ingestion import (
                EPAEmissionFactorsIngestion
            )
            from backend.models import EmissionFactor
        except ImportError:
            pytest.skip("Required modules not yet implemented")

        ingestion_temp = EPAEmissionFactorsIngestion(
            db=MagicMock(),
            data_source_id=data_source.id,
            file_key="fuels"
        )
        mock_url = ingestion_temp.file_config["url"]

        respx.get(mock_url).mock(
            return_value=httpx.Response(200, content=sample_epa_fuel_excel)
        )

        mock_session = create_mock_async_session(db_session)
        ingestion = EPAEmissionFactorsIngestion(
            db=mock_session,
            data_source_id=data_source.id,
            file_key="fuels"
        )

        await ingestion.execute_sync()

        # Verify all emission factors have external_id
        factors = db_session.query(EmissionFactor).filter(
            EmissionFactor.data_source_id == data_source.id
        ).all()

        for factor in factors:
            assert factor.external_id is not None
            assert "EPA" in factor.external_id

    @pytest.mark.asyncio
    @respx.mock
    async def test_egrid_factors_have_correct_unit(
        self, db_session, data_source, sample_epa_egrid_excel
    ):
        """Test that eGRID factors have correct converted unit."""
        try:
            from backend.services.data_ingestion.epa_ingestion import (
                EPAEmissionFactorsIngestion
            )
            from backend.models import EmissionFactor
        except ImportError:
            pytest.skip("Required modules not yet implemented")

        ingestion_temp = EPAEmissionFactorsIngestion(
            db=MagicMock(),
            data_source_id=data_source.id,
            file_key="egrid"
        )
        mock_url = ingestion_temp.file_config["url"]

        respx.get(mock_url).mock(
            return_value=httpx.Response(200, content=sample_epa_egrid_excel)
        )

        mock_session = create_mock_async_session(db_session)
        ingestion = EPAEmissionFactorsIngestion(
            db=mock_session,
            data_source_id=data_source.id,
            file_key="egrid"
        )

        await ingestion.execute_sync()

        # Verify eGRID factors have correct unit (kg CO2e/kWh)
        factors = db_session.query(EmissionFactor).filter(
            EmissionFactor.data_source_id == data_source.id
        ).all()

        for factor in factors:
            assert factor.unit == "kg CO2e/kWh"


# ============================================================================
# Test Scenario 3: Upsert Behavior (Update on Re-Sync)
# ============================================================================

class TestUpsertBehavior:
    """Test upsert behavior for updating existing records."""

    @pytest.mark.asyncio
    @respx.mock
    async def test_resync_updates_existing_records(
        self, db_session, data_source, sample_epa_fuel_excel
    ):
        """Test that re-sync updates existing records."""
        try:
            from backend.services.data_ingestion.epa_ingestion import (
                EPAEmissionFactorsIngestion
            )
            from backend.models import EmissionFactor
        except ImportError:
            pytest.skip("Required modules not yet implemented")

        ingestion_temp = EPAEmissionFactorsIngestion(
            db=MagicMock(),
            data_source_id=data_source.id,
            file_key="fuels"
        )
        mock_url = ingestion_temp.file_config["url"]

        respx.get(mock_url).mock(
            return_value=httpx.Response(200, content=sample_epa_fuel_excel)
        )

        mock_session = create_mock_async_session(db_session)

        # First sync
        ingestion1 = EPAEmissionFactorsIngestion(
            db=mock_session,
            data_source_id=data_source.id,
            file_key="fuels"
        )
        result1 = await ingestion1.execute_sync()

        initial_count = db_session.query(EmissionFactor).filter(
            EmissionFactor.data_source_id == data_source.id
        ).count()

        # Re-sync with same data
        ingestion2 = EPAEmissionFactorsIngestion(
            db=mock_session,
            data_source_id=data_source.id,
            file_key="fuels"
        )
        result2 = await ingestion2.execute_sync()

        final_count = db_session.query(EmissionFactor).filter(
            EmissionFactor.data_source_id == data_source.id
        ).count()

        # Should have same number of records (no duplicates)
        assert final_count == initial_count

    @pytest.mark.asyncio
    @respx.mock
    async def test_resync_does_not_create_duplicates(
        self, db_session, data_source, sample_epa_fuel_excel
    ):
        """Test that re-sync does not create duplicate records."""
        try:
            from backend.services.data_ingestion.epa_ingestion import (
                EPAEmissionFactorsIngestion
            )
            from backend.models import EmissionFactor
        except ImportError:
            pytest.skip("Required modules not yet implemented")

        ingestion_temp = EPAEmissionFactorsIngestion(
            db=MagicMock(),
            data_source_id=data_source.id,
            file_key="fuels"
        )
        mock_url = ingestion_temp.file_config["url"]

        respx.get(mock_url).mock(
            return_value=httpx.Response(200, content=sample_epa_fuel_excel)
        )

        mock_session = create_mock_async_session(db_session)

        # Run sync 3 times
        for _ in range(3):
            ingestion = EPAEmissionFactorsIngestion(
                db=mock_session,
                data_source_id=data_source.id,
                file_key="fuels"
            )
            await ingestion.execute_sync()

        # Count records with EPA external_id pattern
        epa_factors = db_session.query(EmissionFactor).filter(
            EmissionFactor.external_id.like("EPA_fuels_%")
        ).all()

        # Group by external_id to check for duplicates
        external_ids = [f.external_id for f in epa_factors]
        unique_ids = set(external_ids)

        # No duplicates - count should equal unique count
        assert len(external_ids) == len(unique_ids)

    @pytest.mark.asyncio
    @respx.mock
    async def test_updated_records_reflected_in_database(
        self, db_session, data_source
    ):
        """Test that updated records are reflected in database."""
        try:
            from backend.services.data_ingestion.epa_ingestion import (
                EPAEmissionFactorsIngestion
            )
            from backend.models import EmissionFactor
            from openpyxl import Workbook
        except ImportError:
            pytest.skip("Required modules not yet implemented")

        ingestion_temp = EPAEmissionFactorsIngestion(
            db=MagicMock(),
            data_source_id=data_source.id,
            file_key="fuels"
        )
        mock_url = ingestion_temp.file_config["url"]

        # Create initial Excel file with original value
        def create_excel(co2e_value):
            wb = Workbook()
            ws = wb.active
            ws.title = "Table 1 - Fuel"
            headers = ["Fuel Type", "kg CO2e per unit", "Unit", "Category"]
            for col, header in enumerate(headers, 1):
                ws.cell(row=1, column=col, value=header)
            ws.cell(row=2, column=1, value="Test Fuel")
            ws.cell(row=2, column=2, value=co2e_value)
            ws.cell(row=2, column=3, value="kg")
            ws.cell(row=2, column=4, value="Stationary Combustion")
            output = BytesIO()
            wb.save(output)
            output.seek(0)
            return output.getvalue()

        # First sync with original value
        respx.get(mock_url).mock(
            return_value=httpx.Response(200, content=create_excel(1.0))
        )

        mock_session = create_mock_async_session(db_session)
        ingestion1 = EPAEmissionFactorsIngestion(
            db=mock_session,
            data_source_id=data_source.id,
            file_key="fuels"
        )
        await ingestion1.execute_sync()

        # Verify initial value
        factor = db_session.query(EmissionFactor).filter(
            EmissionFactor.external_id.like("%Test_Fuel%")
        ).first()
        assert factor is not None
        initial_factor = float(factor.co2e_factor)
        assert initial_factor == 1.0

        # Re-sync with updated value
        respx.reset()
        respx.get(mock_url).mock(
            return_value=httpx.Response(200, content=create_excel(2.5))
        )

        ingestion2 = EPAEmissionFactorsIngestion(
            db=mock_session,
            data_source_id=data_source.id,
            file_key="fuels"
        )
        await ingestion2.execute_sync()

        # Verify updated value
        db_session.expire_all()
        updated_factor = db_session.query(EmissionFactor).filter(
            EmissionFactor.external_id.like("%Test_Fuel%")
        ).first()
        assert updated_factor is not None
        assert float(updated_factor.co2e_factor) == 2.5


# ============================================================================
# Test Scenario 4: Sync Log Lifecycle
# ============================================================================

class TestSyncLogLifecycle:
    """Test sync log lifecycle (in_progress -> completed)."""

    @pytest.mark.asyncio
    @respx.mock
    async def test_sync_log_created_for_sync(
        self, db_session, data_source, sample_epa_fuel_excel
    ):
        """Test that sync log is created for each sync operation."""
        try:
            from backend.services.data_ingestion.epa_ingestion import (
                EPAEmissionFactorsIngestion
            )
            from backend.models import DataSyncLog
        except ImportError:
            pytest.skip("Required modules not yet implemented")

        ingestion_temp = EPAEmissionFactorsIngestion(
            db=MagicMock(),
            data_source_id=data_source.id,
            file_key="fuels"
        )
        mock_url = ingestion_temp.file_config["url"]

        respx.get(mock_url).mock(
            return_value=httpx.Response(200, content=sample_epa_fuel_excel)
        )

        mock_session = create_mock_async_session(db_session)
        ingestion = EPAEmissionFactorsIngestion(
            db=mock_session,
            data_source_id=data_source.id,
            file_key="fuels",
            sync_type="manual"
        )

        await ingestion.execute_sync()

        # Verify sync log was created
        sync_log = db_session.query(DataSyncLog).filter(
            DataSyncLog.data_source_id == data_source.id
        ).first()

        assert sync_log is not None
        assert sync_log.sync_type == "manual"

    @pytest.mark.asyncio
    @respx.mock
    async def test_sync_log_status_completed_on_success(
        self, db_session, data_source, sample_epa_fuel_excel
    ):
        """Test that sync log status is 'completed' on success."""
        try:
            from backend.services.data_ingestion.epa_ingestion import (
                EPAEmissionFactorsIngestion
            )
            from backend.models import DataSyncLog
        except ImportError:
            pytest.skip("Required modules not yet implemented")

        ingestion_temp = EPAEmissionFactorsIngestion(
            db=MagicMock(),
            data_source_id=data_source.id,
            file_key="fuels"
        )
        mock_url = ingestion_temp.file_config["url"]

        respx.get(mock_url).mock(
            return_value=httpx.Response(200, content=sample_epa_fuel_excel)
        )

        mock_session = create_mock_async_session(db_session)
        ingestion = EPAEmissionFactorsIngestion(
            db=mock_session,
            data_source_id=data_source.id,
            file_key="fuels"
        )

        await ingestion.execute_sync()

        # Verify sync log status
        sync_log = db_session.query(DataSyncLog).filter(
            DataSyncLog.data_source_id == data_source.id
        ).order_by(DataSyncLog.created_at.desc()).first()

        assert sync_log is not None
        assert sync_log.status == "completed"
        assert sync_log.completed_at is not None

    @pytest.mark.asyncio
    @respx.mock
    async def test_sync_log_records_processing_stats(
        self, db_session, data_source, sample_epa_fuel_excel
    ):
        """Test that sync log records processing statistics."""
        try:
            from backend.services.data_ingestion.epa_ingestion import (
                EPAEmissionFactorsIngestion
            )
            from backend.models import DataSyncLog
        except ImportError:
            pytest.skip("Required modules not yet implemented")

        ingestion_temp = EPAEmissionFactorsIngestion(
            db=MagicMock(),
            data_source_id=data_source.id,
            file_key="fuels"
        )
        mock_url = ingestion_temp.file_config["url"]

        respx.get(mock_url).mock(
            return_value=httpx.Response(200, content=sample_epa_fuel_excel)
        )

        mock_session = create_mock_async_session(db_session)
        ingestion = EPAEmissionFactorsIngestion(
            db=mock_session,
            data_source_id=data_source.id,
            file_key="fuels"
        )

        await ingestion.execute_sync()

        # Verify sync log statistics
        sync_log = db_session.query(DataSyncLog).filter(
            DataSyncLog.data_source_id == data_source.id
        ).order_by(DataSyncLog.created_at.desc()).first()

        assert sync_log is not None
        assert sync_log.records_processed >= 10
        assert sync_log.records_created >= 10
        assert sync_log.records_failed == 0

    @pytest.mark.asyncio
    @respx.mock
    async def test_sync_log_status_failed_on_error(
        self, db_session, data_source
    ):
        """Test that sync log status is 'failed' on error."""
        try:
            from backend.services.data_ingestion.epa_ingestion import (
                EPAEmissionFactorsIngestion
            )
            from backend.models import DataSyncLog
        except ImportError:
            pytest.skip("Required modules not yet implemented")

        ingestion_temp = EPAEmissionFactorsIngestion(
            db=MagicMock(),
            data_source_id=data_source.id,
            file_key="fuels"
        )
        mock_url = ingestion_temp.file_config["url"]

        # Mock HTTP 500 error
        respx.get(mock_url).mock(
            return_value=httpx.Response(500)
        )

        mock_session = create_mock_async_session(db_session)
        ingestion = EPAEmissionFactorsIngestion(
            db=mock_session,
            data_source_id=data_source.id,
            file_key="fuels"
        )

        try:
            await ingestion.execute_sync()
        except Exception:
            pass

        # Verify sync log status is failed
        sync_log = db_session.query(DataSyncLog).filter(
            DataSyncLog.data_source_id == data_source.id
        ).order_by(DataSyncLog.created_at.desc()).first()

        assert sync_log is not None
        assert sync_log.status == "failed"
        assert sync_log.error_message is not None

    @pytest.mark.asyncio
    @respx.mock
    async def test_multiple_syncs_create_separate_logs(
        self, db_session, data_source, sample_epa_fuel_excel
    ):
        """Test that multiple syncs create separate sync logs."""
        try:
            from backend.services.data_ingestion.epa_ingestion import (
                EPAEmissionFactorsIngestion
            )
            from backend.models import DataSyncLog
        except ImportError:
            pytest.skip("Required modules not yet implemented")

        ingestion_temp = EPAEmissionFactorsIngestion(
            db=MagicMock(),
            data_source_id=data_source.id,
            file_key="fuels"
        )
        mock_url = ingestion_temp.file_config["url"]

        respx.get(mock_url).mock(
            return_value=httpx.Response(200, content=sample_epa_fuel_excel)
        )

        mock_session = create_mock_async_session(db_session)

        # Run 3 syncs
        for i in range(3):
            ingestion = EPAEmissionFactorsIngestion(
                db=mock_session,
                data_source_id=data_source.id,
                file_key="fuels",
                sync_type=f"sync_{i}"
            )
            await ingestion.execute_sync()

        # Verify 3 sync logs created
        sync_logs = db_session.query(DataSyncLog).filter(
            DataSyncLog.data_source_id == data_source.id
        ).all()

        assert len(sync_logs) == 3


# ============================================================================
# Test Scenario 5: Data Quality Validation
# ============================================================================

class TestDataQuality:
    """Test data quality validation during sync."""

    @pytest.mark.asyncio
    @respx.mock
    async def test_fuel_factors_have_data_quality_rating(
        self, db_session, data_source, sample_epa_fuel_excel
    ):
        """Test that fuel factors have data quality rating set."""
        try:
            from backend.services.data_ingestion.epa_ingestion import (
                EPAEmissionFactorsIngestion
            )
            from backend.models import EmissionFactor
        except ImportError:
            pytest.skip("Required modules not yet implemented")

        ingestion_temp = EPAEmissionFactorsIngestion(
            db=MagicMock(),
            data_source_id=data_source.id,
            file_key="fuels"
        )
        mock_url = ingestion_temp.file_config["url"]

        respx.get(mock_url).mock(
            return_value=httpx.Response(200, content=sample_epa_fuel_excel)
        )

        mock_session = create_mock_async_session(db_session)
        ingestion = EPAEmissionFactorsIngestion(
            db=mock_session,
            data_source_id=data_source.id,
            file_key="fuels"
        )

        await ingestion.execute_sync()

        # Verify data quality rating is set
        factors = db_session.query(EmissionFactor).filter(
            EmissionFactor.data_source_id == data_source.id
        ).all()

        for factor in factors:
            assert factor.data_quality_rating is not None
            # EPA is a high-quality source
            assert float(factor.data_quality_rating) >= 0.85

    @pytest.mark.asyncio
    @respx.mock
    async def test_egrid_factors_have_high_data_quality(
        self, db_session, data_source, sample_epa_egrid_excel
    ):
        """Test that eGRID factors have high data quality rating."""
        try:
            from backend.services.data_ingestion.epa_ingestion import (
                EPAEmissionFactorsIngestion
            )
            from backend.models import EmissionFactor
        except ImportError:
            pytest.skip("Required modules not yet implemented")

        ingestion_temp = EPAEmissionFactorsIngestion(
            db=MagicMock(),
            data_source_id=data_source.id,
            file_key="egrid"
        )
        mock_url = ingestion_temp.file_config["url"]

        respx.get(mock_url).mock(
            return_value=httpx.Response(200, content=sample_epa_egrid_excel)
        )

        mock_session = create_mock_async_session(db_session)
        ingestion = EPAEmissionFactorsIngestion(
            db=mock_session,
            data_source_id=data_source.id,
            file_key="egrid"
        )

        await ingestion.execute_sync()

        # Verify eGRID has highest data quality
        factors = db_session.query(EmissionFactor).filter(
            EmissionFactor.data_source_id == data_source.id
        ).all()

        for factor in factors:
            assert factor.data_quality_rating is not None
            # eGRID is highly reliable
            assert float(factor.data_quality_rating) >= 0.90

    @pytest.mark.asyncio
    @respx.mock
    async def test_factors_have_us_geography(
        self, db_session, data_source, sample_epa_fuel_excel
    ):
        """Test that EPA factors have US geography."""
        try:
            from backend.services.data_ingestion.epa_ingestion import (
                EPAEmissionFactorsIngestion
            )
            from backend.models import EmissionFactor
        except ImportError:
            pytest.skip("Required modules not yet implemented")

        ingestion_temp = EPAEmissionFactorsIngestion(
            db=MagicMock(),
            data_source_id=data_source.id,
            file_key="fuels"
        )
        mock_url = ingestion_temp.file_config["url"]

        respx.get(mock_url).mock(
            return_value=httpx.Response(200, content=sample_epa_fuel_excel)
        )

        mock_session = create_mock_async_session(db_session)
        ingestion = EPAEmissionFactorsIngestion(
            db=mock_session,
            data_source_id=data_source.id,
            file_key="fuels"
        )

        await ingestion.execute_sync()

        # Verify geography is US
        factors = db_session.query(EmissionFactor).filter(
            EmissionFactor.data_source_id == data_source.id
        ).all()

        for factor in factors:
            assert factor.geography == "US"


# ============================================================================
# Test Scenario 6: Unit Conversion Verification
# ============================================================================

class TestUnitConversion:
    """Test unit conversion for eGRID data."""

    @pytest.mark.asyncio
    @respx.mock
    async def test_egrid_lb_per_mwh_converted_to_kg_per_kwh(
        self, db_session, data_source
    ):
        """Test that eGRID lb/MWh is correctly converted to kg/kWh."""
        try:
            from backend.services.data_ingestion.epa_ingestion import (
                EPAEmissionFactorsIngestion
            )
            from backend.models import EmissionFactor
            from openpyxl import Workbook
        except ImportError:
            pytest.skip("Required modules not yet implemented")

        # Create eGRID file with known value
        wb = Workbook()
        ws = wb.active
        ws.title = "SUBRGN22"
        ws.cell(row=1, column=1, value="SUBRGN")
        ws.cell(row=1, column=2, value="SRCO2RTA")
        ws.cell(row=2, column=1, value="TEST")
        ws.cell(row=2, column=2, value=1000.0)  # 1000 lb/MWh

        output = BytesIO()
        wb.save(output)
        output.seek(0)
        test_excel = output.getvalue()

        ingestion_temp = EPAEmissionFactorsIngestion(
            db=MagicMock(),
            data_source_id=data_source.id,
            file_key="egrid"
        )
        mock_url = ingestion_temp.file_config["url"]

        respx.get(mock_url).mock(
            return_value=httpx.Response(200, content=test_excel)
        )

        mock_session = create_mock_async_session(db_session)
        ingestion = EPAEmissionFactorsIngestion(
            db=mock_session,
            data_source_id=data_source.id,
            file_key="egrid"
        )

        await ingestion.execute_sync()

        # Verify conversion: 1000 lb/MWh * 0.453592 / 1000 = 0.453592 kg/kWh
        factor = db_session.query(EmissionFactor).filter(
            EmissionFactor.external_id.like("%TEST%")
        ).first()

        assert factor is not None
        expected_value = 1000.0 * 0.453592 / 1000
        assert abs(float(factor.co2e_factor) - expected_value) < 0.001
        assert factor.unit == "kg CO2e/kWh"
