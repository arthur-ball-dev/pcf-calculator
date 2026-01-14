"""
Integration test suite for External Data Connectors (EPA/DEFRA).

TASK-DATA-P8-003: Activate External Data Connectors
Phase A: Test-Driven Development - Tests First

This test suite validates:
1. Both connectors (EPA, DEFRA) sync successfully
2. Expected emission factor counts by source
3. Data quality validation (positive factors, valid units)
4. DataSyncLog entries created correctly
5. Idempotency - re-running syncs doesn't create duplicates

Expected Counts (from SPEC):
- EPA: 75-125 records
- DEFRA: 200-400 records
- Total: 275-525 records

Test-Driven Development Protocol:
- These tests MUST be committed BEFORE implementation changes
- Tests define the expected behavior, implementation must make them pass

NOTE: Exiobase has been removed from this project due to licensing constraints.
"""

import pytest
from unittest.mock import MagicMock, AsyncMock, patch
from io import BytesIO
from datetime import datetime, timezone

from sqlalchemy import create_engine, text
from sqlalchemy.orm import sessionmaker
from sqlalchemy.pool import StaticPool


# Mark all tests as integration tests
pytestmark = pytest.mark.integration


# =============================================================================
# Test Fixtures
# =============================================================================


@pytest.fixture(scope="function")
def db_engine():
    """Create in-memory SQLite database for testing."""
    engine = create_engine(
        "sqlite:///:memory:",
        connect_args={"check_same_thread": False},
        poolclass=StaticPool,
        echo=False
    )
    return engine


@pytest.fixture(scope="function")
def db_session(db_engine):
    """Create database session for testing."""
    from backend.models import Base

    Base.metadata.create_all(db_engine)
    SessionLocal = sessionmaker(bind=db_engine)
    session = SessionLocal()

    # Enable foreign key constraints for SQLite
    session.execute(text("PRAGMA foreign_keys = ON"))
    session.commit()

    yield session
    session.close()


@pytest.fixture
def epa_data_source(db_session):
    """Create EPA data source for testing."""
    from backend.models import DataSource

    source = DataSource(
        name="EPA GHG Emission Factors Hub",
        source_type="file",
        base_url="https://www.epa.gov/climateleadership/ghg-emission-factors-hub",
        sync_frequency="biweekly",
        is_active=True,
        license_type="Public Domain",
        requires_attribution=False,
        allows_commercial_use=True,
    )
    db_session.add(source)
    db_session.commit()
    db_session.refresh(source)
    return source


@pytest.fixture
def defra_data_source(db_session):
    """Create DEFRA data source for testing."""
    from backend.models import DataSource

    source = DataSource(
        name="DEFRA Conversion Factors",
        source_type="file",
        base_url="https://www.gov.uk/government/publications/greenhouse-gas-reporting-conversion-factors",
        sync_frequency="biweekly",
        is_active=True,
        license_type="Open Government Licence v3.0",
        requires_attribution=True,
        allows_commercial_use=True,
    )
    db_session.add(source)
    db_session.commit()
    db_session.refresh(source)
    return source




@pytest.fixture
def all_data_sources(db_session, epa_data_source, defra_data_source):
    """Create all data sources."""
    return {
        "EPA": epa_data_source,
        "DEFRA": defra_data_source,
    }


def create_mock_async_session(db_session):
    """Create an async-compatible mock that delegates to sync session."""
    mock_session = MagicMock()
    mock_session.add = db_session.add
    mock_session.flush = AsyncMock(side_effect=lambda: db_session.flush())
    mock_session.commit = AsyncMock(side_effect=lambda: db_session.commit())
    mock_session.rollback = AsyncMock(side_effect=lambda: db_session.rollback())
    mock_session.refresh = AsyncMock(side_effect=lambda obj: db_session.refresh(obj))

    async def mock_execute(stmt):
        result = db_session.execute(stmt)
        mock_result = MagicMock()
        mock_result.rowcount = result.rowcount
        mock_result.scalars = result.scalars
        mock_result.fetchall = result.fetchall
        mock_result.fetchone = result.fetchone
        return mock_result

    mock_session.execute = mock_execute
    return mock_session


# =============================================================================
# Mock Data Fixtures
# =============================================================================


@pytest.fixture
def sample_epa_fuel_excel():
    """Create sample EPA fuel Excel with 10+ records."""
    try:
        from openpyxl import Workbook
    except ImportError:
        pytest.skip("openpyxl not available")

    wb = Workbook()
    ws = wb.active
    ws.title = "Table 1 - Fuel"

    headers = ["Fuel Type", "kg CO2e per unit", "Unit", "Category"]
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)

    fuel_data = [
        ("Natural Gas", 2.75, "kg", "Stationary Combustion"),
        ("Diesel", 10.21, "L", "Stationary Combustion"),
        ("Gasoline", 8.89, "L", "Mobile"),
        ("Propane", 6.35, "kg", "Stationary Combustion"),
        ("Residual Fuel Oil", 11.27, "L", "Stationary Combustion"),
        ("Coal (Bituminous)", 2.563, "kg", "Stationary Combustion"),
        ("Kerosene", 9.75, "L", "Stationary Combustion"),
        ("Aviation Gasoline", 8.31, "L", "Mobile - Aviation"),
        ("Jet Fuel", 9.57, "L", "Mobile - Aviation"),
        ("LPG", 5.68, "L", "Stationary Combustion"),
        ("Wood", 1.89, "kg", "Biomass"),
        ("Ethanol", 1.53, "L", "Mobile"),
    ]

    for row_idx, (fuel_type, factor, unit, category) in enumerate(fuel_data, 2):
        ws.cell(row=row_idx, column=1, value=fuel_type)
        ws.cell(row=row_idx, column=2, value=factor)
        ws.cell(row=row_idx, column=3, value=unit)
        ws.cell(row=row_idx, column=4, value=category)

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


@pytest.fixture
def sample_defra_excel():
    """Create sample DEFRA Excel with multiple sheets."""
    try:
        from openpyxl import Workbook
    except ImportError:
        pytest.skip("openpyxl not available")

    wb = Workbook()

    # Fuels sheet
    ws_fuels = wb.active
    ws_fuels.title = "Fuels"

    # Add some header rows (DEFRA has metadata at top)
    ws_fuels.cell(row=1, column=1, value="DEFRA 2024 Conversion Factors")
    ws_fuels.cell(row=2, column=1, value="Fuels")

    # Actual header row
    headers = ["Fuel", "Unit", "kg CO2e"]
    for col, header in enumerate(headers, 1):
        ws_fuels.cell(row=3, column=col, value=header)

    fuels_data = [
        ("Natural Gas", "kWh", 0.183),
        ("Diesel", "litre", 2.513),
        ("Petrol", "litre", 2.194),
        ("Coal (domestic)", "tonne", 2883.5),
        ("LPG", "litre", 1.513),
    ]

    for row_idx, (fuel, unit, factor) in enumerate(fuels_data, 4):
        ws_fuels.cell(row=row_idx, column=1, value=fuel)
        ws_fuels.cell(row=row_idx, column=2, value=unit)
        ws_fuels.cell(row=row_idx, column=3, value=factor)

    # Electricity sheet
    ws_elec = wb.create_sheet("Electricity")
    ws_elec.cell(row=1, column=1, value="Electricity")
    elec_headers = ["Activity", "Unit", "kg CO2e"]
    for col, header in enumerate(elec_headers, 1):
        ws_elec.cell(row=2, column=col, value=header)

    elec_data = [
        ("UK Grid Electricity", "kWh", 0.212),
        ("Renewable Electricity", "kWh", 0.0),
    ]

    for row_idx, (activity, unit, factor) in enumerate(elec_data, 3):
        ws_elec.cell(row=row_idx, column=1, value=activity)
        ws_elec.cell(row=row_idx, column=2, value=unit)
        ws_elec.cell(row=row_idx, column=3, value=factor)

    # Material use sheet
    ws_mat = wb.create_sheet("Material use")
    ws_mat.cell(row=1, column=1, value="Material use")
    mat_headers = ["Material", "Unit", "kg CO2e"]
    for col, header in enumerate(mat_headers, 1):
        ws_mat.cell(row=2, column=col, value=header)

    mat_data = [
        ("Aluminium", "tonne", 9.14),
        ("Steel", "tonne", 1.77),
        ("Paper", "tonne", 0.919),
    ]

    for row_idx, (material, unit, factor) in enumerate(mat_data, 3):
        ws_mat.cell(row=row_idx, column=1, value=material)
        ws_mat.cell(row=row_idx, column=2, value=unit)
        ws_mat.cell(row=row_idx, column=3, value=factor)

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()






# =============================================================================
# Test Class: EPA Sync
# =============================================================================


class TestEPASync:
    """Test EPA emission factor synchronization."""

    @pytest.mark.asyncio
    async def test_epa_sync_creates_emission_factors(
        self, db_session, epa_data_source, sample_epa_fuel_excel
    ):
        """Test that EPA sync creates emission factors in database."""
        import respx
        import httpx
        from backend.services.data_ingestion.epa_ingestion import (
            EPAEmissionFactorsIngestion
        )
        from backend.models import EmissionFactor

        ingestion = EPAEmissionFactorsIngestion(
            db=MagicMock(),
            data_source_id=epa_data_source.id,
            file_key="fuels"
        )
        mock_url = ingestion.file_config["url"]

        with respx.mock:
            respx.get(mock_url).mock(
                return_value=httpx.Response(200, content=sample_epa_fuel_excel)
            )

            mock_session = create_mock_async_session(db_session)
            ingestion = EPAEmissionFactorsIngestion(
                db=mock_session,
                data_source_id=epa_data_source.id,
                file_key="fuels",
                sync_type="initial"
            )

            result = await ingestion.execute_sync()

        # Verify sync completed
        assert result.status == "completed"
        assert result.records_processed >= 10
        assert result.records_created >= 10
        assert result.records_failed == 0

        # Verify factors in database
        factors = db_session.query(EmissionFactor).filter(
            EmissionFactor.data_source_id == epa_data_source.id
        ).all()
        assert len(factors) >= 10

    @pytest.mark.asyncio
    async def test_epa_sync_sets_correct_geography(
        self, db_session, epa_data_source, sample_epa_fuel_excel
    ):
        """Test that EPA factors have US geography."""
        import respx
        import httpx
        from backend.services.data_ingestion.epa_ingestion import (
            EPAEmissionFactorsIngestion
        )
        from backend.models import EmissionFactor

        ingestion_temp = EPAEmissionFactorsIngestion(
            db=MagicMock(),
            data_source_id=epa_data_source.id,
            file_key="fuels"
        )
        mock_url = ingestion_temp.file_config["url"]

        with respx.mock:
            respx.get(mock_url).mock(
                return_value=httpx.Response(200, content=sample_epa_fuel_excel)
            )

            mock_session = create_mock_async_session(db_session)
            ingestion = EPAEmissionFactorsIngestion(
                db=mock_session,
                data_source_id=epa_data_source.id,
                file_key="fuels"
            )
            await ingestion.execute_sync()

        factors = db_session.query(EmissionFactor).filter(
            EmissionFactor.data_source_id == epa_data_source.id
        ).all()

        for factor in factors:
            assert factor.geography == "US", f"EPA factor {factor.activity_name} has wrong geography"

    @pytest.mark.asyncio
    async def test_epa_factors_have_positive_values(
        self, db_session, epa_data_source, sample_epa_fuel_excel
    ):
        """Test that all EPA factors have positive CO2e values."""
        import respx
        import httpx
        from backend.services.data_ingestion.epa_ingestion import (
            EPAEmissionFactorsIngestion
        )
        from backend.models import EmissionFactor

        ingestion_temp = EPAEmissionFactorsIngestion(
            db=MagicMock(),
            data_source_id=epa_data_source.id,
            file_key="fuels"
        )
        mock_url = ingestion_temp.file_config["url"]

        with respx.mock:
            respx.get(mock_url).mock(
                return_value=httpx.Response(200, content=sample_epa_fuel_excel)
            )

            mock_session = create_mock_async_session(db_session)
            ingestion = EPAEmissionFactorsIngestion(
                db=mock_session,
                data_source_id=epa_data_source.id,
                file_key="fuels"
            )
            await ingestion.execute_sync()

        factors = db_session.query(EmissionFactor).filter(
            EmissionFactor.data_source_id == epa_data_source.id
        ).all()

        for factor in factors:
            assert float(factor.co2e_factor) > 0, (
                f"EPA factor {factor.activity_name} has non-positive value"
            )


# =============================================================================
# Test Class: DEFRA Sync
# =============================================================================


class TestDEFRASync:
    """Test DEFRA emission factor synchronization."""

    @pytest.mark.asyncio
    async def test_defra_sync_creates_emission_factors(
        self, db_session, defra_data_source, sample_defra_excel
    ):
        """Test that DEFRA sync creates emission factors in database."""
        import respx
        import httpx
        from backend.services.data_ingestion.defra_ingestion import (
            DEFRAEmissionFactorsIngestion
        )
        from backend.models import EmissionFactor

        with respx.mock:
            respx.get(url__regex=r".*gov\.uk.*").mock(
                return_value=httpx.Response(200, content=sample_defra_excel)
            )

            mock_session = create_mock_async_session(db_session)
            ingestion = DEFRAEmissionFactorsIngestion(
                db=mock_session,
                data_source_id=defra_data_source.id,
                sync_type="initial"
            )

            result = await ingestion.execute_sync()

        assert result.status == "completed"
        # DEFRA has multiple sheets - should have several factors
        assert result.records_processed >= 5
        assert result.records_created >= 5

    @pytest.mark.asyncio
    async def test_defra_sync_sets_gb_geography(
        self, db_session, defra_data_source, sample_defra_excel
    ):
        """Test that DEFRA factors have GB geography."""
        import respx
        import httpx
        from backend.services.data_ingestion.defra_ingestion import (
            DEFRAEmissionFactorsIngestion
        )
        from backend.models import EmissionFactor

        with respx.mock:
            respx.get(url__regex=r".*gov\.uk.*|.*service\.gov\.uk.*").mock(
                return_value=httpx.Response(200, content=sample_defra_excel)
            )

            mock_session = create_mock_async_session(db_session)
            ingestion = DEFRAEmissionFactorsIngestion(
                db=mock_session,
                data_source_id=defra_data_source.id
            )
            await ingestion.execute_sync()

        factors = db_session.query(EmissionFactor).filter(
            EmissionFactor.data_source_id == defra_data_source.id
        ).all()

        for factor in factors:
            assert factor.geography == "GB", (
                f"DEFRA factor {factor.activity_name} has wrong geography"
            )




# =============================================================================
# Test Class: DataSyncLog Lifecycle
# =============================================================================


class TestDataSyncLogLifecycle:
    """Test DataSyncLog creation and updates."""

    @pytest.mark.asyncio
    async def test_sync_creates_log_entry(
        self, db_session, epa_data_source, sample_epa_fuel_excel
    ):
        """Test that sync creates a DataSyncLog entry."""
        import respx
        import httpx
        from backend.services.data_ingestion.epa_ingestion import (
            EPAEmissionFactorsIngestion
        )
        from backend.models import DataSyncLog

        ingestion_temp = EPAEmissionFactorsIngestion(
            db=MagicMock(),
            data_source_id=epa_data_source.id,
            file_key="fuels"
        )
        mock_url = ingestion_temp.file_config["url"]

        with respx.mock:
            respx.get(mock_url).mock(
                return_value=httpx.Response(200, content=sample_epa_fuel_excel)
            )

            mock_session = create_mock_async_session(db_session)
            ingestion = EPAEmissionFactorsIngestion(
                db=mock_session,
                data_source_id=epa_data_source.id,
                file_key="fuels",
                sync_type="initial"
            )
            await ingestion.execute_sync()

        sync_log = db_session.query(DataSyncLog).filter(
            DataSyncLog.data_source_id == epa_data_source.id
        ).first()

        assert sync_log is not None
        assert sync_log.status == "completed"
        assert sync_log.sync_type == "initial"
        assert sync_log.records_created >= 10

    @pytest.mark.asyncio
    async def test_sync_log_records_statistics(
        self, db_session, epa_data_source, sample_epa_fuel_excel
    ):
        """Test that sync log records correct statistics."""
        import respx
        import httpx
        from backend.services.data_ingestion.epa_ingestion import (
            EPAEmissionFactorsIngestion
        )
        from backend.models import DataSyncLog

        ingestion_temp = EPAEmissionFactorsIngestion(
            db=MagicMock(),
            data_source_id=epa_data_source.id,
            file_key="fuels"
        )
        mock_url = ingestion_temp.file_config["url"]

        with respx.mock:
            respx.get(mock_url).mock(
                return_value=httpx.Response(200, content=sample_epa_fuel_excel)
            )

            mock_session = create_mock_async_session(db_session)
            ingestion = EPAEmissionFactorsIngestion(
                db=mock_session,
                data_source_id=epa_data_source.id,
                file_key="fuels"
            )
            await ingestion.execute_sync()

        sync_log = db_session.query(DataSyncLog).filter(
            DataSyncLog.data_source_id == epa_data_source.id
        ).order_by(DataSyncLog.created_at.desc()).first()

        assert sync_log.records_processed >= 10
        assert sync_log.records_created >= 10
        assert sync_log.records_failed == 0


# =============================================================================
# Test Class: Idempotency
# =============================================================================


class TestSyncIdempotency:
    """Test that syncs are idempotent (no duplicates on re-run)."""

    @pytest.mark.asyncio
    async def test_double_sync_no_duplicates(
        self, db_session, epa_data_source, sample_epa_fuel_excel
    ):
        """Test that running sync twice doesn't create duplicates."""
        import respx
        import httpx
        from backend.services.data_ingestion.epa_ingestion import (
            EPAEmissionFactorsIngestion
        )
        from backend.models import EmissionFactor

        ingestion_temp = EPAEmissionFactorsIngestion(
            db=MagicMock(),
            data_source_id=epa_data_source.id,
            file_key="fuels"
        )
        mock_url = ingestion_temp.file_config["url"]

        with respx.mock:
            respx.get(mock_url).mock(
                return_value=httpx.Response(200, content=sample_epa_fuel_excel)
            )

            mock_session = create_mock_async_session(db_session)

            # First sync
            ingestion1 = EPAEmissionFactorsIngestion(
                db=mock_session,
                data_source_id=epa_data_source.id,
                file_key="fuels"
            )
            await ingestion1.execute_sync()

            first_count = db_session.query(EmissionFactor).filter(
                EmissionFactor.data_source_id == epa_data_source.id
            ).count()

            # Second sync
            ingestion2 = EPAEmissionFactorsIngestion(
                db=mock_session,
                data_source_id=epa_data_source.id,
                file_key="fuels"
            )
            await ingestion2.execute_sync()

            second_count = db_session.query(EmissionFactor).filter(
                EmissionFactor.data_source_id == epa_data_source.id
            ).count()

        # Count should be the same (upsert, not insert)
        assert second_count == first_count, (
            f"Duplicate records created: {first_count} -> {second_count}"
        )


# =============================================================================
# Test Class: Data Quality Validation
# =============================================================================


class TestDataQualityValidation:
    """Test data quality validation during sync."""

    @pytest.mark.asyncio
    async def test_all_factors_have_valid_co2e(
        self, db_session, epa_data_source, sample_epa_fuel_excel
    ):
        """Test that all synced factors have valid (positive) CO2e values."""
        import respx
        import httpx
        from backend.services.data_ingestion.epa_ingestion import (
            EPAEmissionFactorsIngestion
        )
        from backend.models import EmissionFactor

        ingestion_temp = EPAEmissionFactorsIngestion(
            db=MagicMock(),
            data_source_id=epa_data_source.id,
            file_key="fuels"
        )
        mock_url = ingestion_temp.file_config["url"]

        with respx.mock:
            respx.get(mock_url).mock(
                return_value=httpx.Response(200, content=sample_epa_fuel_excel)
            )

            mock_session = create_mock_async_session(db_session)
            ingestion = EPAEmissionFactorsIngestion(
                db=mock_session,
                data_source_id=epa_data_source.id,
                file_key="fuels"
            )
            await ingestion.execute_sync()

        factors = db_session.query(EmissionFactor).filter(
            EmissionFactor.data_source_id == epa_data_source.id
        ).all()

        invalid_factors = [
            f for f in factors
            if f.co2e_factor is None or float(f.co2e_factor) <= 0
        ]

        assert len(invalid_factors) == 0, (
            f"Found {len(invalid_factors)} factors with invalid CO2e values"
        )

    @pytest.mark.asyncio
    async def test_all_factors_have_external_id(
        self, db_session, epa_data_source, sample_epa_fuel_excel
    ):
        """Test that all synced factors have external_id set."""
        import respx
        import httpx
        from backend.services.data_ingestion.epa_ingestion import (
            EPAEmissionFactorsIngestion
        )
        from backend.models import EmissionFactor

        ingestion_temp = EPAEmissionFactorsIngestion(
            db=MagicMock(),
            data_source_id=epa_data_source.id,
            file_key="fuels"
        )
        mock_url = ingestion_temp.file_config["url"]

        with respx.mock:
            respx.get(mock_url).mock(
                return_value=httpx.Response(200, content=sample_epa_fuel_excel)
            )

            mock_session = create_mock_async_session(db_session)
            ingestion = EPAEmissionFactorsIngestion(
                db=mock_session,
                data_source_id=epa_data_source.id,
                file_key="fuels"
            )
            await ingestion.execute_sync()

        factors = db_session.query(EmissionFactor).filter(
            EmissionFactor.data_source_id == epa_data_source.id
        ).all()

        factors_without_id = [f for f in factors if not f.external_id]

        assert len(factors_without_id) == 0, (
            f"Found {len(factors_without_id)} factors without external_id"
        )


# =============================================================================
# Test Class: Data Source Registration
# =============================================================================


class TestDataSourceRegistration:
    """Test data source seed data correctness."""

    def test_seed_data_has_all_sources(self):
        """Test that seed data includes EPA and DEFRA sources."""
        from backend.database.seeds.data_sources import SEED_DATA_SOURCES

        source_names = {s["name"] for s in SEED_DATA_SOURCES}

        assert "EPA GHG Emission Factors Hub" in source_names
        assert "DEFRA Conversion Factors" in source_names

    def test_defra_requires_attribution(self):
        """Test that DEFRA seed data has requires_attribution=True."""
        from backend.database.seeds.data_sources import SEED_DATA_SOURCES

        defra_seed = next(
            (s for s in SEED_DATA_SOURCES if "DEFRA" in s["name"]),
            None
        )

        assert defra_seed is not None
        assert defra_seed.get("requires_attribution", False) is True, (
            "DEFRA requires attribution under OGL v3.0"
        )



# =============================================================================
# Test Class: Verification Script Output Format
# =============================================================================


class TestVerificationOutputFormat:
    """Test the expected output format of verification script."""

    def test_factor_count_query_structure(self, db_session, all_data_sources):
        """Test that we can query factor counts by source."""
        from backend.models import EmissionFactor, DataSource

        # Create some test factors with required data_source field
        for source_name, source in all_data_sources.items():
            for i in range(3):
                factor = EmissionFactor(
                    activity_name=f"Test {source_name} {i}",
                    co2e_factor=1.0,
                    unit="kg",
                    geography="US",
                    data_source=source.name,  # Required legacy field
                    data_source_id=source.id,
                    external_id=f"{source_name}_test_{i}",
                )
                db_session.add(factor)

        db_session.commit()

        # Query counts by source
        result = db_session.execute(text("""
            SELECT ds.name, COUNT(ef.id) as factor_count
            FROM data_sources ds
            LEFT JOIN emission_factors ef ON ds.id = ef.data_source_id
            GROUP BY ds.name
            ORDER BY ds.name
        """)).fetchall()

        # Should have 2 rows (one per src: EPA, DEFRA)
        assert len(result) == 2

        # Each should have 3 factors
        for name, count in result:
            assert count == 3, f"Source {name} should have 3 factors"

    def test_no_orphaned_factors_query(self, db_session, epa_data_source):
        """Test query to find orphaned factors."""
        from backend.models import EmissionFactor

        # Create a valid factor with required data_source field
        factor = EmissionFactor(
            activity_name="Test Factor",
            co2e_factor=1.0,
            unit="kg",
            geography="US",
            data_source=epa_data_source.name,  # Required legacy field
            data_source_id=epa_data_source.id,
            external_id="test_factor_1",
        )
        db_session.add(factor)
        db_session.commit()

        # Query for orphaned factors
        result = db_session.execute(text("""
            SELECT COUNT(*) FROM emission_factors ef
            LEFT JOIN data_sources ds ON ef.data_source_id = ds.id
            WHERE ds.id IS NULL
        """)).scalar()

        assert result == 0, "Should have no orphaned factors"
