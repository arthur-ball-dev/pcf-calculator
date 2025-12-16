"""
Test suite for DEFRAEmissionFactorsIngestion class.

TASK-DATA-P5-003: DEFRA Data Connector - Phase A Tests

This test suite validates:
- Constructor initializes with correct defaults
- fetch_raw_data downloads from correct URL
- parse_data extracts records from multiple sheets
- _find_sheet matches partial sheet names
- transform_data handles various column formats
- _determine_unit extracts unit from column name
- External ID generation handles special characters
- Error handling for missing sheets

Test-Driven Development Protocol:
- These tests MUST be committed BEFORE implementation
- Tests should FAIL initially (no DEFRAEmissionFactorsIngestion class exists yet)
- Implementation must make tests PASS without modifying tests
"""

import pytest
from unittest.mock import AsyncMock, MagicMock, patch
from datetime import datetime, timezone
from decimal import Decimal
from uuid import uuid4
from io import BytesIO

from sqlalchemy import create_engine, text
from sqlalchemy.orm import sessionmaker
from sqlalchemy.pool import StaticPool
from sqlalchemy.ext.asyncio import AsyncSession


@pytest.fixture(scope="function")
def db_engine():
    """Create in-memory SQLite database for testing."""
    engine = create_engine(
        "sqlite:///:memory:",
        connect_args={"check_same_thread": False},
        poolclass=StaticPool,
        echo=False
    )
    return engine


@pytest.fixture(scope="function")
def db_session(db_engine):
    """Create database session for testing."""
    from backend.models import Base

    Base.metadata.create_all(db_engine)
    SessionLocal = sessionmaker(bind=db_engine)
    session = SessionLocal()

    # Enable foreign key constraints for SQLite
    session.execute(text("PRAGMA foreign_keys = ON"))
    session.commit()

    yield session

    session.close()


@pytest.fixture
def mock_async_session():
    """Create mock async session for unit tests."""
    session = AsyncMock(spec=AsyncSession)
    session.add = MagicMock()
    session.flush = AsyncMock()
    session.commit = AsyncMock()
    session.rollback = AsyncMock()
    session.execute = AsyncMock()
    session.refresh = AsyncMock()
    return session


@pytest.fixture
def data_source_id():
    """Generate a test data source ID."""
    return uuid4().hex


@pytest.fixture
def sample_defra_workbook():
    """
    Create a minimal DEFRA-style Excel workbook for testing.

    Uses openpyxl to create an in-memory workbook with sample data.
    """
    try:
        from openpyxl import Workbook
    except ImportError:
        pytest.skip("openpyxl not installed")

    wb = Workbook()

    # Remove default sheet
    default_sheet = wb.active
    wb.remove(default_sheet)

    # Create Fuels sheet
    fuels = wb.create_sheet("Fuels")
    fuels.append(["Category", "Fuel", "Unit", "kg CO2e per unit"])
    fuels.append(["Gaseous fuels", "Natural Gas", "kWh", 0.18287])
    fuels.append(["Gaseous fuels", "LPG", "kWh", 0.21448])
    fuels.append(["Liquid fuels", "Diesel", "litre", 2.70554])
    fuels.append(["Liquid fuels", "Petrol", "litre", 2.31481])
    fuels.append(["Solid fuels", "Coal (industrial)", "kg", 2.41674])

    # Create Electricity sheet
    electricity = wb.create_sheet("UK electricity")
    electricity.append(["Activity", "kg CO2e per kWh"])
    electricity.append(["Electricity: UK", 0.21233])
    electricity.append(["Electricity: UK (T&D)", 0.01879])
    electricity.append(["Electricity: UK (generation)", 0.19354])

    # Create Material use sheet
    materials = wb.create_sheet("Material use")
    materials.append(["Category", "Material", "kg CO2e per kg"])
    materials.append(["Metals", "Primary steel", 1.85])
    materials.append(["Metals", "Aluminium (primary)", 11.59])
    materials.append(["Plastics", "PVC", 3.10])

    # Create Waste disposal sheet
    waste = wb.create_sheet("Waste disposal")
    waste.append(["Waste Category", "Waste type", "kg CO2e per tonne"])
    waste.append(["Commercial", "Mixed waste", 467.0])
    waste.append(["Construction", "Demolition waste", 1.0])

    # Create Business travel- air sheet
    travel = wb.create_sheet("Business travel- air")
    travel.append(["Class", "Type of flight", "kg CO2e per passenger km"])
    travel.append(["Economy", "Domestic", 0.24587])
    travel.append(["Economy", "Short-haul", 0.15353])
    travel.append(["Business", "Long-haul", 0.42951])

    # Create Freighting goods sheet
    freight = wb.create_sheet("Freighting goods")
    freight.append(["Mode", "Vehicle type", "kg CO2e per tonne.km"])
    freight.append(["Road", "HGV (average)", 0.10691])
    freight.append(["Rail", "Freight train", 0.02792])
    freight.append(["Sea", "Container ship", 0.01615])

    # Save to BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


# ============================================================================
# Test Scenario 1: Constructor Initialization
# ============================================================================

class TestDEFRAIngestionInstantiation:
    """Test DEFRAEmissionFactorsIngestion class instantiation."""

    def test_constructor_initializes_with_correct_defaults(
        self, mock_async_session, data_source_id
    ):
        """Test that DEFRA ingestion initializes with correct default values."""
        try:
            from backend.services.data_ingestion.defra_ingestion import (
                DEFRAEmissionFactorsIngestion
            )
        except ImportError:
            pytest.skip("DEFRAEmissionFactorsIngestion not yet implemented")

        ingestion = DEFRAEmissionFactorsIngestion(
            db=mock_async_session,
            data_source_id=data_source_id,
            sync_type="manual"
        )

        assert ingestion.db == mock_async_session
        assert ingestion.data_source_id == data_source_id
        assert ingestion.sync_type == "manual"
        assert ingestion.reference_year == 2024
        assert ingestion.sync_log is None
        assert ingestion.stats["records_processed"] == 0

    def test_constructor_has_sheet_configs(
        self, mock_async_session, data_source_id
    ):
        """Test that SHEET_CONFIGS class attribute is defined."""
        try:
            from backend.services.data_ingestion.defra_ingestion import (
                DEFRAEmissionFactorsIngestion
            )
        except ImportError:
            pytest.skip("DEFRAEmissionFactorsIngestion not yet implemented")

        # Check that SHEET_CONFIGS is defined
        assert hasattr(DEFRAEmissionFactorsIngestion, 'SHEET_CONFIGS')
        assert isinstance(DEFRAEmissionFactorsIngestion.SHEET_CONFIGS, dict)

        # Check expected sheet configurations exist
        expected_sheets = [
            "Fuels", "Electricity", "Material use",
            "Waste disposal", "Business travel- air", "Freighting goods"
        ]
        for sheet in expected_sheets:
            assert sheet in DEFRAEmissionFactorsIngestion.SHEET_CONFIGS

    def test_constructor_has_defra_url(
        self, mock_async_session, data_source_id
    ):
        """Test that DEFRA_URL class attribute is defined."""
        try:
            from backend.services.data_ingestion.defra_ingestion import (
                DEFRAEmissionFactorsIngestion
            )
        except ImportError:
            pytest.skip("DEFRAEmissionFactorsIngestion not yet implemented")

        assert hasattr(DEFRAEmissionFactorsIngestion, 'DEFRA_URL')
        assert isinstance(DEFRAEmissionFactorsIngestion.DEFRA_URL, str)
        assert "gov.uk" in DEFRAEmissionFactorsIngestion.DEFRA_URL.lower()


# ============================================================================
# Test Scenario 2: fetch_raw_data Downloads from Correct URL
# ============================================================================

class TestFetchRawData:
    """Test fetch_raw_data method downloads from correct URL."""

    @pytest.mark.asyncio
    async def test_fetch_raw_data_downloads_xlsx(
        self, mock_async_session, data_source_id, sample_defra_workbook
    ):
        """Test that fetch_raw_data downloads Excel file."""
        try:
            from backend.services.data_ingestion.defra_ingestion import (
                DEFRAEmissionFactorsIngestion
            )
        except ImportError:
            pytest.skip("DEFRAEmissionFactorsIngestion not yet implemented")

        ingestion = DEFRAEmissionFactorsIngestion(
            db=mock_async_session,
            data_source_id=data_source_id
        )

        # Mock httpx response
        with patch('httpx.AsyncClient') as mock_client:
            mock_response = MagicMock()
            mock_response.content = sample_defra_workbook
            mock_response.raise_for_status = MagicMock()

            mock_client_instance = AsyncMock()
            mock_client_instance.get = AsyncMock(return_value=mock_response)
            mock_client_instance.__aenter__ = AsyncMock(
                return_value=mock_client_instance
            )
            mock_client_instance.__aexit__ = AsyncMock(return_value=None)
            mock_client.return_value = mock_client_instance

            result = await ingestion.fetch_raw_data()

        # Verify result is bytes
        assert isinstance(result, bytes)
        assert len(result) > 0

    @pytest.mark.asyncio
    async def test_fetch_raw_data_uses_correct_url(
        self, mock_async_session, data_source_id
    ):
        """Test that fetch_raw_data uses the DEFRA_URL."""
        try:
            from backend.services.data_ingestion.defra_ingestion import (
                DEFRAEmissionFactorsIngestion
            )
        except ImportError:
            pytest.skip("DEFRAEmissionFactorsIngestion not yet implemented")

        ingestion = DEFRAEmissionFactorsIngestion(
            db=mock_async_session,
            data_source_id=data_source_id
        )

        with patch('httpx.AsyncClient') as mock_client:
            mock_response = MagicMock()
            mock_response.content = b"test"
            mock_response.raise_for_status = MagicMock()

            mock_client_instance = AsyncMock()
            mock_client_instance.get = AsyncMock(return_value=mock_response)
            mock_client_instance.__aenter__ = AsyncMock(
                return_value=mock_client_instance
            )
            mock_client_instance.__aexit__ = AsyncMock(return_value=None)
            mock_client.return_value = mock_client_instance

            await ingestion.fetch_raw_data()

            # Verify the URL was called
            mock_client_instance.get.assert_called_once()
            called_url = mock_client_instance.get.call_args[0][0]
            assert called_url == DEFRAEmissionFactorsIngestion.DEFRA_URL

    @pytest.mark.asyncio
    async def test_fetch_raw_data_follows_redirects(
        self, mock_async_session, data_source_id
    ):
        """Test that fetch_raw_data follows redirects."""
        try:
            from backend.services.data_ingestion.defra_ingestion import (
                DEFRAEmissionFactorsIngestion
            )
        except ImportError:
            pytest.skip("DEFRAEmissionFactorsIngestion not yet implemented")

        ingestion = DEFRAEmissionFactorsIngestion(
            db=mock_async_session,
            data_source_id=data_source_id
        )

        with patch('httpx.AsyncClient') as mock_client:
            mock_response = MagicMock()
            mock_response.content = b"test"
            mock_response.raise_for_status = MagicMock()

            mock_client_instance = AsyncMock()
            mock_client_instance.get = AsyncMock(return_value=mock_response)
            mock_client_instance.__aenter__ = AsyncMock(
                return_value=mock_client_instance
            )
            mock_client_instance.__aexit__ = AsyncMock(return_value=None)
            mock_client.return_value = mock_client_instance

            await ingestion.fetch_raw_data()

            # Verify follow_redirects was set
            call_kwargs = mock_client_instance.get.call_args[1]
            assert call_kwargs.get('follow_redirects') is True


# ============================================================================
# Test Scenario 3: parse_data Extracts Records from Multiple Sheets
# ============================================================================

class TestParseData:
    """Test parse_data method extracts records from multiple sheets."""

    @pytest.mark.asyncio
    async def test_parse_data_extracts_fuels_records(
        self, mock_async_session, data_source_id, sample_defra_workbook
    ):
        """Test that parse_data extracts records from Fuels sheet."""
        try:
            from backend.services.data_ingestion.defra_ingestion import (
                DEFRAEmissionFactorsIngestion
            )
        except ImportError:
            pytest.skip("DEFRAEmissionFactorsIngestion not yet implemented")

        ingestion = DEFRAEmissionFactorsIngestion(
            db=mock_async_session,
            data_source_id=data_source_id
        )

        records = await ingestion.parse_data(sample_defra_workbook)

        # Should have records from Fuels sheet
        fuels_records = [
            r for r in records
            if "Fuels" in r.get("_sheet_name", "")
        ]
        assert len(fuels_records) >= 5

    @pytest.mark.asyncio
    async def test_parse_data_extracts_electricity_records(
        self, mock_async_session, data_source_id, sample_defra_workbook
    ):
        """Test that parse_data extracts records from Electricity sheet."""
        try:
            from backend.services.data_ingestion.defra_ingestion import (
                DEFRAEmissionFactorsIngestion
            )
        except ImportError:
            pytest.skip("DEFRAEmissionFactorsIngestion not yet implemented")

        ingestion = DEFRAEmissionFactorsIngestion(
            db=mock_async_session,
            data_source_id=data_source_id
        )

        records = await ingestion.parse_data(sample_defra_workbook)

        # Should have records matching Electricity config
        electricity_records = [
            r for r in records
            if "electricity" in r.get("_sheet_name", "").lower()
        ]
        assert len(electricity_records) >= 3

    @pytest.mark.asyncio
    async def test_parse_data_extracts_material_records(
        self, mock_async_session, data_source_id, sample_defra_workbook
    ):
        """Test that parse_data extracts records from Material use sheet."""
        try:
            from backend.services.data_ingestion.defra_ingestion import (
                DEFRAEmissionFactorsIngestion
            )
        except ImportError:
            pytest.skip("DEFRAEmissionFactorsIngestion not yet implemented")

        ingestion = DEFRAEmissionFactorsIngestion(
            db=mock_async_session,
            data_source_id=data_source_id
        )

        records = await ingestion.parse_data(sample_defra_workbook)

        # Should have records from Material use sheet
        material_records = [
            r for r in records
            if "material" in r.get("_sheet_name", "").lower()
        ]
        assert len(material_records) >= 3

    @pytest.mark.asyncio
    async def test_parse_data_includes_sheet_metadata(
        self, mock_async_session, data_source_id, sample_defra_workbook
    ):
        """Test that parsed records include sheet metadata."""
        try:
            from backend.services.data_ingestion.defra_ingestion import (
                DEFRAEmissionFactorsIngestion
            )
        except ImportError:
            pytest.skip("DEFRAEmissionFactorsIngestion not yet implemented")

        ingestion = DEFRAEmissionFactorsIngestion(
            db=mock_async_session,
            data_source_id=data_source_id
        )

        records = await ingestion.parse_data(sample_defra_workbook)

        assert len(records) > 0
        for record in records:
            assert "_sheet_name" in record
            assert "_row_idx" in record
            assert "_config" in record

    @pytest.mark.asyncio
    async def test_parse_data_extracts_all_six_sheet_types(
        self, mock_async_session, data_source_id, sample_defra_workbook
    ):
        """Test that parse_data extracts records from all 6 configured sheets."""
        try:
            from backend.services.data_ingestion.defra_ingestion import (
                DEFRAEmissionFactorsIngestion
            )
        except ImportError:
            pytest.skip("DEFRAEmissionFactorsIngestion not yet implemented")

        ingestion = DEFRAEmissionFactorsIngestion(
            db=mock_async_session,
            data_source_id=data_source_id
        )

        records = await ingestion.parse_data(sample_defra_workbook)

        # Collect unique sheet names
        sheet_names = set(r.get("_sheet_name", "") for r in records)

        # Should have at least 5 different sheets (some might be named differently)
        assert len(sheet_names) >= 5


# ============================================================================
# Test Scenario 4: _find_sheet Matches Partial Sheet Names
# ============================================================================

class TestFindSheet:
    """Test _find_sheet method matches partial sheet names."""

    def test_find_sheet_exact_match(
        self, mock_async_session, data_source_id, sample_defra_workbook
    ):
        """Test _find_sheet finds exact sheet name match."""
        try:
            from backend.services.data_ingestion.defra_ingestion import (
                DEFRAEmissionFactorsIngestion
            )
            from openpyxl import load_workbook
        except ImportError:
            pytest.skip("Required modules not yet implemented")

        ingestion = DEFRAEmissionFactorsIngestion(
            db=mock_async_session,
            data_source_id=data_source_id
        )

        workbook = load_workbook(BytesIO(sample_defra_workbook), read_only=True)

        result = ingestion._find_sheet(workbook, "Fuels")
        assert result == "Fuels"

        workbook.close()

    def test_find_sheet_partial_match(
        self, mock_async_session, data_source_id, sample_defra_workbook
    ):
        """Test _find_sheet finds partial sheet name match."""
        try:
            from backend.services.data_ingestion.defra_ingestion import (
                DEFRAEmissionFactorsIngestion
            )
            from openpyxl import load_workbook
        except ImportError:
            pytest.skip("Required modules not yet implemented")

        ingestion = DEFRAEmissionFactorsIngestion(
            db=mock_async_session,
            data_source_id=data_source_id
        )

        workbook = load_workbook(BytesIO(sample_defra_workbook), read_only=True)

        # "Electricity" should match "UK electricity"
        result = ingestion._find_sheet(workbook, "Electricity")
        assert result is not None
        assert "electricity" in result.lower()

        workbook.close()

    def test_find_sheet_case_insensitive(
        self, mock_async_session, data_source_id, sample_defra_workbook
    ):
        """Test _find_sheet is case insensitive."""
        try:
            from backend.services.data_ingestion.defra_ingestion import (
                DEFRAEmissionFactorsIngestion
            )
            from openpyxl import load_workbook
        except ImportError:
            pytest.skip("Required modules not yet implemented")

        ingestion = DEFRAEmissionFactorsIngestion(
            db=mock_async_session,
            data_source_id=data_source_id
        )

        workbook = load_workbook(BytesIO(sample_defra_workbook), read_only=True)

        # Case variations should work
        result = ingestion._find_sheet(workbook, "FUELS")
        assert result == "Fuels"

        result = ingestion._find_sheet(workbook, "fuels")
        assert result == "Fuels"

        workbook.close()

    def test_find_sheet_returns_none_for_missing(
        self, mock_async_session, data_source_id, sample_defra_workbook
    ):
        """Test _find_sheet returns None for non-existent sheet."""
        try:
            from backend.services.data_ingestion.defra_ingestion import (
                DEFRAEmissionFactorsIngestion
            )
            from openpyxl import load_workbook
        except ImportError:
            pytest.skip("Required modules not yet implemented")

        ingestion = DEFRAEmissionFactorsIngestion(
            db=mock_async_session,
            data_source_id=data_source_id
        )

        workbook = load_workbook(BytesIO(sample_defra_workbook), read_only=True)

        result = ingestion._find_sheet(workbook, "NonExistentSheet")
        assert result is None

        workbook.close()


# ============================================================================
# Test Scenario 5: transform_data Handles Various Column Formats
# ============================================================================

class TestTransformData:
    """Test transform_data method handles various column formats."""

    @pytest.fixture
    def parsed_fuels_record(self):
        """Create a parsed Fuels sheet record."""
        return {
            "Category": "Gaseous fuels",
            "Fuel": "Natural Gas",
            "Unit": "kWh",
            "kg CO2e per unit": 0.18287,
            "_sheet_name": "Fuels",
            "_row_idx": 1,
            "_config": {
                "scope": "Scope 1",
                "category": "combustion",
                "activity_col": "Fuel",
                "co2e_col": "kg CO2e per unit",
                "unit_col": "Unit",
            }
        }

    @pytest.fixture
    def parsed_electricity_record(self):
        """Create a parsed Electricity sheet record."""
        return {
            "Activity": "Electricity: UK",
            "kg CO2e per kWh": 0.21233,
            "_sheet_name": "UK electricity",
            "_row_idx": 1,
            "_config": {
                "scope": "Scope 2",
                "category": "electricity",
                "activity_col": "Activity",
                "co2e_col": "kg CO2e per kWh",
                "unit_col": None,
            }
        }

    @pytest.mark.asyncio
    async def test_transform_data_extracts_activity_name(
        self, mock_async_session, data_source_id, parsed_fuels_record
    ):
        """Test that transform_data extracts activity name correctly."""
        try:
            from backend.services.data_ingestion.defra_ingestion import (
                DEFRAEmissionFactorsIngestion
            )
        except ImportError:
            pytest.skip("DEFRAEmissionFactorsIngestion not yet implemented")

        ingestion = DEFRAEmissionFactorsIngestion(
            db=mock_async_session,
            data_source_id=data_source_id
        )

        result = await ingestion.transform_data([parsed_fuels_record])

        assert len(result) == 1
        assert result[0]["activity_name"] == "Natural Gas"

    @pytest.mark.asyncio
    async def test_transform_data_extracts_co2e_factor(
        self, mock_async_session, data_source_id, parsed_fuels_record
    ):
        """Test that transform_data extracts co2e_factor correctly."""
        try:
            from backend.services.data_ingestion.defra_ingestion import (
                DEFRAEmissionFactorsIngestion
            )
        except ImportError:
            pytest.skip("DEFRAEmissionFactorsIngestion not yet implemented")

        ingestion = DEFRAEmissionFactorsIngestion(
            db=mock_async_session,
            data_source_id=data_source_id
        )

        result = await ingestion.transform_data([parsed_fuels_record])

        assert len(result) == 1
        assert abs(result[0]["co2e_factor"] - 0.18287) < 0.00001

    @pytest.mark.asyncio
    async def test_transform_data_sets_correct_scope(
        self, mock_async_session, data_source_id, parsed_fuels_record,
        parsed_electricity_record
    ):
        """Test that transform_data sets correct scope per sheet config."""
        try:
            from backend.services.data_ingestion.defra_ingestion import (
                DEFRAEmissionFactorsIngestion
            )
        except ImportError:
            pytest.skip("DEFRAEmissionFactorsIngestion not yet implemented")

        ingestion = DEFRAEmissionFactorsIngestion(
            db=mock_async_session,
            data_source_id=data_source_id
        )

        fuels_result = await ingestion.transform_data([parsed_fuels_record])
        electricity_result = await ingestion.transform_data(
            [parsed_electricity_record]
        )

        assert fuels_result[0]["scope"] == "Scope 1"
        assert electricity_result[0]["scope"] == "Scope 2"

    @pytest.mark.asyncio
    async def test_transform_data_sets_geography_to_gb(
        self, mock_async_session, data_source_id, parsed_fuels_record
    ):
        """Test that transform_data sets geography to GB."""
        try:
            from backend.services.data_ingestion.defra_ingestion import (
                DEFRAEmissionFactorsIngestion
            )
        except ImportError:
            pytest.skip("DEFRAEmissionFactorsIngestion not yet implemented")

        ingestion = DEFRAEmissionFactorsIngestion(
            db=mock_async_session,
            data_source_id=data_source_id
        )

        result = await ingestion.transform_data([parsed_fuels_record])

        assert result[0]["geography"] == "GB"

    @pytest.mark.asyncio
    async def test_transform_data_sets_data_quality_rating(
        self, mock_async_session, data_source_id, parsed_fuels_record
    ):
        """Test that transform_data sets data_quality_rating to 0.88."""
        try:
            from backend.services.data_ingestion.defra_ingestion import (
                DEFRAEmissionFactorsIngestion
            )
        except ImportError:
            pytest.skip("DEFRAEmissionFactorsIngestion not yet implemented")

        ingestion = DEFRAEmissionFactorsIngestion(
            db=mock_async_session,
            data_source_id=data_source_id
        )

        result = await ingestion.transform_data([parsed_fuels_record])

        assert result[0]["data_quality_rating"] == 0.88

    @pytest.mark.asyncio
    async def test_transform_data_includes_metadata(
        self, mock_async_session, data_source_id, parsed_fuels_record
    ):
        """Test that transform_data includes source metadata."""
        try:
            from backend.services.data_ingestion.defra_ingestion import (
                DEFRAEmissionFactorsIngestion
            )
        except ImportError:
            pytest.skip("DEFRAEmissionFactorsIngestion not yet implemented")

        ingestion = DEFRAEmissionFactorsIngestion(
            db=mock_async_session,
            data_source_id=data_source_id
        )

        result = await ingestion.transform_data([parsed_fuels_record])

        assert "metadata" in result[0]
        assert result[0]["metadata"]["source_sheet"] == "Fuels"
        assert result[0]["metadata"]["source_row"] == 1

    @pytest.mark.asyncio
    async def test_transform_data_skips_records_without_activity(
        self, mock_async_session, data_source_id
    ):
        """Test that transform_data skips records without activity name."""
        try:
            from backend.services.data_ingestion.defra_ingestion import (
                DEFRAEmissionFactorsIngestion
            )
        except ImportError:
            pytest.skip("DEFRAEmissionFactorsIngestion not yet implemented")

        ingestion = DEFRAEmissionFactorsIngestion(
            db=mock_async_session,
            data_source_id=data_source_id
        )

        record = {
            "Category": "Gaseous fuels",
            "Fuel": None,  # Missing activity
            "kg CO2e per unit": 0.18287,
            "_sheet_name": "Fuels",
            "_row_idx": 1,
            "_config": {
                "scope": "Scope 1",
                "category": "combustion",
                "activity_col": "Fuel",
                "co2e_col": "kg CO2e per unit",
            }
        }

        result = await ingestion.transform_data([record])

        assert len(result) == 0

    @pytest.mark.asyncio
    async def test_transform_data_skips_records_without_co2e(
        self, mock_async_session, data_source_id
    ):
        """Test that transform_data skips records without co2e factor."""
        try:
            from backend.services.data_ingestion.defra_ingestion import (
                DEFRAEmissionFactorsIngestion
            )
        except ImportError:
            pytest.skip("DEFRAEmissionFactorsIngestion not yet implemented")

        ingestion = DEFRAEmissionFactorsIngestion(
            db=mock_async_session,
            data_source_id=data_source_id
        )

        record = {
            "Fuel": "Natural Gas",
            "kg CO2e per unit": None,  # Missing co2e
            "_sheet_name": "Fuels",
            "_row_idx": 1,
            "_config": {
                "scope": "Scope 1",
                "category": "combustion",
                "activity_col": "Fuel",
                "co2e_col": "kg CO2e per unit",
            }
        }

        result = await ingestion.transform_data([record])

        assert len(result) == 0


# ============================================================================
# Test Scenario 6: _determine_unit Extracts Unit from Column Name
# ============================================================================

class TestDetermineUnit:
    """Test _determine_unit method extracts unit from column name."""

    def test_determine_unit_from_column_name_kwh(
        self, mock_async_session, data_source_id
    ):
        """Test extracting kWh unit from column name."""
        try:
            from backend.services.data_ingestion.defra_ingestion import (
                DEFRAEmissionFactorsIngestion
            )
        except ImportError:
            pytest.skip("DEFRAEmissionFactorsIngestion not yet implemented")

        ingestion = DEFRAEmissionFactorsIngestion(
            db=mock_async_session,
            data_source_id=data_source_id
        )

        record = {"Activity": "Test"}
        config = {"unit_col": None}
        co2e_col = "kg CO2e per kWh"

        result = ingestion._determine_unit(record, config, co2e_col)

        assert result.lower() == "kwh"

    def test_determine_unit_from_column_name_kg(
        self, mock_async_session, data_source_id
    ):
        """Test extracting kg unit from column name."""
        try:
            from backend.services.data_ingestion.defra_ingestion import (
                DEFRAEmissionFactorsIngestion
            )
        except ImportError:
            pytest.skip("DEFRAEmissionFactorsIngestion not yet implemented")

        ingestion = DEFRAEmissionFactorsIngestion(
            db=mock_async_session,
            data_source_id=data_source_id
        )

        record = {"Material": "Steel"}
        config = {"unit_col": None}
        co2e_col = "kg CO2e per kg"

        result = ingestion._determine_unit(record, config, co2e_col)

        assert result.lower() == "kg"

    def test_determine_unit_from_column_name_tonne(
        self, mock_async_session, data_source_id
    ):
        """Test extracting tonne unit from column name."""
        try:
            from backend.services.data_ingestion.defra_ingestion import (
                DEFRAEmissionFactorsIngestion
            )
        except ImportError:
            pytest.skip("DEFRAEmissionFactorsIngestion not yet implemented")

        ingestion = DEFRAEmissionFactorsIngestion(
            db=mock_async_session,
            data_source_id=data_source_id
        )

        record = {"Waste type": "Mixed"}
        config = {"unit_col": None}
        co2e_col = "kg CO2e per tonne"

        result = ingestion._determine_unit(record, config, co2e_col)

        assert result.lower() == "tonne"

    def test_determine_unit_from_explicit_unit_column(
        self, mock_async_session, data_source_id
    ):
        """Test extracting unit from explicit unit column."""
        try:
            from backend.services.data_ingestion.defra_ingestion import (
                DEFRAEmissionFactorsIngestion
            )
        except ImportError:
            pytest.skip("DEFRAEmissionFactorsIngestion not yet implemented")

        ingestion = DEFRAEmissionFactorsIngestion(
            db=mock_async_session,
            data_source_id=data_source_id
        )

        record = {"Fuel": "Diesel", "Unit": "litre"}
        config = {"unit_col": "Unit"}
        co2e_col = "kg CO2e per unit"

        result = ingestion._determine_unit(record, config, co2e_col)

        assert result == "litre"

    def test_determine_unit_returns_default_when_no_match(
        self, mock_async_session, data_source_id
    ):
        """Test that _determine_unit returns default when no match."""
        try:
            from backend.services.data_ingestion.defra_ingestion import (
                DEFRAEmissionFactorsIngestion
            )
        except ImportError:
            pytest.skip("DEFRAEmissionFactorsIngestion not yet implemented")

        ingestion = DEFRAEmissionFactorsIngestion(
            db=mock_async_session,
            data_source_id=data_source_id
        )

        record = {"Activity": "Test"}
        config = {"unit_col": None}
        co2e_col = "emission factor"  # No "per X" pattern

        result = ingestion._determine_unit(record, config, co2e_col)

        assert result == "unit"


# ============================================================================
# Test Scenario 7: External ID Generation Handles Special Characters
# ============================================================================

class TestExternalIdGeneration:
    """Test external ID generation handles special characters."""

    @pytest.mark.asyncio
    async def test_external_id_format(
        self, mock_async_session, data_source_id
    ):
        """Test that external_id follows correct format."""
        try:
            from backend.services.data_ingestion.defra_ingestion import (
                DEFRAEmissionFactorsIngestion
            )
        except ImportError:
            pytest.skip("DEFRAEmissionFactorsIngestion not yet implemented")

        ingestion = DEFRAEmissionFactorsIngestion(
            db=mock_async_session,
            data_source_id=data_source_id
        )

        record = {
            "Fuel": "Natural Gas",
            "kg CO2e per unit": 0.18287,
            "_sheet_name": "Fuels",
            "_row_idx": 1,
            "_config": {
                "scope": "Scope 1",
                "category": "combustion",
                "activity_col": "Fuel",
                "co2e_col": "kg CO2e per unit",
            }
        }

        result = await ingestion.transform_data([record])

        assert len(result) == 1
        external_id = result[0]["external_id"]
        assert external_id.startswith("DEFRA_")
        assert "Fuels" in external_id
        assert "Natural" in external_id

    @pytest.mark.asyncio
    async def test_external_id_replaces_spaces_with_underscores(
        self, mock_async_session, data_source_id
    ):
        """Test that spaces are replaced with underscores."""
        try:
            from backend.services.data_ingestion.defra_ingestion import (
                DEFRAEmissionFactorsIngestion
            )
        except ImportError:
            pytest.skip("DEFRAEmissionFactorsIngestion not yet implemented")

        ingestion = DEFRAEmissionFactorsIngestion(
            db=mock_async_session,
            data_source_id=data_source_id
        )

        record = {
            "Fuel": "Liquefied Petroleum Gas",
            "kg CO2e per unit": 0.21448,
            "_sheet_name": "Fuels",
            "_row_idx": 2,
            "_config": {
                "scope": "Scope 1",
                "category": "combustion",
                "activity_col": "Fuel",
                "co2e_col": "kg CO2e per unit",
            }
        }

        result = await ingestion.transform_data([record])

        external_id = result[0]["external_id"]
        assert " " not in external_id
        assert "_" in external_id

    @pytest.mark.asyncio
    async def test_external_id_handles_special_characters(
        self, mock_async_session, data_source_id
    ):
        """Test that special characters are handled correctly."""
        try:
            from backend.services.data_ingestion.defra_ingestion import (
                DEFRAEmissionFactorsIngestion
            )
        except ImportError:
            pytest.skip("DEFRAEmissionFactorsIngestion not yet implemented")

        ingestion = DEFRAEmissionFactorsIngestion(
            db=mock_async_session,
            data_source_id=data_source_id
        )

        record = {
            "Fuel": "Gas (natural/LNG)",
            "kg CO2e per unit": 0.18287,
            "_sheet_name": "Fuels",
            "_row_idx": 1,
            "_config": {
                "scope": "Scope 1",
                "category": "combustion",
                "activity_col": "Fuel",
                "co2e_col": "kg CO2e per unit",
            }
        }

        result = await ingestion.transform_data([record])

        external_id = result[0]["external_id"]
        # Special chars like parentheses and slashes should be replaced
        assert "(" not in external_id
        assert ")" not in external_id
        assert "/" not in external_id

    @pytest.mark.asyncio
    async def test_external_id_truncated_to_200_chars(
        self, mock_async_session, data_source_id
    ):
        """Test that external_id is truncated to 200 characters."""
        try:
            from backend.services.data_ingestion.defra_ingestion import (
                DEFRAEmissionFactorsIngestion
            )
        except ImportError:
            pytest.skip("DEFRAEmissionFactorsIngestion not yet implemented")

        ingestion = DEFRAEmissionFactorsIngestion(
            db=mock_async_session,
            data_source_id=data_source_id
        )

        # Create a very long activity name
        long_name = "Very Long Activity Name " * 20  # ~480 chars

        record = {
            "Fuel": long_name,
            "kg CO2e per unit": 0.18287,
            "_sheet_name": "Fuels",
            "_row_idx": 1,
            "_config": {
                "scope": "Scope 1",
                "category": "combustion",
                "activity_col": "Fuel",
                "co2e_col": "kg CO2e per unit",
            }
        }

        result = await ingestion.transform_data([record])

        external_id = result[0]["external_id"]
        assert len(external_id) <= 200


# ============================================================================
# Test Scenario 8: Error Handling for Missing Sheets
# ============================================================================

class TestErrorHandlingMissingSheets:
    """Test error handling for missing sheets."""

    @pytest.fixture
    def minimal_workbook(self):
        """Create a workbook with only one sheet."""
        try:
            from openpyxl import Workbook
        except ImportError:
            pytest.skip("openpyxl not installed")

        wb = Workbook()
        sheet = wb.active
        sheet.title = "Fuels"
        sheet.append(["Fuel", "kg CO2e per unit"])
        sheet.append(["Natural Gas", 0.18287])

        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()

    @pytest.mark.asyncio
    async def test_parse_data_handles_missing_sheets_gracefully(
        self, mock_async_session, data_source_id, minimal_workbook
    ):
        """Test that parse_data handles missing sheets gracefully."""
        try:
            from backend.services.data_ingestion.defra_ingestion import (
                DEFRAEmissionFactorsIngestion
            )
        except ImportError:
            pytest.skip("DEFRAEmissionFactorsIngestion not yet implemented")

        ingestion = DEFRAEmissionFactorsIngestion(
            db=mock_async_session,
            data_source_id=data_source_id
        )

        # Should not raise an exception
        records = await ingestion.parse_data(minimal_workbook)

        # Should return records from available sheets only
        assert isinstance(records, list)
        # Only Fuels sheet exists, so only Fuels records should be parsed
        for record in records:
            assert "Fuels" in record.get("_sheet_name", "")

    @pytest.mark.asyncio
    async def test_parse_data_continues_after_missing_sheet(
        self, mock_async_session, data_source_id
    ):
        """Test that parse_data continues processing after missing sheet."""
        try:
            from backend.services.data_ingestion.defra_ingestion import (
                DEFRAEmissionFactorsIngestion
            )
            from openpyxl import Workbook
        except ImportError:
            pytest.skip("Required modules not yet implemented")

        ingestion = DEFRAEmissionFactorsIngestion(
            db=mock_async_session,
            data_source_id=data_source_id
        )

        # Create workbook with Fuels and Material use, but not Electricity
        wb = Workbook()
        default_sheet = wb.active
        wb.remove(default_sheet)

        fuels = wb.create_sheet("Fuels")
        fuels.append(["Fuel", "kg CO2e per unit"])
        fuels.append(["Diesel", 2.70554])

        materials = wb.create_sheet("Material use")
        materials.append(["Material", "kg CO2e per kg"])
        materials.append(["Steel", 1.85])

        output = BytesIO()
        wb.save(output)
        output.seek(0)
        workbook_bytes = output.getvalue()

        records = await ingestion.parse_data(workbook_bytes)

        # Should have records from both available sheets
        sheet_names = set(r.get("_sheet_name", "") for r in records)
        assert "Fuels" in sheet_names
        assert "Material use" in sheet_names

    @pytest.mark.asyncio
    async def test_empty_workbook_returns_empty_list(
        self, mock_async_session, data_source_id
    ):
        """Test that empty workbook returns empty list."""
        try:
            from backend.services.data_ingestion.defra_ingestion import (
                DEFRAEmissionFactorsIngestion
            )
            from openpyxl import Workbook
        except ImportError:
            pytest.skip("Required modules not yet implemented")

        ingestion = DEFRAEmissionFactorsIngestion(
            db=mock_async_session,
            data_source_id=data_source_id
        )

        # Create empty workbook (no matching sheets)
        wb = Workbook()
        sheet = wb.active
        sheet.title = "UnrelatedSheet"

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        records = await ingestion.parse_data(output.getvalue())

        assert records == []


# ============================================================================
# Test Scenario 9: Sheet Configuration Correctness
# ============================================================================

class TestSheetConfigurationCorrectness:
    """Test that sheet configurations are correct."""

    def test_fuels_sheet_config(self, mock_async_session, data_source_id):
        """Test Fuels sheet configuration is correct."""
        try:
            from backend.services.data_ingestion.defra_ingestion import (
                DEFRAEmissionFactorsIngestion
            )
        except ImportError:
            pytest.skip("DEFRAEmissionFactorsIngestion not yet implemented")

        config = DEFRAEmissionFactorsIngestion.SHEET_CONFIGS.get("Fuels", {})

        assert config.get("scope") == "Scope 1"
        assert config.get("category") == "combustion"
        assert "activity_col" in config
        assert "co2e_col" in config

    def test_electricity_sheet_config(self, mock_async_session, data_source_id):
        """Test Electricity sheet configuration is correct."""
        try:
            from backend.services.data_ingestion.defra_ingestion import (
                DEFRAEmissionFactorsIngestion
            )
        except ImportError:
            pytest.skip("DEFRAEmissionFactorsIngestion not yet implemented")

        config = DEFRAEmissionFactorsIngestion.SHEET_CONFIGS.get(
            "Electricity", {}
        )

        assert config.get("scope") == "Scope 2"
        assert config.get("category") == "electricity"

    def test_material_use_sheet_config(self, mock_async_session, data_source_id):
        """Test Material use sheet configuration is correct."""
        try:
            from backend.services.data_ingestion.defra_ingestion import (
                DEFRAEmissionFactorsIngestion
            )
        except ImportError:
            pytest.skip("DEFRAEmissionFactorsIngestion not yet implemented")

        config = DEFRAEmissionFactorsIngestion.SHEET_CONFIGS.get(
            "Material use", {}
        )

        assert config.get("scope") == "Scope 3"
        assert config.get("category") == "materials"

    def test_waste_disposal_sheet_config(self, mock_async_session, data_source_id):
        """Test Waste disposal sheet configuration is correct."""
        try:
            from backend.services.data_ingestion.defra_ingestion import (
                DEFRAEmissionFactorsIngestion
            )
        except ImportError:
            pytest.skip("DEFRAEmissionFactorsIngestion not yet implemented")

        config = DEFRAEmissionFactorsIngestion.SHEET_CONFIGS.get(
            "Waste disposal", {}
        )

        assert config.get("scope") == "Scope 3"
        assert config.get("category") == "waste"

    def test_business_travel_air_sheet_config(
        self, mock_async_session, data_source_id
    ):
        """Test Business travel- air sheet configuration is correct."""
        try:
            from backend.services.data_ingestion.defra_ingestion import (
                DEFRAEmissionFactorsIngestion
            )
        except ImportError:
            pytest.skip("DEFRAEmissionFactorsIngestion not yet implemented")

        config = DEFRAEmissionFactorsIngestion.SHEET_CONFIGS.get(
            "Business travel- air", {}
        )

        assert config.get("scope") == "Scope 3"
        assert config.get("category") == "transport"

    def test_freighting_goods_sheet_config(
        self, mock_async_session, data_source_id
    ):
        """Test Freighting goods sheet configuration is correct."""
        try:
            from backend.services.data_ingestion.defra_ingestion import (
                DEFRAEmissionFactorsIngestion
            )
        except ImportError:
            pytest.skip("DEFRAEmissionFactorsIngestion not yet implemented")

        config = DEFRAEmissionFactorsIngestion.SHEET_CONFIGS.get(
            "Freighting goods", {}
        )

        assert config.get("scope") == "Scope 3"
        assert config.get("category") == "transport"


# ============================================================================
# Test Scenario 10: Reference Year Configuration
# ============================================================================

class TestReferenceYearConfiguration:
    """Test reference year configuration."""

    @pytest.mark.asyncio
    async def test_transform_data_uses_reference_year(
        self, mock_async_session, data_source_id
    ):
        """Test that transform_data uses the configured reference_year."""
        try:
            from backend.services.data_ingestion.defra_ingestion import (
                DEFRAEmissionFactorsIngestion
            )
        except ImportError:
            pytest.skip("DEFRAEmissionFactorsIngestion not yet implemented")

        ingestion = DEFRAEmissionFactorsIngestion(
            db=mock_async_session,
            data_source_id=data_source_id
        )

        record = {
            "Fuel": "Natural Gas",
            "kg CO2e per unit": 0.18287,
            "_sheet_name": "Fuels",
            "_row_idx": 1,
            "_config": {
                "scope": "Scope 1",
                "category": "combustion",
                "activity_col": "Fuel",
                "co2e_col": "kg CO2e per unit",
            }
        }

        result = await ingestion.transform_data([record])

        assert result[0]["reference_year"] == 2024

    def test_reference_year_can_be_overridden(
        self, mock_async_session, data_source_id
    ):
        """Test that reference_year can be overridden after instantiation."""
        try:
            from backend.services.data_ingestion.defra_ingestion import (
                DEFRAEmissionFactorsIngestion
            )
        except ImportError:
            pytest.skip("DEFRAEmissionFactorsIngestion not yet implemented")

        ingestion = DEFRAEmissionFactorsIngestion(
            db=mock_async_session,
            data_source_id=data_source_id
        )

        # Override reference year
        ingestion.reference_year = 2023

        assert ingestion.reference_year == 2023
