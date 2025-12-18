"""
Integration tests for DEFRA connector URL fixes.

TASK-DATA-P7-008: Fix DEFRA Connector URL

This test suite validates:
- DEFRA URL is accessible (BUG-DATA-002 fix)
- DEFRA file contains expected sheets
- DEFRA_URL constant has correct CDN URL

BUG-DATA-002 (CRITICAL): DEFRA URL returns 404
Root Cause: UK Government Digital Service migrated content to a new CDN structure.
The old /government/uploads/system/uploads/attachment_data/ path no longer resolves.
The new path uses /media/{content-id}/ format.

Test-Driven Development Protocol:
- These tests are written BEFORE implementation fixes
- Tests verify the CORRECT values that should be in the code
- Implementation must make tests PASS without modifying tests
"""

import pytest
from unittest.mock import AsyncMock, MagicMock
from uuid import uuid4

from sqlalchemy.ext.asyncio import AsyncSession


@pytest.fixture(scope="function")
def mock_async_session():
    """Create mock async session for unit tests."""
    session = AsyncMock(spec=AsyncSession)
    session.add = MagicMock()
    session.flush = AsyncMock()
    session.commit = AsyncMock()
    session.rollback = AsyncMock()
    session.execute = AsyncMock()
    session.refresh = AsyncMock()
    return session


@pytest.fixture
def data_source_id():
    """Generate a test data source ID."""
    return uuid4().hex


# ============================================================================
# TASK-DATA-P7-008: Integration Tests for DEFRA URL Accessibility
# ============================================================================

class TestDEFRAURLAccessibility:
    """
    Integration tests for DEFRA data source URL accessibility.

    TASK-DATA-P7-008: Fix DEFRA Connector URL

    These tests verify that the DEFRA URL is accessible and returns a valid
    Excel file with the expected sheet names.
    """

    # The correct URL for DEFRA GHG Conversion Factors 2024
    CORRECT_DEFRA_URL = (
        "https://assets.publishing.service.gov.uk/media/6722567487df31a87d8c497e/"
        "ghg-conversion-factors-2024-full_set__for_advanced_users__v1_1.xlsx"
    )

    @pytest.mark.integration
    @pytest.mark.asyncio
    async def test_defra_url_accessible(self):
        """
        Verify DEFRA URL returns 200 OK.

        BUG-DATA-002 (CRITICAL): DEFRA URL 404
        Expected URL: https://assets.publishing.service.gov.uk/media/6722567487df31a87d8c497e/
                      ghg-conversion-factors-2024-full_set__for_advanced_users__v1_1.xlsx
        """
        import httpx

        async with httpx.AsyncClient(timeout=120.0, follow_redirects=True) as client:
            response = await client.head(self.CORRECT_DEFRA_URL)
            assert response.status_code == 200, (
                f"DEFRA URL returned {response.status_code}. "
                f"URL: {self.CORRECT_DEFRA_URL}"
            )

    @pytest.mark.integration
    @pytest.mark.asyncio
    async def test_defra_contains_expected_sheets(self):
        """
        Verify DEFRA file contains expected sheets.

        DEFRA conversion factors file should contain sheets for:
        - Fuels (Scope 1)
        - Electricity (Scope 2)
        - Material use (Scope 3)
        - Waste disposal (Scope 3)
        - Business travel (Scope 3)
        - Freighting goods (Scope 3)
        """
        import httpx
        from openpyxl import load_workbook
        import io

        async with httpx.AsyncClient(timeout=120.0, follow_redirects=True) as client:
            response = await client.get(self.CORRECT_DEFRA_URL)
            response.raise_for_status()

        workbook = load_workbook(io.BytesIO(response.content), read_only=True)
        sheet_names = workbook.sheetnames
        workbook.close()

        # DEFRA sheets should include patterns for Fuels, Electricity
        expected_patterns = ["Fuel", "Electricity"]
        for pattern in expected_patterns:
            found = any(pattern.lower() in s.lower() for s in sheet_names)
            assert found, (
                f"No sheet matching '{pattern}' found in DEFRA file. "
                f"Available sheets: {sheet_names}"
            )

    @pytest.mark.integration
    @pytest.mark.asyncio
    async def test_defra_file_is_valid_excel(self):
        """
        Verify DEFRA URL returns a valid Excel file.

        This test downloads the file and verifies it can be parsed by openpyxl.
        """
        import httpx
        from openpyxl import load_workbook
        import io

        async with httpx.AsyncClient(timeout=120.0, follow_redirects=True) as client:
            response = await client.get(self.CORRECT_DEFRA_URL)
            response.raise_for_status()

        # Verify the file can be loaded as Excel
        workbook = load_workbook(io.BytesIO(response.content), read_only=True)
        sheet_names = workbook.sheetnames
        workbook.close()

        # Should have multiple sheets
        assert len(sheet_names) > 5, (
            f"DEFRA file should have multiple sheets, got {len(sheet_names)}: {sheet_names}"
        )

    @pytest.mark.integration
    @pytest.mark.asyncio
    async def test_defra_file_content_type_is_excel(self):
        """
        Verify DEFRA URL returns correct content type for Excel.
        """
        import httpx

        async with httpx.AsyncClient(timeout=120.0, follow_redirects=True) as client:
            response = await client.head(self.CORRECT_DEFRA_URL)
            response.raise_for_status()

        content_type = response.headers.get("content-type", "")
        # Excel files may have various content types
        valid_excel_types = [
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            "application/vnd.ms-excel",
            "application/octet-stream",  # Some CDNs return this
            "binary/octet-stream",
        ]
        assert any(t in content_type for t in valid_excel_types), (
            f"DEFRA URL should return Excel content type, got: {content_type}"
        )


# ============================================================================
# TASK-DATA-P7-008: Tests for DEFRA_URL Constant Correctness
# ============================================================================

class TestDEFRAURLConstantCorrectness:
    """
    Tests to verify the DEFRA_URL constant in DEFRAEmissionFactorsIngestion
    contains the correct CDN URL.

    TASK-DATA-P7-008: Fix DEFRA Connector URL
    """

    # The correct URL for DEFRA GHG Conversion Factors 2024
    CORRECT_DEFRA_URL = (
        "https://assets.publishing.service.gov.uk/media/6722567487df31a87d8c497e/"
        "ghg-conversion-factors-2024-full_set__for_advanced_users__v1_1.xlsx"
    )

    def test_defra_url_constant_is_correct(self, mock_async_session, data_source_id):
        """
        Test that DEFRA_URL constant has the correct CDN URL.

        BUG-DATA-002 fix: URL should point to new CDN location with media/ path
        instead of old government/uploads/system/uploads/attachment_data/ path.
        """
        try:
            from backend.services.data_ingestion.defra_ingestion import (
                DEFRAEmissionFactorsIngestion
            )
        except ImportError:
            pytest.skip("DEFRAEmissionFactorsIngestion not yet implemented")

        assert DEFRAEmissionFactorsIngestion.DEFRA_URL == self.CORRECT_DEFRA_URL, (
            f"DEFRA_URL is incorrect. "
            f"Expected: {self.CORRECT_DEFRA_URL}, "
            f"Got: {DEFRAEmissionFactorsIngestion.DEFRA_URL}"
        )

    def test_defra_url_uses_new_cdn_path(self, mock_async_session, data_source_id):
        """
        Test that DEFRA_URL uses new /media/ CDN path format.

        After BUG-DATA-002 fix, URL should use /media/{content-id}/ format
        instead of old /government/uploads/system/uploads/attachment_data/ path.
        """
        try:
            from backend.services.data_ingestion.defra_ingestion import (
                DEFRAEmissionFactorsIngestion
            )
        except ImportError:
            pytest.skip("DEFRAEmissionFactorsIngestion not yet implemented")

        # Verify the URL uses new /media/ CDN path
        assert "/media/" in DEFRAEmissionFactorsIngestion.DEFRA_URL, (
            f"DEFRA_URL should use /media/ CDN path format, "
            f"got: {DEFRAEmissionFactorsIngestion.DEFRA_URL}"
        )

    def test_defra_url_does_not_use_old_path(self, mock_async_session, data_source_id):
        """
        Test that DEFRA_URL does NOT use old government uploads path.

        The old path /government/uploads/system/uploads/attachment_data/
        no longer works and returns 404.
        """
        try:
            from backend.services.data_ingestion.defra_ingestion import (
                DEFRAEmissionFactorsIngestion
            )
        except ImportError:
            pytest.skip("DEFRAEmissionFactorsIngestion not yet implemented")

        # Verify the URL does NOT use old path patterns
        old_patterns = [
            "/government/uploads/",
            "/attachment_data/",
            "system/uploads/",
        ]
        for pattern in old_patterns:
            assert pattern not in DEFRAEmissionFactorsIngestion.DEFRA_URL, (
                f"DEFRA_URL should NOT contain old path pattern '{pattern}', "
                f"got: {DEFRAEmissionFactorsIngestion.DEFRA_URL}"
            )

    def test_defra_url_contains_correct_filename(self, mock_async_session, data_source_id):
        """
        Test that DEFRA_URL contains correct 2024 filename.

        The filename should be:
        ghg-conversion-factors-2024-full_set__for_advanced_users__v1_1.xlsx
        """
        try:
            from backend.services.data_ingestion.defra_ingestion import (
                DEFRAEmissionFactorsIngestion
            )
        except ImportError:
            pytest.skip("DEFRAEmissionFactorsIngestion not yet implemented")

        expected_filename = "ghg-conversion-factors-2024-full_set__for_advanced_users__v1_1.xlsx"
        assert expected_filename in DEFRAEmissionFactorsIngestion.DEFRA_URL, (
            f"DEFRA_URL should contain filename '{expected_filename}', "
            f"got: {DEFRAEmissionFactorsIngestion.DEFRA_URL}"
        )

    def test_defra_url_uses_assets_publishing_domain(self, mock_async_session, data_source_id):
        """
        Test that DEFRA_URL uses assets.publishing.service.gov.uk domain.
        """
        try:
            from backend.services.data_ingestion.defra_ingestion import (
                DEFRAEmissionFactorsIngestion
            )
        except ImportError:
            pytest.skip("DEFRAEmissionFactorsIngestion not yet implemented")

        expected_domain = "assets.publishing.service.gov.uk"
        assert expected_domain in DEFRAEmissionFactorsIngestion.DEFRA_URL, (
            f"DEFRA_URL should use domain '{expected_domain}', "
            f"got: {DEFRAEmissionFactorsIngestion.DEFRA_URL}"
        )


# ============================================================================
# TASK-DATA-P7-008: Regression Test for Old URL
# ============================================================================

class TestDEFRAOldURLRegression:
    """
    Regression tests to verify the old DEFRA URL is indeed broken.

    This confirms that BUG-DATA-002 is a real issue and the fix is necessary.
    """

    @pytest.mark.integration
    @pytest.mark.asyncio
    async def test_old_defra_url_returns_404(self):
        """
        Verify old DEFRA URL returns 404 (confirming BUG-DATA-002).

        This test documents the broken old URL for regression purposes.
        """
        import httpx

        # The OLD broken URL
        old_url = (
            "https://www.gov.uk/government/uploads/system/uploads/"
            "attachment_data/file/ghg-conversion-factors-2024.xlsx"
        )

        async with httpx.AsyncClient(timeout=60.0, follow_redirects=True) as client:
            response = await client.head(old_url)
            # Old URL should NOT be accessible (404 or similar error)
            assert response.status_code != 200, (
                f"Old DEFRA URL unexpectedly returned {response.status_code}. "
                f"This suggests the old URL might be working again. "
                f"URL: {old_url}"
            )

    @pytest.mark.integration
    @pytest.mark.asyncio
    async def test_old_assets_url_pattern_returns_404(self):
        """
        Verify old assets.publishing URL pattern returns 404.

        This is the URL pattern that was in the code before the fix.
        """
        import httpx

        # The OLD broken URL with assets.publishing domain
        old_url = (
            "https://assets.publishing.service.gov.uk/government/uploads/system/uploads/"
            "attachment_data/file/ghg-conversion-factors-2024.xlsx"
        )

        async with httpx.AsyncClient(timeout=60.0, follow_redirects=True) as client:
            response = await client.head(old_url)
            # Old URL should NOT be accessible (404 or similar error)
            assert response.status_code != 200, (
                f"Old assets DEFRA URL unexpectedly returned {response.status_code}. "
                f"This suggests the old URL might be working again. "
                f"URL: {old_url}"
            )
