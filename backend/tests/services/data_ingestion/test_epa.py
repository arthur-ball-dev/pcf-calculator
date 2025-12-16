"""
Test suite for EPAEmissionFactorsIngestion class.

TASK-DATA-P5-002: EPA Data Connector - Phase A Tests

This test suite validates:
- Constructor initializes correctly with file_key (fuels, egrid)
- fetch_raw_data downloads from correct URL
- parse_data extracts records from Excel sheets
- transform_data for fuel records (correct fields and structure)
- transform_data for eGRID records (unit conversion lb/MWh to kg/kWh)
- _determine_scope returns correct scope based on category
- Error handling for malformed Excel file

Test-Driven Development Protocol:
- These tests MUST be committed BEFORE implementation
- Tests should FAIL initially (no EPAEmissionFactorsIngestion class exists yet)
- Implementation must make tests PASS without modifying tests
"""

import pytest
from unittest.mock import AsyncMock, MagicMock, patch
from datetime import datetime, timezone
from decimal import Decimal
from uuid import uuid4
from io import BytesIO

from sqlalchemy import create_engine, text
from sqlalchemy.orm import sessionmaker
from sqlalchemy.pool import StaticPool
from sqlalchemy.ext.asyncio import AsyncSession


@pytest.fixture(scope="function")
def mock_async_session():
    """Create mock async session for unit tests."""
    session = AsyncMock(spec=AsyncSession)
    session.add = MagicMock()
    session.flush = AsyncMock()
    session.commit = AsyncMock()
    session.rollback = AsyncMock()
    session.execute = AsyncMock()
    session.refresh = AsyncMock()
    return session


@pytest.fixture
def data_source_id():
    """Generate a test data source ID."""
    return uuid4().hex


@pytest.fixture
def sample_fuel_excel_bytes():
    """
    Create a minimal sample EPA fuel emission factors Excel file.

    This fixture creates a valid Excel file in memory using openpyxl
    with the "Table 1 - Fuel" sheet containing sample fuel records.
    """
    try:
        from openpyxl import Workbook
    except ImportError:
        pytest.skip("openpyxl not available")

    wb = Workbook()

    # Create "Table 1 - Fuel" sheet
    ws = wb.active
    ws.title = "Table 1 - Fuel"

    # Header row
    headers = ["Fuel Type", "kg CO2e per unit", "Unit", "Category"]
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)

    # Sample fuel data
    fuel_data = [
        ("Natural Gas", 2.75, "kg", "Stationary Combustion"),
        ("Diesel", 10.21, "L", "Stationary Combustion"),
        ("Gasoline", 8.89, "L", "Mobile"),
        ("Propane", 6.35, "kg", "Stationary Combustion"),
        ("Residual Fuel Oil", 11.27, "L", "Stationary Combustion"),
        ("Coal (Bituminous)", 2.563, "kg", "Stationary Combustion"),
        ("Kerosene", 9.75, "L", "Stationary Combustion"),
        ("Aviation Gasoline", 8.31, "L", "Mobile - Aviation"),
    ]

    for row_idx, (fuel_type, factor, unit, category) in enumerate(fuel_data, 2):
        ws.cell(row=row_idx, column=1, value=fuel_type)
        ws.cell(row=row_idx, column=2, value=factor)
        ws.cell(row=row_idx, column=3, value=unit)
        ws.cell(row=row_idx, column=4, value=category)

    # Save to bytes
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


@pytest.fixture
def sample_egrid_excel_bytes():
    """
    Create a minimal sample EPA eGRID Excel file.

    This fixture creates a valid Excel file in memory with eGRID
    subregion data for electricity emission factors.
    """
    try:
        from openpyxl import Workbook
    except ImportError:
        pytest.skip("openpyxl not available")

    wb = Workbook()

    # Create "SUBRGN22" sheet (subregion data)
    ws = wb.active
    ws.title = "SUBRGN22"

    # Header row - eGRID uses specific column names
    headers = ["SUBRGN", "SRCO2RTA", "SRNOXRTA", "SRSO2RTA"]
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)

    # Sample eGRID subregion data (SRCO2RTA is in lb/MWh)
    egrid_data = [
        ("AKGD", 1098.5, 2.3, 1.2),  # Alaska Grid
        ("CAMX", 531.2, 0.8, 0.4),   # WECC California
        ("ERCT", 925.3, 1.5, 0.9),   # ERCOT
        ("RFCW", 1245.7, 2.1, 1.8),  # RFC West
        ("SRSO", 1012.8, 1.9, 1.1),  # SERC South
    ]

    for row_idx, (subrgn, co2_rate, nox_rate, so2_rate) in enumerate(egrid_data, 2):
        ws.cell(row=row_idx, column=1, value=subrgn)
        ws.cell(row=row_idx, column=2, value=co2_rate)
        ws.cell(row=row_idx, column=3, value=nox_rate)
        ws.cell(row=row_idx, column=4, value=so2_rate)

    # Save to bytes
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


@pytest.fixture
def malformed_excel_bytes():
    """Create a malformed/empty Excel file for error testing."""
    try:
        from openpyxl import Workbook
    except ImportError:
        pytest.skip("openpyxl not available")

    wb = Workbook()
    ws = wb.active
    ws.title = "Wrong Sheet Name"
    # Empty sheet - no data

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


# ============================================================================
# Test Scenario 1: Constructor Initialization
# ============================================================================

class TestEPAIngestionInstantiation:
    """Test EPAEmissionFactorsIngestion class instantiation."""

    def test_constructor_initializes_with_default_file_key(
        self, mock_async_session, data_source_id
    ):
        """Test that constructor defaults to 'fuels' file_key."""
        try:
            from backend.services.data_ingestion.epa_ingestion import (
                EPAEmissionFactorsIngestion
            )
        except ImportError:
            pytest.skip("EPAEmissionFactorsIngestion not yet implemented")

        ingestion = EPAEmissionFactorsIngestion(
            db=mock_async_session,
            data_source_id=data_source_id
        )

        assert ingestion.file_key == "fuels"
        assert ingestion.file_config is not None
        assert ingestion.file_config["type"] == "combustion"

    def test_constructor_initializes_with_fuels_file_key(
        self, mock_async_session, data_source_id
    ):
        """Test that constructor accepts 'fuels' file_key explicitly."""
        try:
            from backend.services.data_ingestion.epa_ingestion import (
                EPAEmissionFactorsIngestion
            )
        except ImportError:
            pytest.skip("EPAEmissionFactorsIngestion not yet implemented")

        ingestion = EPAEmissionFactorsIngestion(
            db=mock_async_session,
            data_source_id=data_source_id,
            file_key="fuels"
        )

        assert ingestion.file_key == "fuels"
        assert "url" in ingestion.file_config
        assert "sheets" in ingestion.file_config

    def test_constructor_initializes_with_egrid_file_key(
        self, mock_async_session, data_source_id
    ):
        """Test that constructor accepts 'egrid' file_key."""
        try:
            from backend.services.data_ingestion.epa_ingestion import (
                EPAEmissionFactorsIngestion
            )
        except ImportError:
            pytest.skip("EPAEmissionFactorsIngestion not yet implemented")

        ingestion = EPAEmissionFactorsIngestion(
            db=mock_async_session,
            data_source_id=data_source_id,
            file_key="egrid"
        )

        assert ingestion.file_key == "egrid"
        assert ingestion.file_config["type"] == "electricity"

    def test_constructor_raises_error_for_invalid_file_key(
        self, mock_async_session, data_source_id
    ):
        """Test that constructor raises KeyError for invalid file_key."""
        try:
            from backend.services.data_ingestion.epa_ingestion import (
                EPAEmissionFactorsIngestion
            )
        except ImportError:
            pytest.skip("EPAEmissionFactorsIngestion not yet implemented")

        with pytest.raises(KeyError):
            EPAEmissionFactorsIngestion(
                db=mock_async_session,
                data_source_id=data_source_id,
                file_key="invalid_key"
            )

    def test_constructor_inherits_from_base_data_ingestion(
        self, mock_async_session, data_source_id
    ):
        """Test that EPAEmissionFactorsIngestion inherits from BaseDataIngestion."""
        try:
            from backend.services.data_ingestion.epa_ingestion import (
                EPAEmissionFactorsIngestion
            )
            from backend.services.data_ingestion.base import BaseDataIngestion
        except ImportError:
            pytest.skip("Required modules not yet implemented")

        ingestion = EPAEmissionFactorsIngestion(
            db=mock_async_session,
            data_source_id=data_source_id
        )

        assert isinstance(ingestion, BaseDataIngestion)
        assert hasattr(ingestion, 'db')
        assert hasattr(ingestion, 'data_source_id')
        assert hasattr(ingestion, 'stats')
        assert hasattr(ingestion, 'errors')


# ============================================================================
# Test Scenario 2: fetch_raw_data Downloads from Correct URL
# ============================================================================

class TestFetchRawData:
    """Test fetch_raw_data downloads from correct EPA URL."""

    @pytest.mark.asyncio
    async def test_fetch_raw_data_uses_correct_fuels_url(
        self, mock_async_session, data_source_id
    ):
        """Test that fetch_raw_data uses the correct URL for fuels."""
        try:
            from backend.services.data_ingestion.epa_ingestion import (
                EPAEmissionFactorsIngestion
            )
        except ImportError:
            pytest.skip("EPAEmissionFactorsIngestion not yet implemented")

        ingestion = EPAEmissionFactorsIngestion(
            db=mock_async_session,
            data_source_id=data_source_id,
            file_key="fuels"
        )

        # Verify the URL is correctly configured
        expected_url_pattern = "emission-factors"
        assert expected_url_pattern in ingestion.file_config["url"]
        assert ingestion.file_config["url"].startswith("https://")

    @pytest.mark.asyncio
    async def test_fetch_raw_data_uses_correct_egrid_url(
        self, mock_async_session, data_source_id
    ):
        """Test that fetch_raw_data uses the correct URL for eGRID."""
        try:
            from backend.services.data_ingestion.epa_ingestion import (
                EPAEmissionFactorsIngestion
            )
        except ImportError:
            pytest.skip("EPAEmissionFactorsIngestion not yet implemented")

        ingestion = EPAEmissionFactorsIngestion(
            db=mock_async_session,
            data_source_id=data_source_id,
            file_key="egrid"
        )

        # Verify the URL is correctly configured for eGRID
        expected_url_pattern = "egrid"
        assert expected_url_pattern in ingestion.file_config["url"].lower()

    @pytest.mark.asyncio
    async def test_fetch_raw_data_returns_bytes(
        self, mock_async_session, data_source_id, sample_fuel_excel_bytes
    ):
        """Test that fetch_raw_data returns bytes."""
        try:
            import respx
            import httpx
            from backend.services.data_ingestion.epa_ingestion import (
                EPAEmissionFactorsIngestion
            )
        except ImportError:
            pytest.skip("Required modules not yet implemented")

        ingestion = EPAEmissionFactorsIngestion(
            db=mock_async_session,
            data_source_id=data_source_id,
            file_key="fuels"
        )

        with respx.mock:
            respx.get(ingestion.file_config["url"]).mock(
                return_value=httpx.Response(200, content=sample_fuel_excel_bytes)
            )

            result = await ingestion.fetch_raw_data()

            assert isinstance(result, bytes)
            assert len(result) > 0

    @pytest.mark.asyncio
    async def test_fetch_raw_data_raises_on_http_error(
        self, mock_async_session, data_source_id
    ):
        """Test that fetch_raw_data raises on HTTP error."""
        try:
            import respx
            import httpx
            from backend.services.data_ingestion.epa_ingestion import (
                EPAEmissionFactorsIngestion
            )
        except ImportError:
            pytest.skip("Required modules not yet implemented")

        ingestion = EPAEmissionFactorsIngestion(
            db=mock_async_session,
            data_source_id=data_source_id,
            file_key="fuels"
        )

        with respx.mock:
            respx.get(ingestion.file_config["url"]).mock(
                return_value=httpx.Response(404)
            )

            with pytest.raises(httpx.HTTPStatusError):
                await ingestion.fetch_raw_data()


# ============================================================================
# Test Scenario 3: parse_data Extracts Records from Excel Sheets
# ============================================================================

class TestParseData:
    """Test parse_data extracts records from Excel sheets."""

    @pytest.mark.asyncio
    async def test_parse_data_extracts_fuel_records(
        self, mock_async_session, data_source_id, sample_fuel_excel_bytes
    ):
        """Test that parse_data extracts records from fuel Excel file."""
        try:
            from backend.services.data_ingestion.epa_ingestion import (
                EPAEmissionFactorsIngestion
            )
        except ImportError:
            pytest.skip("EPAEmissionFactorsIngestion not yet implemented")

        ingestion = EPAEmissionFactorsIngestion(
            db=mock_async_session,
            data_source_id=data_source_id,
            file_key="fuels"
        )

        records = await ingestion.parse_data(sample_fuel_excel_bytes)

        assert isinstance(records, list)
        assert len(records) == 8  # 8 fuel records in fixture

        # Each record should have the header columns
        first_record = records[0]
        assert "Fuel Type" in first_record
        assert "kg CO2e per unit" in first_record
        assert "Unit" in first_record

    @pytest.mark.asyncio
    async def test_parse_data_extracts_egrid_records(
        self, mock_async_session, data_source_id, sample_egrid_excel_bytes
    ):
        """Test that parse_data extracts records from eGRID Excel file."""
        try:
            from backend.services.data_ingestion.epa_ingestion import (
                EPAEmissionFactorsIngestion
            )
        except ImportError:
            pytest.skip("EPAEmissionFactorsIngestion not yet implemented")

        ingestion = EPAEmissionFactorsIngestion(
            db=mock_async_session,
            data_source_id=data_source_id,
            file_key="egrid"
        )

        records = await ingestion.parse_data(sample_egrid_excel_bytes)

        assert isinstance(records, list)
        assert len(records) == 5  # 5 subregion records in fixture

        # Each record should have eGRID columns
        first_record = records[0]
        assert "SUBRGN" in first_record
        assert "SRCO2RTA" in first_record

    @pytest.mark.asyncio
    async def test_parse_data_includes_source_metadata(
        self, mock_async_session, data_source_id, sample_fuel_excel_bytes
    ):
        """Test that parse_data includes source sheet and row metadata."""
        try:
            from backend.services.data_ingestion.epa_ingestion import (
                EPAEmissionFactorsIngestion
            )
        except ImportError:
            pytest.skip("EPAEmissionFactorsIngestion not yet implemented")

        ingestion = EPAEmissionFactorsIngestion(
            db=mock_async_session,
            data_source_id=data_source_id,
            file_key="fuels"
        )

        records = await ingestion.parse_data(sample_fuel_excel_bytes)

        first_record = records[0]
        assert "_source_sheet" in first_record
        assert "_source_row" in first_record
        assert first_record["_source_sheet"] == "Table 1 - Fuel"

    @pytest.mark.asyncio
    async def test_parse_data_handles_empty_sheet(
        self, mock_async_session, data_source_id, malformed_excel_bytes
    ):
        """Test that parse_data handles Excel with no matching sheets."""
        try:
            from backend.services.data_ingestion.epa_ingestion import (
                EPAEmissionFactorsIngestion
            )
        except ImportError:
            pytest.skip("EPAEmissionFactorsIngestion not yet implemented")

        ingestion = EPAEmissionFactorsIngestion(
            db=mock_async_session,
            data_source_id=data_source_id,
            file_key="fuels"
        )

        records = await ingestion.parse_data(malformed_excel_bytes)

        # Should return empty list for missing sheets
        assert isinstance(records, list)
        assert len(records) == 0

    @pytest.mark.asyncio
    async def test_parse_data_handles_corrupted_file(
        self, mock_async_session, data_source_id
    ):
        """Test that parse_data raises error for corrupted Excel file."""
        try:
            from backend.services.data_ingestion.epa_ingestion import (
                EPAEmissionFactorsIngestion
            )
        except ImportError:
            pytest.skip("EPAEmissionFactorsIngestion not yet implemented")

        ingestion = EPAEmissionFactorsIngestion(
            db=mock_async_session,
            data_source_id=data_source_id,
            file_key="fuels"
        )

        # Invalid Excel bytes
        corrupted_data = b"this is not a valid excel file"

        with pytest.raises(Exception):  # openpyxl raises various exceptions
            await ingestion.parse_data(corrupted_data)


# ============================================================================
# Test Scenario 4: transform_data for Fuel Records
# ============================================================================

class TestTransformDataFuel:
    """Test transform_data for fuel emission factor records."""

    @pytest.mark.asyncio
    async def test_transform_fuel_record_extracts_activity_name(
        self, mock_async_session, data_source_id
    ):
        """Test that fuel records have correct activity_name."""
        try:
            from backend.services.data_ingestion.epa_ingestion import (
                EPAEmissionFactorsIngestion
            )
        except ImportError:
            pytest.skip("EPAEmissionFactorsIngestion not yet implemented")

        ingestion = EPAEmissionFactorsIngestion(
            db=mock_async_session,
            data_source_id=data_source_id,
            file_key="fuels"
        )

        parsed_data = [
            {
                "Fuel Type": "Natural Gas",
                "kg CO2e per unit": 2.75,
                "Unit": "kg",
                "Category": "Stationary Combustion",
                "_source_sheet": "Table 1 - Fuel",
                "_source_row": 2
            }
        ]

        transformed = await ingestion.transform_data(parsed_data)

        assert len(transformed) == 1
        assert transformed[0]["activity_name"] == "Natural Gas"

    @pytest.mark.asyncio
    async def test_transform_fuel_record_extracts_co2e_factor(
        self, mock_async_session, data_source_id
    ):
        """Test that fuel records have correct co2e_factor."""
        try:
            from backend.services.data_ingestion.epa_ingestion import (
                EPAEmissionFactorsIngestion
            )
        except ImportError:
            pytest.skip("EPAEmissionFactorsIngestion not yet implemented")

        ingestion = EPAEmissionFactorsIngestion(
            db=mock_async_session,
            data_source_id=data_source_id,
            file_key="fuels"
        )

        parsed_data = [
            {
                "Fuel Type": "Diesel",
                "kg CO2e per unit": 10.21,
                "Unit": "L",
                "Category": "Stationary Combustion",
                "_source_sheet": "Table 1 - Fuel",
                "_source_row": 3
            }
        ]

        transformed = await ingestion.transform_data(parsed_data)

        assert len(transformed) == 1
        assert float(transformed[0]["co2e_factor"]) == 10.21

    @pytest.mark.asyncio
    async def test_transform_fuel_record_includes_required_fields(
        self, mock_async_session, data_source_id
    ):
        """Test that transformed fuel records include all required fields."""
        try:
            from backend.services.data_ingestion.epa_ingestion import (
                EPAEmissionFactorsIngestion
            )
        except ImportError:
            pytest.skip("EPAEmissionFactorsIngestion not yet implemented")

        ingestion = EPAEmissionFactorsIngestion(
            db=mock_async_session,
            data_source_id=data_source_id,
            file_key="fuels"
        )

        parsed_data = [
            {
                "Fuel Type": "Propane",
                "kg CO2e per unit": 6.35,
                "Unit": "kg",
                "Category": "Stationary Combustion",
                "_source_sheet": "Table 1 - Fuel",
                "_source_row": 4
            }
        ]

        transformed = await ingestion.transform_data(parsed_data)

        assert len(transformed) == 1
        record = transformed[0]

        # Required fields from schema
        assert "activity_name" in record
        assert "co2e_factor" in record
        assert "unit" in record
        assert "external_id" in record
        assert "geography" in record
        assert "scope" in record
        assert "category" in record
        assert "data_quality_rating" in record

    @pytest.mark.asyncio
    async def test_transform_fuel_record_generates_external_id(
        self, mock_async_session, data_source_id
    ):
        """Test that fuel records have correctly formatted external_id."""
        try:
            from backend.services.data_ingestion.epa_ingestion import (
                EPAEmissionFactorsIngestion
            )
        except ImportError:
            pytest.skip("EPAEmissionFactorsIngestion not yet implemented")

        ingestion = EPAEmissionFactorsIngestion(
            db=mock_async_session,
            data_source_id=data_source_id,
            file_key="fuels"
        )

        parsed_data = [
            {
                "Fuel Type": "Natural Gas",
                "kg CO2e per unit": 2.75,
                "Unit": "kg",
                "Category": "Stationary Combustion",
                "_source_sheet": "Table 1 - Fuel",
                "_source_row": 2
            }
        ]

        transformed = await ingestion.transform_data(parsed_data)

        assert "external_id" in transformed[0]
        assert "EPA" in transformed[0]["external_id"]
        assert "fuels" in transformed[0]["external_id"]
        # External ID should not have spaces
        assert " " not in transformed[0]["external_id"]

    @pytest.mark.asyncio
    async def test_transform_fuel_record_skips_missing_fuel_type(
        self, mock_async_session, data_source_id
    ):
        """Test that records without Fuel Type are skipped."""
        try:
            from backend.services.data_ingestion.epa_ingestion import (
                EPAEmissionFactorsIngestion
            )
        except ImportError:
            pytest.skip("EPAEmissionFactorsIngestion not yet implemented")

        ingestion = EPAEmissionFactorsIngestion(
            db=mock_async_session,
            data_source_id=data_source_id,
            file_key="fuels"
        )

        parsed_data = [
            {
                "Fuel Type": None,  # Missing
                "kg CO2e per unit": 2.75,
                "Unit": "kg",
                "_source_sheet": "Table 1 - Fuel",
                "_source_row": 2
            }
        ]

        transformed = await ingestion.transform_data(parsed_data)

        assert len(transformed) == 0

    @pytest.mark.asyncio
    async def test_transform_fuel_record_skips_missing_co2e_factor(
        self, mock_async_session, data_source_id
    ):
        """Test that records without CO2e factor are skipped."""
        try:
            from backend.services.data_ingestion.epa_ingestion import (
                EPAEmissionFactorsIngestion
            )
        except ImportError:
            pytest.skip("EPAEmissionFactorsIngestion not yet implemented")

        ingestion = EPAEmissionFactorsIngestion(
            db=mock_async_session,
            data_source_id=data_source_id,
            file_key="fuels"
        )

        parsed_data = [
            {
                "Fuel Type": "Natural Gas",
                "kg CO2e per unit": None,  # Missing
                "Unit": "kg",
                "_source_sheet": "Table 1 - Fuel",
                "_source_row": 2
            }
        ]

        transformed = await ingestion.transform_data(parsed_data)

        assert len(transformed) == 0

    @pytest.mark.asyncio
    async def test_transform_fuel_record_handles_invalid_co2e_factor(
        self, mock_async_session, data_source_id
    ):
        """Test that records with invalid CO2e factor are skipped."""
        try:
            from backend.services.data_ingestion.epa_ingestion import (
                EPAEmissionFactorsIngestion
            )
        except ImportError:
            pytest.skip("EPAEmissionFactorsIngestion not yet implemented")

        ingestion = EPAEmissionFactorsIngestion(
            db=mock_async_session,
            data_source_id=data_source_id,
            file_key="fuels"
        )

        parsed_data = [
            {
                "Fuel Type": "Natural Gas",
                "kg CO2e per unit": "not a number",  # Invalid
                "Unit": "kg",
                "_source_sheet": "Table 1 - Fuel",
                "_source_row": 2
            }
        ]

        transformed = await ingestion.transform_data(parsed_data)

        assert len(transformed) == 0

    @pytest.mark.asyncio
    async def test_transform_multiple_fuel_records(
        self, mock_async_session, data_source_id
    ):
        """Test that multiple fuel records are transformed correctly."""
        try:
            from backend.services.data_ingestion.epa_ingestion import (
                EPAEmissionFactorsIngestion
            )
        except ImportError:
            pytest.skip("EPAEmissionFactorsIngestion not yet implemented")

        ingestion = EPAEmissionFactorsIngestion(
            db=mock_async_session,
            data_source_id=data_source_id,
            file_key="fuels"
        )

        parsed_data = [
            {
                "Fuel Type": "Natural Gas",
                "kg CO2e per unit": 2.75,
                "Unit": "kg",
                "Category": "Stationary Combustion",
                "_source_sheet": "Table 1 - Fuel",
                "_source_row": 2
            },
            {
                "Fuel Type": "Diesel",
                "kg CO2e per unit": 10.21,
                "Unit": "L",
                "Category": "Stationary Combustion",
                "_source_sheet": "Table 1 - Fuel",
                "_source_row": 3
            },
            {
                "Fuel Type": "Gasoline",
                "kg CO2e per unit": 8.89,
                "Unit": "L",
                "Category": "Mobile",
                "_source_sheet": "Table 1 - Fuel",
                "_source_row": 4
            }
        ]

        transformed = await ingestion.transform_data(parsed_data)

        assert len(transformed) == 3
        activity_names = [r["activity_name"] for r in transformed]
        assert "Natural Gas" in activity_names
        assert "Diesel" in activity_names
        assert "Gasoline" in activity_names


# ============================================================================
# Test Scenario 5: transform_data for eGRID Records (Unit Conversion)
# ============================================================================

class TestTransformDataEGRID:
    """Test transform_data for eGRID electricity emission factor records."""

    @pytest.mark.asyncio
    async def test_transform_egrid_record_converts_lb_per_mwh_to_kg_per_kwh(
        self, mock_async_session, data_source_id
    ):
        """Test that eGRID records convert lb/MWh to kg/kWh correctly."""
        try:
            from backend.services.data_ingestion.epa_ingestion import (
                EPAEmissionFactorsIngestion
            )
        except ImportError:
            pytest.skip("EPAEmissionFactorsIngestion not yet implemented")

        ingestion = EPAEmissionFactorsIngestion(
            db=mock_async_session,
            data_source_id=data_source_id,
            file_key="egrid"
        )

        parsed_data = [
            {
                "SUBRGN": "CAMX",
                "SRCO2RTA": 1000.0,  # 1000 lb/MWh
                "_source_sheet": "SUBRGN22",
                "_source_row": 2
            }
        ]

        transformed = await ingestion.transform_data(parsed_data)

        assert len(transformed) == 1
        # Conversion: 1000 lb/MWh * 0.453592 kg/lb / 1000 kWh/MWh = 0.453592 kg/kWh
        expected_factor = 1000.0 * 0.453592 / 1000
        assert abs(float(transformed[0]["co2e_factor"]) - expected_factor) < 0.001

    @pytest.mark.asyncio
    async def test_transform_egrid_record_unit_is_kg_co2e_per_kwh(
        self, mock_async_session, data_source_id
    ):
        """Test that eGRID records have correct unit after conversion."""
        try:
            from backend.services.data_ingestion.epa_ingestion import (
                EPAEmissionFactorsIngestion
            )
        except ImportError:
            pytest.skip("EPAEmissionFactorsIngestion not yet implemented")

        ingestion = EPAEmissionFactorsIngestion(
            db=mock_async_session,
            data_source_id=data_source_id,
            file_key="egrid"
        )

        parsed_data = [
            {
                "SUBRGN": "ERCT",
                "SRCO2RTA": 925.3,
                "_source_sheet": "SUBRGN22",
                "_source_row": 3
            }
        ]

        transformed = await ingestion.transform_data(parsed_data)

        assert len(transformed) == 1
        assert transformed[0]["unit"] == "kg CO2e/kWh"

    @pytest.mark.asyncio
    async def test_transform_egrid_record_activity_name_includes_subregion(
        self, mock_async_session, data_source_id
    ):
        """Test that eGRID activity name includes subregion code."""
        try:
            from backend.services.data_ingestion.epa_ingestion import (
                EPAEmissionFactorsIngestion
            )
        except ImportError:
            pytest.skip("EPAEmissionFactorsIngestion not yet implemented")

        ingestion = EPAEmissionFactorsIngestion(
            db=mock_async_session,
            data_source_id=data_source_id,
            file_key="egrid"
        )

        parsed_data = [
            {
                "SUBRGN": "RFCW",
                "SRCO2RTA": 1245.7,
                "_source_sheet": "SUBRGN22",
                "_source_row": 4
            }
        ]

        transformed = await ingestion.transform_data(parsed_data)

        assert len(transformed) == 1
        assert "RFCW" in transformed[0]["activity_name"]
        assert "Electricity" in transformed[0]["activity_name"] or "Grid" in transformed[0]["activity_name"]

    @pytest.mark.asyncio
    async def test_transform_egrid_record_scope_is_scope_2(
        self, mock_async_session, data_source_id
    ):
        """Test that eGRID records have Scope 2 designation."""
        try:
            from backend.services.data_ingestion.epa_ingestion import (
                EPAEmissionFactorsIngestion
            )
        except ImportError:
            pytest.skip("EPAEmissionFactorsIngestion not yet implemented")

        ingestion = EPAEmissionFactorsIngestion(
            db=mock_async_session,
            data_source_id=data_source_id,
            file_key="egrid"
        )

        parsed_data = [
            {
                "SUBRGN": "SRSO",
                "SRCO2RTA": 1012.8,
                "_source_sheet": "SUBRGN22",
                "_source_row": 5
            }
        ]

        transformed = await ingestion.transform_data(parsed_data)

        assert len(transformed) == 1
        assert transformed[0]["scope"] == "Scope 2"

    @pytest.mark.asyncio
    async def test_transform_egrid_record_category_is_electricity(
        self, mock_async_session, data_source_id
    ):
        """Test that eGRID records have 'electricity' category."""
        try:
            from backend.services.data_ingestion.epa_ingestion import (
                EPAEmissionFactorsIngestion
            )
        except ImportError:
            pytest.skip("EPAEmissionFactorsIngestion not yet implemented")

        ingestion = EPAEmissionFactorsIngestion(
            db=mock_async_session,
            data_source_id=data_source_id,
            file_key="egrid"
        )

        parsed_data = [
            {
                "SUBRGN": "AKGD",
                "SRCO2RTA": 1098.5,
                "_source_sheet": "SUBRGN22",
                "_source_row": 2
            }
        ]

        transformed = await ingestion.transform_data(parsed_data)

        assert len(transformed) == 1
        assert transformed[0]["category"] == "electricity"

    @pytest.mark.asyncio
    async def test_transform_egrid_record_data_quality_rating(
        self, mock_async_session, data_source_id
    ):
        """Test that eGRID records have high data quality rating."""
        try:
            from backend.services.data_ingestion.epa_ingestion import (
                EPAEmissionFactorsIngestion
            )
        except ImportError:
            pytest.skip("EPAEmissionFactorsIngestion not yet implemented")

        ingestion = EPAEmissionFactorsIngestion(
            db=mock_async_session,
            data_source_id=data_source_id,
            file_key="egrid"
        )

        parsed_data = [
            {
                "SUBRGN": "CAMX",
                "SRCO2RTA": 531.2,
                "_source_sheet": "SUBRGN22",
                "_source_row": 2
            }
        ]

        transformed = await ingestion.transform_data(parsed_data)

        assert len(transformed) == 1
        # eGRID is a high-quality source
        assert float(transformed[0]["data_quality_rating"]) >= 0.90

    @pytest.mark.asyncio
    async def test_transform_egrid_record_external_id_format(
        self, mock_async_session, data_source_id
    ):
        """Test that eGRID records have correctly formatted external_id."""
        try:
            from backend.services.data_ingestion.epa_ingestion import (
                EPAEmissionFactorsIngestion
            )
        except ImportError:
            pytest.skip("EPAEmissionFactorsIngestion not yet implemented")

        ingestion = EPAEmissionFactorsIngestion(
            db=mock_async_session,
            data_source_id=data_source_id,
            file_key="egrid"
        )

        parsed_data = [
            {
                "SUBRGN": "ERCT",
                "SRCO2RTA": 925.3,
                "_source_sheet": "SUBRGN22",
                "_source_row": 3
            }
        ]

        transformed = await ingestion.transform_data(parsed_data)

        assert len(transformed) == 1
        assert "EPA" in transformed[0]["external_id"]
        assert "eGRID" in transformed[0]["external_id"]
        assert "ERCT" in transformed[0]["external_id"]

    @pytest.mark.asyncio
    async def test_transform_egrid_record_skips_missing_subregion(
        self, mock_async_session, data_source_id
    ):
        """Test that eGRID records without subregion are skipped."""
        try:
            from backend.services.data_ingestion.epa_ingestion import (
                EPAEmissionFactorsIngestion
            )
        except ImportError:
            pytest.skip("EPAEmissionFactorsIngestion not yet implemented")

        ingestion = EPAEmissionFactorsIngestion(
            db=mock_async_session,
            data_source_id=data_source_id,
            file_key="egrid"
        )

        parsed_data = [
            {
                "SUBRGN": None,  # Missing
                "SRCO2RTA": 925.3,
                "_source_sheet": "SUBRGN22",
                "_source_row": 3
            }
        ]

        transformed = await ingestion.transform_data(parsed_data)

        assert len(transformed) == 0

    @pytest.mark.asyncio
    async def test_transform_multiple_egrid_records(
        self, mock_async_session, data_source_id
    ):
        """Test that multiple eGRID records are transformed correctly."""
        try:
            from backend.services.data_ingestion.epa_ingestion import (
                EPAEmissionFactorsIngestion
            )
        except ImportError:
            pytest.skip("EPAEmissionFactorsIngestion not yet implemented")

        ingestion = EPAEmissionFactorsIngestion(
            db=mock_async_session,
            data_source_id=data_source_id,
            file_key="egrid"
        )

        parsed_data = [
            {
                "SUBRGN": "AKGD",
                "SRCO2RTA": 1098.5,
                "_source_sheet": "SUBRGN22",
                "_source_row": 2
            },
            {
                "SUBRGN": "CAMX",
                "SRCO2RTA": 531.2,
                "_source_sheet": "SUBRGN22",
                "_source_row": 3
            },
            {
                "SUBRGN": "ERCT",
                "SRCO2RTA": 925.3,
                "_source_sheet": "SUBRGN22",
                "_source_row": 4
            }
        ]

        transformed = await ingestion.transform_data(parsed_data)

        assert len(transformed) == 3
        subregions = [r["external_id"] for r in transformed]
        assert any("AKGD" in s for s in subregions)
        assert any("CAMX" in s for s in subregions)
        assert any("ERCT" in s for s in subregions)


# ============================================================================
# Test Scenario 6: _determine_scope Returns Correct Scope
# ============================================================================

class TestDetermineScope:
    """Test _determine_scope method returns correct GHG Protocol scope."""

    def test_determine_scope_returns_scope_2_for_electricity(
        self, mock_async_session, data_source_id
    ):
        """Test that electricity category returns Scope 2."""
        try:
            from backend.services.data_ingestion.epa_ingestion import (
                EPAEmissionFactorsIngestion
            )
        except ImportError:
            pytest.skip("EPAEmissionFactorsIngestion not yet implemented")

        ingestion = EPAEmissionFactorsIngestion(
            db=mock_async_session,
            data_source_id=data_source_id,
            file_key="fuels"
        )

        record = {"Category": "Electricity"}
        scope = ingestion._determine_scope(record)

        assert scope == "Scope 2"

    def test_determine_scope_returns_scope_1_for_mobile(
        self, mock_async_session, data_source_id
    ):
        """Test that mobile/transport category returns Scope 1."""
        try:
            from backend.services.data_ingestion.epa_ingestion import (
                EPAEmissionFactorsIngestion
            )
        except ImportError:
            pytest.skip("EPAEmissionFactorsIngestion not yet implemented")

        ingestion = EPAEmissionFactorsIngestion(
            db=mock_async_session,
            data_source_id=data_source_id,
            file_key="fuels"
        )

        record_mobile = {"Category": "Mobile"}
        record_transport = {"Category": "Transportation"}

        assert ingestion._determine_scope(record_mobile) == "Scope 1"
        assert ingestion._determine_scope(record_transport) == "Scope 1"

    def test_determine_scope_returns_scope_1_for_stationary_combustion(
        self, mock_async_session, data_source_id
    ):
        """Test that stationary combustion returns Scope 1."""
        try:
            from backend.services.data_ingestion.epa_ingestion import (
                EPAEmissionFactorsIngestion
            )
        except ImportError:
            pytest.skip("EPAEmissionFactorsIngestion not yet implemented")

        ingestion = EPAEmissionFactorsIngestion(
            db=mock_async_session,
            data_source_id=data_source_id,
            file_key="fuels"
        )

        record = {"Category": "Stationary Combustion"}
        scope = ingestion._determine_scope(record)

        assert scope == "Scope 1"

    def test_determine_scope_returns_scope_1_for_unknown_category(
        self, mock_async_session, data_source_id
    ):
        """Test that unknown categories default to Scope 1."""
        try:
            from backend.services.data_ingestion.epa_ingestion import (
                EPAEmissionFactorsIngestion
            )
        except ImportError:
            pytest.skip("EPAEmissionFactorsIngestion not yet implemented")

        ingestion = EPAEmissionFactorsIngestion(
            db=mock_async_session,
            data_source_id=data_source_id,
            file_key="fuels"
        )

        record = {"Category": "Unknown Category"}
        scope = ingestion._determine_scope(record)

        assert scope == "Scope 1"

    def test_determine_scope_handles_missing_category(
        self, mock_async_session, data_source_id
    ):
        """Test that missing category defaults to Scope 1."""
        try:
            from backend.services.data_ingestion.epa_ingestion import (
                EPAEmissionFactorsIngestion
            )
        except ImportError:
            pytest.skip("EPAEmissionFactorsIngestion not yet implemented")

        ingestion = EPAEmissionFactorsIngestion(
            db=mock_async_session,
            data_source_id=data_source_id,
            file_key="fuels"
        )

        record = {}  # No Category field
        scope = ingestion._determine_scope(record)

        assert scope == "Scope 1"

    def test_determine_scope_case_insensitive(
        self, mock_async_session, data_source_id
    ):
        """Test that scope determination is case-insensitive."""
        try:
            from backend.services.data_ingestion.epa_ingestion import (
                EPAEmissionFactorsIngestion
            )
        except ImportError:
            pytest.skip("EPAEmissionFactorsIngestion not yet implemented")

        ingestion = EPAEmissionFactorsIngestion(
            db=mock_async_session,
            data_source_id=data_source_id,
            file_key="fuels"
        )

        record_upper = {"Category": "ELECTRICITY"}
        record_lower = {"Category": "electricity"}
        record_mixed = {"Category": "ElEcTrIcItY"}

        assert ingestion._determine_scope(record_upper) == "Scope 2"
        assert ingestion._determine_scope(record_lower) == "Scope 2"
        assert ingestion._determine_scope(record_mixed) == "Scope 2"


# ============================================================================
# Test Scenario 7: Error Handling for Malformed Files
# ============================================================================

class TestErrorHandling:
    """Test error handling for malformed Excel files."""

    @pytest.mark.asyncio
    async def test_parse_raises_error_for_non_excel_content(
        self, mock_async_session, data_source_id
    ):
        """Test that parse_data raises error for non-Excel content."""
        try:
            from backend.services.data_ingestion.epa_ingestion import (
                EPAEmissionFactorsIngestion
            )
        except ImportError:
            pytest.skip("EPAEmissionFactorsIngestion not yet implemented")

        ingestion = EPAEmissionFactorsIngestion(
            db=mock_async_session,
            data_source_id=data_source_id,
            file_key="fuels"
        )

        # Plain text content
        invalid_content = b"This is not an Excel file"

        with pytest.raises(Exception):
            await ingestion.parse_data(invalid_content)

    @pytest.mark.asyncio
    async def test_parse_raises_error_for_empty_bytes(
        self, mock_async_session, data_source_id
    ):
        """Test that parse_data raises error for empty bytes."""
        try:
            from backend.services.data_ingestion.epa_ingestion import (
                EPAEmissionFactorsIngestion
            )
        except ImportError:
            pytest.skip("EPAEmissionFactorsIngestion not yet implemented")

        ingestion = EPAEmissionFactorsIngestion(
            db=mock_async_session,
            data_source_id=data_source_id,
            file_key="fuels"
        )

        empty_content = b""

        with pytest.raises(Exception):
            await ingestion.parse_data(empty_content)

    @pytest.mark.asyncio
    async def test_transform_handles_non_numeric_co2e_gracefully(
        self, mock_async_session, data_source_id
    ):
        """Test that transform skips records with non-numeric co2e values."""
        try:
            from backend.services.data_ingestion.epa_ingestion import (
                EPAEmissionFactorsIngestion
            )
        except ImportError:
            pytest.skip("EPAEmissionFactorsIngestion not yet implemented")

        ingestion = EPAEmissionFactorsIngestion(
            db=mock_async_session,
            data_source_id=data_source_id,
            file_key="fuels"
        )

        parsed_data = [
            {
                "Fuel Type": "Natural Gas",
                "kg CO2e per unit": "N/A",  # Non-numeric
                "Unit": "kg",
                "_source_sheet": "Table 1 - Fuel",
                "_source_row": 2
            },
            {
                "Fuel Type": "Diesel",
                "kg CO2e per unit": 10.21,  # Valid
                "Unit": "L",
                "_source_sheet": "Table 1 - Fuel",
                "_source_row": 3
            }
        ]

        # Should not raise - just skip invalid record
        transformed = await ingestion.transform_data(parsed_data)

        assert len(transformed) == 1
        assert transformed[0]["activity_name"] == "Diesel"

    @pytest.mark.asyncio
    async def test_files_constant_exists_with_required_keys(
        self, mock_async_session, data_source_id
    ):
        """Test that FILES constant has required keys and structure."""
        try:
            from backend.services.data_ingestion.epa_ingestion import (
                EPAEmissionFactorsIngestion
            )
        except ImportError:
            pytest.skip("EPAEmissionFactorsIngestion not yet implemented")

        # Check FILES constant exists and has expected structure
        assert hasattr(EPAEmissionFactorsIngestion, 'FILES')

        files = EPAEmissionFactorsIngestion.FILES

        # Should have fuels and egrid keys
        assert "fuels" in files
        assert "egrid" in files

        # Each file config should have required fields
        for key in ["fuels", "egrid"]:
            assert "url" in files[key]
            assert "sheets" in files[key]
            assert "type" in files[key]
            assert isinstance(files[key]["sheets"], list)
            assert files[key]["url"].startswith("https://")
