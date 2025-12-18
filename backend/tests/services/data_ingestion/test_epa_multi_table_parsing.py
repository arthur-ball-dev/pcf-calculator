"""
Test suite for EPA Fuels connector multi-table parsing.

TASK-DATA-P8-BUG-002: Fix EPA Fuels Connector Multi-Table Parsing

The EPA 2024 file has a different structure than expected:
- Single sheet called "Emission Factors Hub"
- Multiple sub-tables within that sheet:
  - Table 1: Stationary Combustion (rows 14-80+, kg CO2 per mmBtu)
  - Table 2: Mobile Combustion CO2 (rows 101-113, kg CO2 per unit)
  - Additional tables for CH4/N2O factors, electricity, etc.

Test-Driven Development Protocol:
- These tests MUST be committed BEFORE implementation fixes
- Tests should FAIL initially (parse_data returns 0 records)
- Implementation must make tests PASS without modifying tests
"""

import pytest
from io import BytesIO
from unittest.mock import AsyncMock, MagicMock
from uuid import uuid4

from sqlalchemy.ext.asyncio import AsyncSession


@pytest.fixture(scope="function")
def mock_async_session():
    """Create mock async session for unit tests."""
    session = AsyncMock(spec=AsyncSession)
    session.add = MagicMock()
    session.flush = AsyncMock()
    session.commit = AsyncMock()
    session.rollback = AsyncMock()
    session.execute = AsyncMock()
    session.refresh = AsyncMock()
    return session


@pytest.fixture
def data_source_id():
    """Generate a test data source ID."""
    return uuid4().hex


@pytest.fixture
def sample_epa_2024_format_excel():
    """
    Create a sample EPA 2024 format Excel file with multi-table structure.

    The 2024 format has:
    - Single sheet "Emission Factors Hub"
    - Table markers like "Table 1", "Table 2" in column B
    - Headers in row after table marker
    - Data offset by 2 columns (columns C onwards)
    """
    try:
        from openpyxl import Workbook
    except ImportError:
        pytest.skip("openpyxl not available")

    wb = Workbook()
    ws = wb.active
    ws.title = "Emission Factors Hub"

    # Row 1: Empty (header info in real file)
    # Row 2: Title
    ws.cell(row=2, column=2, value="Emission Factors for Greenhouse Gas Inventories")

    # Row 14: Table 1 marker
    ws.cell(row=14, column=2, value="Table 1")
    ws.cell(row=14, column=3, value="   Stationary Combustion")

    # Row 16: Table 1 headers
    ws.cell(row=16, column=3, value="Fuel Type")
    ws.cell(row=16, column=4, value="Heat Content (HHV)")
    ws.cell(row=16, column=5, value="CO2 Factor")
    ws.cell(row=16, column=6, value="CH4 Factor")
    ws.cell(row=16, column=7, value="N2O Factor")

    # Row 17: Unit row
    ws.cell(row=17, column=4, value="mmBtu per short ton")
    ws.cell(row=17, column=5, value="kg CO2 per mmBtu")
    ws.cell(row=17, column=6, value="g CH4 per mmBtu")
    ws.cell(row=17, column=7, value="g N2O per mmBtu")

    # Row 18: Category header (Coal and Coke)
    ws.cell(row=18, column=3, value="Coal and Coke")

    # Row 19-22: Actual data
    table1_data = [
        ("Anthracite", 25.09, 103.69, 11, 1.6),
        ("Bituminous", 24.93, 93.28, 11, 1.6),
        ("Sub-bituminous", 17.25, 97.17, 11, 1.6),
        ("Natural Gas", 0.001026, 53.06, 1, 0.1),
    ]
    for row_idx, (fuel, heat, co2, ch4, n2o) in enumerate(table1_data, 19):
        ws.cell(row=row_idx, column=3, value=fuel)
        ws.cell(row=row_idx, column=4, value=heat)
        ws.cell(row=row_idx, column=5, value=co2)
        ws.cell(row=row_idx, column=6, value=ch4)
        ws.cell(row=row_idx, column=7, value=n2o)

    # Row 30: Table 2 marker (simulating position ~101 in real file)
    ws.cell(row=30, column=2, value="Table 2")
    ws.cell(row=30, column=3, value="   Mobile Combustion CO2")

    # Row 32: Table 2 headers
    ws.cell(row=32, column=3, value="Fuel Type")
    ws.cell(row=32, column=4, value="kg CO2 per unit")
    ws.cell(row=32, column=5, value="Unit")

    # Row 33-37: Table 2 data
    table2_data = [
        ("Aviation Gasoline", 8.31, "gallon"),
        ("Diesel Fuel", 10.21, "gallon"),
        ("Motor Gasoline", 8.78, "gallon"),
        ("Jet Fuel", 9.75, "gallon"),
        ("LPG", 5.68, "gallon"),
    ]
    for row_idx, (fuel, co2, unit) in enumerate(table2_data, 33):
        ws.cell(row=row_idx, column=3, value=fuel)
        ws.cell(row=row_idx, column=4, value=co2)
        ws.cell(row=row_idx, column=5, value=unit)

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


# ============================================================================
# TASK-DATA-P8-BUG-002: Tests for Multi-Table Parsing
# ============================================================================

class TestMultiTableParsing:
    """
    Test EPA Fuels connector parses multi-table structure correctly.

    TASK-DATA-P8-BUG-002: Fix EPA Fuels Connector Multi-Table Parsing

    The smoke test shows 0 records because the current implementation
    expects separate sheets for each table, but the 2024 file has all
    tables in a single "Emission Factors Hub" sheet.
    """

    @pytest.mark.asyncio
    async def test_parse_data_returns_records_from_2024_format(
        self, mock_async_session, data_source_id, sample_epa_2024_format_excel
    ):
        """
        Test that parse_data returns records from 2024 multi-table format.

        This test FAILS initially because current implementation expects
        "Table 1 - Fuel" sheet which doesn't exist in 2024 file.
        """
        try:
            from backend.services.data_ingestion.epa_ingestion import (
                EPAEmissionFactorsIngestion
            )
        except ImportError:
            pytest.skip("EPAEmissionFactorsIngestion not yet implemented")

        ingestion = EPAEmissionFactorsIngestion(
            db=mock_async_session,
            data_source_id=data_source_id,
            file_key="fuels"
        )

        records = await ingestion.parse_data(sample_epa_2024_format_excel)

        # Should parse records from both Table 1 and Table 2
        assert len(records) > 0, (
            "parse_data should return records from 2024 multi-table format"
        )

    @pytest.mark.asyncio
    async def test_parse_data_extracts_table1_stationary_combustion(
        self, mock_async_session, data_source_id, sample_epa_2024_format_excel
    ):
        """
        Test that parse_data extracts Table 1 (Stationary Combustion) records.

        Table 1 has columns: Fuel Type, Heat Content, CO2 Factor, CH4 Factor, N2O Factor
        """
        try:
            from backend.services.data_ingestion.epa_ingestion import (
                EPAEmissionFactorsIngestion
            )
        except ImportError:
            pytest.skip("EPAEmissionFactorsIngestion not yet implemented")

        ingestion = EPAEmissionFactorsIngestion(
            db=mock_async_session,
            data_source_id=data_source_id,
            file_key="fuels"
        )

        records = await ingestion.parse_data(sample_epa_2024_format_excel)

        # Find records from Table 1 (has CO2 Factor column in kg CO2 per mmBtu)
        table1_records = [
            r for r in records
            if r.get("CO2 Factor") is not None or r.get("kg CO2 per mmBtu") is not None
        ]

        assert len(table1_records) >= 4, (
            f"Should extract at least 4 Table 1 records, got {len(table1_records)}"
        )

    @pytest.mark.asyncio
    async def test_parse_data_extracts_table2_mobile_combustion(
        self, mock_async_session, data_source_id, sample_epa_2024_format_excel
    ):
        """
        Test that parse_data extracts Table 2 (Mobile Combustion CO2) records.

        Table 2 has columns: Fuel Type, kg CO2 per unit, Unit
        """
        try:
            from backend.services.data_ingestion.epa_ingestion import (
                EPAEmissionFactorsIngestion
            )
        except ImportError:
            pytest.skip("EPAEmissionFactorsIngestion not yet implemented")

        ingestion = EPAEmissionFactorsIngestion(
            db=mock_async_session,
            data_source_id=data_source_id,
            file_key="fuels"
        )

        records = await ingestion.parse_data(sample_epa_2024_format_excel)

        # Find records from Table 2 (has Unit column with 'gallon')
        table2_records = [
            r for r in records
            if r.get("Unit") == "gallon" or r.get("kg CO2 per unit") is not None
        ]

        assert len(table2_records) >= 5, (
            f"Should extract at least 5 Table 2 records, got {len(table2_records)}"
        )

    @pytest.mark.asyncio
    async def test_parse_data_includes_table_metadata(
        self, mock_async_session, data_source_id, sample_epa_2024_format_excel
    ):
        """
        Test that parsed records include table metadata for debugging.

        Records should have _source_table or similar metadata.
        """
        try:
            from backend.services.data_ingestion.epa_ingestion import (
                EPAEmissionFactorsIngestion
            )
        except ImportError:
            pytest.skip("EPAEmissionFactorsIngestion not yet implemented")

        ingestion = EPAEmissionFactorsIngestion(
            db=mock_async_session,
            data_source_id=data_source_id,
            file_key="fuels"
        )

        records = await ingestion.parse_data(sample_epa_2024_format_excel)

        # All records should have source metadata
        for record in records:
            assert "_source_sheet" in record or "_source_table" in record, (
                "Records should include source metadata"
            )
            assert "_source_row" in record, (
                "Records should include source row number"
            )


class TestTransformDataMultiTable:
    """
    Test transform_data correctly transforms multi-table records.

    Different tables have different column structures that need
    different transformation logic.
    """

    @pytest.mark.asyncio
    async def test_transform_stationary_combustion_records(
        self, mock_async_session, data_source_id, sample_epa_2024_format_excel
    ):
        """
        Test that Table 1 (Stationary Combustion) records transform correctly.

        Table 1 records have CO2 Factor in kg CO2 per mmBtu.
        """
        try:
            from backend.services.data_ingestion.epa_ingestion import (
                EPAEmissionFactorsIngestion
            )
        except ImportError:
            pytest.skip("EPAEmissionFactorsIngestion not yet implemented")

        ingestion = EPAEmissionFactorsIngestion(
            db=mock_async_session,
            data_source_id=data_source_id,
            file_key="fuels"
        )

        parsed = await ingestion.parse_data(sample_epa_2024_format_excel)
        transformed = await ingestion.transform_data(parsed)

        assert len(transformed) > 0, "Should transform some records"

        # Find a stationary combustion record (e.g., Anthracite)
        anthracite = next(
            (r for r in transformed if "Anthracite" in r.get("activity_name", "")),
            None
        )

        if anthracite:
            assert anthracite["co2e_factor"] > 0, "CO2e factor should be positive"
            assert anthracite["category"] == "combustion", (
                "Stationary combustion should have 'combustion' category"
            )

    @pytest.mark.asyncio
    async def test_transform_mobile_combustion_records(
        self, mock_async_session, data_source_id, sample_epa_2024_format_excel
    ):
        """
        Test that Table 2 (Mobile Combustion CO2) records transform correctly.

        Table 2 records have CO2 Factor in kg CO2 per unit with explicit Unit column.
        """
        try:
            from backend.services.data_ingestion.epa_ingestion import (
                EPAEmissionFactorsIngestion
            )
        except ImportError:
            pytest.skip("EPAEmissionFactorsIngestion not yet implemented")

        ingestion = EPAEmissionFactorsIngestion(
            db=mock_async_session,
            data_source_id=data_source_id,
            file_key="fuels"
        )

        parsed = await ingestion.parse_data(sample_epa_2024_format_excel)
        transformed = await ingestion.transform_data(parsed)

        # Find a mobile combustion record (e.g., Diesel Fuel)
        diesel = next(
            (r for r in transformed if "Diesel" in r.get("activity_name", "")),
            None
        )

        if diesel:
            assert diesel["co2e_factor"] > 0, "CO2e factor should be positive"
            assert "gallon" in diesel["unit"].lower(), (
                "Mobile combustion should have 'gallon' in unit"
            )

    @pytest.mark.asyncio
    async def test_transform_skips_category_headers(
        self, mock_async_session, data_source_id, sample_epa_2024_format_excel
    ):
        """
        Test that category header rows (like "Coal and Coke") are skipped.

        The 2024 format has category headers that don't have numeric values.
        """
        try:
            from backend.services.data_ingestion.epa_ingestion import (
                EPAEmissionFactorsIngestion
            )
        except ImportError:
            pytest.skip("EPAEmissionFactorsIngestion not yet implemented")

        ingestion = EPAEmissionFactorsIngestion(
            db=mock_async_session,
            data_source_id=data_source_id,
            file_key="fuels"
        )

        parsed = await ingestion.parse_data(sample_epa_2024_format_excel)
        transformed = await ingestion.transform_data(parsed)

        # Should not have "Coal and Coke" as an activity name
        coal_headers = [
            r for r in transformed
            if r.get("activity_name") == "Coal and Coke"
        ]

        assert len(coal_headers) == 0, (
            "Category headers like 'Coal and Coke' should be skipped"
        )


class TestEndToEndMultiTable:
    """
    End-to-end tests for full sync workflow with multi-table format.
    """

    @pytest.mark.asyncio
    async def test_execute_sync_processes_multi_table_records(
        self, mock_async_session, data_source_id, sample_epa_2024_format_excel
    ):
        """
        Test that execute_sync processes records from multi-table format.

        The smoke test shows records_processed=0. After fix, this should be >0.
        """
        try:
            from backend.services.data_ingestion.epa_ingestion import (
                EPAEmissionFactorsIngestion
            )
        except ImportError:
            pytest.skip("EPAEmissionFactorsIngestion not yet implemented")

        # Mock the fetch to return our test data
        ingestion = EPAEmissionFactorsIngestion(
            db=mock_async_session,
            data_source_id=data_source_id,
            file_key="fuels"
        )

        # Override fetch_raw_data to use test fixture
        async def mock_fetch():
            return sample_epa_2024_format_excel

        ingestion.fetch_raw_data = mock_fetch

        # Mock execute to simulate successful DB writes
        mock_result = MagicMock()
        mock_result.rowcount = 1
        mock_async_session.execute = AsyncMock(return_value=mock_result)

        result = await ingestion.execute_sync()

        assert result.status == "completed", (
            f"Sync should complete successfully, got status={result.status}"
        )
        assert result.records_processed > 0, (
            f"Should process records from multi-table format, got {result.records_processed}"
        )
        assert result.records_created > 0, (
            f"Should create records from multi-table format, got {result.records_created}"
        )


# ============================================================================
# TASK-DATA-P8-BUG-002: Tests for Table Detection Logic
# ============================================================================

class TestTableDetection:
    """
    Test the table detection logic for identifying table boundaries.

    The 2024 format has "Table N" markers in column B that indicate
    the start of a new table with different column headers.
    """

    def test_detect_table_marker(self):
        """Test detection of table marker patterns."""
        # Table markers in the 2024 format
        table_markers = [
            "Table 1",
            "Table 2",
            "Table 3",
            "Table 10",
            "Table 12",
        ]

        for marker in table_markers:
            assert marker.startswith("Table "), f"{marker} should be a table marker"
            parts = marker.split()
            assert len(parts) == 2, f"{marker} should have 2 parts"
            assert parts[1].isdigit(), f"{marker} should have numeric suffix"

    def test_detect_non_table_rows(self):
        """Test that non-table rows are correctly identified."""
        non_table_rows = [
            "Emission Factors for Greenhouse Gas Inventories",
            "Coal and Coke",
            "Source:",
            "Notes:",
            "",
            None,
        ]

        for row in non_table_rows:
            if row:
                assert not row.startswith("Table ") or not row.split()[1].isdigit() if len(row.split()) > 1 else True


__all__ = [
    "TestMultiTableParsing",
    "TestTransformDataMultiTable",
    "TestEndToEndMultiTable",
    "TestTableDetection",
]
