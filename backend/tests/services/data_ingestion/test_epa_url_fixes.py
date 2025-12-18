"""
Integration tests for EPA connector URL and sheet name fixes.

TASK-DATA-P7-007: Fix EPA Connector URLs and Sheet Names

This test suite validates:
- EPA Fuels URL is accessible (BUG-DATA-001 fix)
- EPA eGRID URL is accessible
- eGRID file contains expected sheets (BUG-DATA-003 fix)
- FILES constant has correct URLs and sheet names

Test-Driven Development Protocol:
- These tests are written BEFORE implementation fixes
- Tests verify the CORRECT values that should be in the code
- Implementation must make tests PASS without modifying tests
"""

import pytest
from unittest.mock import AsyncMock, MagicMock
from uuid import uuid4

from sqlalchemy.ext.asyncio import AsyncSession


@pytest.fixture(scope="function")
def mock_async_session():
    """Create mock async session for unit tests."""
    session = AsyncMock(spec=AsyncSession)
    session.add = MagicMock()
    session.flush = AsyncMock()
    session.commit = AsyncMock()
    session.rollback = AsyncMock()
    session.execute = AsyncMock()
    session.refresh = AsyncMock()
    return session


@pytest.fixture
def data_source_id():
    """Generate a test data source ID."""
    return uuid4().hex


# ============================================================================
# TASK-DATA-P7-007: Integration Tests for URL Accessibility
# ============================================================================

class TestEPAURLAccessibility:
    """
    Integration tests for EPA data source URL accessibility.

    TASK-DATA-P7-007: Fix EPA Connector URLs and Sheet Names

    These tests verify that the EPA URLs are accessible and return valid
    Excel files with the expected sheet names.
    """

    @pytest.mark.integration
    @pytest.mark.asyncio
    async def test_epa_fuels_url_accessible(self):
        """
        Verify EPA Fuels URL returns 200 OK.

        BUG-DATA-001 (CRITICAL): EPA Fuels URL 404
        Expected URL: https://www.epa.gov/system/files/documents/2024-02/ghg-emission-factors-hub-2024.xlsx
        """
        import httpx

        # The correct URL for EPA GHG Emission Factors Hub 2024
        expected_url = "https://www.epa.gov/system/files/documents/2024-02/ghg-emission-factors-hub-2024.xlsx"

        async with httpx.AsyncClient(timeout=60.0, follow_redirects=True) as client:
            response = await client.head(expected_url)
            assert response.status_code == 200, (
                f"EPA Fuels URL returned {response.status_code}. "
                f"URL: {expected_url}"
            )

    @pytest.mark.integration
    @pytest.mark.asyncio
    async def test_epa_egrid_url_accessible(self):
        """
        Verify EPA eGRID URL returns 200 OK.

        eGRID URL should be accessible for electricity emission factors.
        """
        import httpx

        # The eGRID data URL
        expected_url = "https://www.epa.gov/system/files/documents/2024-01/egrid2022_data.xlsx"

        async with httpx.AsyncClient(timeout=60.0, follow_redirects=True) as client:
            response = await client.head(expected_url)
            assert response.status_code == 200, (
                f"EPA eGRID URL returned {response.status_code}. "
                f"URL: {expected_url}"
            )

    @pytest.mark.integration
    @pytest.mark.asyncio
    async def test_epa_egrid_contains_expected_sheets(self):
        """
        Verify eGRID file contains SRL22 and US22 sheets.

        BUG-DATA-003 (HIGH): eGRID Sheet Name Mismatch
        Expected sheets: ["SRL22", "US22"] (NOT SUBRGN22)

        The eGRID file uses SRL22 (Subregion Level) instead of SUBRGN22.
        """
        import httpx
        from openpyxl import load_workbook
        import io

        url = "https://www.epa.gov/system/files/documents/2024-01/egrid2022_data.xlsx"

        async with httpx.AsyncClient(timeout=120.0, follow_redirects=True) as client:
            response = await client.get(url)
            response.raise_for_status()

        workbook = load_workbook(io.BytesIO(response.content), read_only=True)
        sheet_names = workbook.sheetnames
        workbook.close()

        # Verify expected sheets exist
        assert "SRL22" in sheet_names, (
            f"SRL22 sheet not found in eGRID file. "
            f"Available sheets: {sheet_names}"
        )
        assert "US22" in sheet_names, (
            f"US22 sheet not found in eGRID file. "
            f"Available sheets: {sheet_names}"
        )

    @pytest.mark.integration
    @pytest.mark.asyncio
    async def test_epa_fuels_file_is_valid_excel(self):
        """
        Verify EPA Fuels URL returns a valid Excel file.

        This test downloads the file and verifies it can be parsed by openpyxl.
        """
        import httpx
        from openpyxl import load_workbook
        import io

        url = "https://www.epa.gov/system/files/documents/2024-02/ghg-emission-factors-hub-2024.xlsx"

        async with httpx.AsyncClient(timeout=120.0, follow_redirects=True) as client:
            response = await client.get(url)
            response.raise_for_status()

        # Verify the file can be loaded as Excel
        workbook = load_workbook(io.BytesIO(response.content), read_only=True)
        sheet_names = workbook.sheetnames
        workbook.close()

        # Should have at least one sheet
        assert len(sheet_names) > 0, "EPA Fuels file has no sheets"


# ============================================================================
# TASK-DATA-P7-007: Tests for FILES Constant Correctness
# ============================================================================

class TestEPAFilesConstantCorrectness:
    """
    Tests to verify the FILES constant in EPAEmissionFactorsIngestion
    contains the correct URLs and sheet names.

    TASK-DATA-P7-007: Fix EPA Connector URLs and Sheet Names
    """

    def test_files_constant_fuels_url_is_correct(self, mock_async_session, data_source_id):
        """
        Test that FILES constant has the correct fuels URL.

        BUG-DATA-001 fix: URL should point to ghg-emission-factors-hub-2024.xlsx
        """
        try:
            from backend.services.data_ingestion.epa_ingestion import (
                EPAEmissionFactorsIngestion
            )
        except ImportError:
            pytest.skip("EPAEmissionFactorsIngestion not yet implemented")

        files = EPAEmissionFactorsIngestion.FILES

        # The correct URL for EPA Fuels
        expected_url = "https://www.epa.gov/system/files/documents/2024-02/ghg-emission-factors-hub-2024.xlsx"

        assert files["fuels"]["url"] == expected_url, (
            f"Fuels URL is incorrect. "
            f"Expected: {expected_url}, "
            f"Got: {files['fuels']['url']}"
        )

    def test_files_constant_egrid_sheets_are_correct(self, mock_async_session, data_source_id):
        """
        Test that FILES constant has the correct eGRID sheet names.

        BUG-DATA-003 fix: Sheets should be ["SRL22", "US22"] not ["SUBRGN22", "US22"]
        """
        try:
            from backend.services.data_ingestion.epa_ingestion import (
                EPAEmissionFactorsIngestion
            )
        except ImportError:
            pytest.skip("EPAEmissionFactorsIngestion not yet implemented")

        files = EPAEmissionFactorsIngestion.FILES

        # The correct sheet names for eGRID
        expected_sheets = ["SRL22", "US22"]

        assert files["egrid"]["sheets"] == expected_sheets, (
            f"eGRID sheets are incorrect. "
            f"Expected: {expected_sheets}, "
            f"Got: {files['egrid']['sheets']}"
        )

    def test_files_constant_egrid_url_is_correct(self, mock_async_session, data_source_id):
        """
        Test that FILES constant has the correct eGRID URL.
        """
        try:
            from backend.services.data_ingestion.epa_ingestion import (
                EPAEmissionFactorsIngestion
            )
        except ImportError:
            pytest.skip("EPAEmissionFactorsIngestion not yet implemented")

        files = EPAEmissionFactorsIngestion.FILES

        # The correct URL for eGRID
        expected_url = "https://www.epa.gov/system/files/documents/2024-01/egrid2022_data.xlsx"

        assert files["egrid"]["url"] == expected_url, (
            f"eGRID URL is incorrect. "
            f"Expected: {expected_url}, "
            f"Got: {files['egrid']['url']}"
        )

    def test_ingestion_uses_correct_fuels_url_pattern(self, mock_async_session, data_source_id):
        """
        Test that ingestion instance uses URL with 'ghg-emission-factors-hub'.

        After BUG-DATA-001 fix, URL should contain 'ghg-emission-factors-hub'
        instead of 'emission-factors_apr2024'.
        """
        try:
            from backend.services.data_ingestion.epa_ingestion import (
                EPAEmissionFactorsIngestion
            )
        except ImportError:
            pytest.skip("EPAEmissionFactorsIngestion not yet implemented")

        ingestion = EPAEmissionFactorsIngestion(
            db=mock_async_session,
            data_source_id=data_source_id,
            file_key="fuels"
        )

        # Verify the URL pattern is correct
        assert "ghg-emission-factors-hub" in ingestion.file_config["url"], (
            f"Fuels URL should contain 'ghg-emission-factors-hub', "
            f"got: {ingestion.file_config['url']}"
        )

        # Should NOT contain old filename pattern
        assert "emission-factors_apr" not in ingestion.file_config["url"], (
            f"Fuels URL should NOT contain old 'emission-factors_apr' pattern, "
            f"got: {ingestion.file_config['url']}"
        )

    def test_ingestion_uses_correct_egrid_sheet_names(self, mock_async_session, data_source_id):
        """
        Test that eGRID ingestion uses SRL22 sheet instead of SUBRGN22.

        After BUG-DATA-003 fix, sheets should be ["SRL22", "US22"].
        """
        try:
            from backend.services.data_ingestion.epa_ingestion import (
                EPAEmissionFactorsIngestion
            )
        except ImportError:
            pytest.skip("EPAEmissionFactorsIngestion not yet implemented")

        ingestion = EPAEmissionFactorsIngestion(
            db=mock_async_session,
            data_source_id=data_source_id,
            file_key="egrid"
        )

        # Verify SRL22 is in sheets
        assert "SRL22" in ingestion.file_config["sheets"], (
            f"eGRID sheets should contain 'SRL22', "
            f"got: {ingestion.file_config['sheets']}"
        )

        # Should NOT contain old SUBRGN22 sheet name
        assert "SUBRGN22" not in ingestion.file_config["sheets"], (
            f"eGRID sheets should NOT contain old 'SUBRGN22' sheet, "
            f"got: {ingestion.file_config['sheets']}"
        )
